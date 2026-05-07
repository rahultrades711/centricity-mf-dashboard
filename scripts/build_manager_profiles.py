"""
One-time data extraction for manual manager-profile enrichment
(Fund Detail Fix-List 5 §D).

Replaces the previous `scrape_manager_profiles.py` approach. The auto-scrape
was producing thin / inconsistent bios that didn't meet the partner-presentable
quality bar. New approach: emit a CSV with the auto-derivable fields
pre-filled, send to the Products Team for manual enrichment of bios + co-
manager + previously-managed funds, re-load when the human pass is done.

Output:    data/manager-profiles-template.csv
Sort:      Total AUM Under Management descending (highest-AUM managers first
           — easier to prioritise the manual fill)

CSV columns:
  Manager Name
  AMC
  Funds Currently Managed              (pipe-separated fund_names from screener JSON)
  Total AUM Under Management (₹ Cr)    (sum of aum_cr across the manager's funds)
  Tenure in Fund (yrs)                 (per-fund tenure, pipe-separated to match
                                        the funds column 1:1)
  Co-Manager                           [FILL MANUALLY]
  Previously Managed Funds             [FILL MANUALLY]
  Brief Bio                            [FILL MANUALLY]
  Source URL                           [FILL MANUALLY]
  Last Verified Date                   [FILL MANUALLY]

Usage:
    python scripts/build_manager_profiles.py
       [--cycle-json data/screener-YYYY-MM-DD.json]
       [--monitor-xlsx <path>]

If --cycle-json is omitted, the latest data/screener-*.json is used.
The --monitor-xlsx arg is reserved for the future Co-Manager column —
right now Monitor doesn't carry a [Fund Manager 2] column on every
sheet, so co-manager stays a manual fill.
"""
from __future__ import annotations

import csv
import datetime as _dt
import json
import sys
from collections import defaultdict
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
OUTPUT_PATH = DATA_DIR / "manager-profiles-template.csv"


def _latest_screener_json() -> Path | None:
    candidates = sorted(DATA_DIR.glob("screener-*.json"), reverse=True)
    return candidates[0] if candidates else None


def _detect_monitor_co_manager_column(monitor_xlsx: Path | None) -> bool:
    """
    Probe the Monitor workbook to see if any sheet has a '[Fund Manager 2]'
    or 'Co-Manager' header. Right now (15 Apr 2026 file) only Fund Manager 1
    is present. If a future Monitor adds the column, the user can wire it
    up here without changing the CSV columns.
    """
    if monitor_xlsx is None or not monitor_xlsx.exists():
        return False
    try:
        import openpyxl
        wb = openpyxl.load_workbook(monitor_xlsx, read_only=True, data_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=12, column=c).value
                if v and ("Fund Manager 2" in str(v) or "Co-Manager" in str(v) or "Co Manager" in str(v)):
                    return True
        wb.close()
    except Exception:
        pass
    return False


def main(argv: list[str] | None = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    cycle_json: Path | None = None
    monitor_xlsx: Path | None = None
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--cycle-json" and i + 1 < len(argv):
            cycle_json = Path(argv[i + 1]); i += 2; continue
        if a == "--monitor-xlsx" and i + 1 < len(argv):
            monitor_xlsx = Path(argv[i + 1]); i += 2; continue
        i += 1

    if cycle_json is None:
        cycle_json = _latest_screener_json()
    if cycle_json is None or not cycle_json.exists():
        print(
            "[mgr-csv] No screener-*.json in data/. Run "
            "scripts/excel_to_json_screener.py first.",
            file=sys.stderr,
        )
        return 1

    has_co_manager_column = _detect_monitor_co_manager_column(monitor_xlsx)
    if has_co_manager_column:
        print("[mgr-csv] Monitor file carries a co-manager column — use it.", file=sys.stderr)
    else:
        print(
            "[mgr-csv] Monitor file has no co-manager column — Co-Manager "
            "stays a manual fill.",
            file=sys.stderr,
        )

    with open(cycle_json, "r", encoding="utf-8") as f:
        cycle = json.load(f)

    # Group fund records by manager_name
    by_mgr: dict[str, list[dict]] = defaultdict(list)
    for fund in cycle.get("funds", []):
        name = (fund.get("manager_name") or "").strip()
        if not name:
            continue
        by_mgr[name].append(fund)

    rows: list[dict] = []
    for manager, funds in by_mgr.items():
        funds_sorted = sorted(funds, key=lambda x: -(x.get("aum_cr") or 0))
        amcs = sorted({(f.get("amc") or "—") for f in funds_sorted})
        amc_str = " / ".join(amcs)
        names = "|".join(f.get("fund_name") or "—" for f in funds_sorted)
        total_aum = round(sum((f.get("aum_cr") or 0) for f in funds_sorted), 2)
        tenures = "|".join(
            (f"{f.get('manager_tenure_yrs'):.1f}" if f.get("manager_tenure_yrs") is not None else "—")
            for f in funds_sorted
        )
        rows.append({
            "Manager Name": manager,
            "AMC": amc_str,
            "Funds Currently Managed": names,
            "Total AUM Under Management (₹ Cr)": total_aum,
            "Tenure in Fund (yrs)": tenures,
            "Co-Manager": "",
            "Previously Managed Funds": "",
            "Brief Bio": "",
            "Source URL": "",
            "Last Verified Date": "",
        })

    rows.sort(key=lambda r: -r["Total AUM Under Management (₹ Cr)"])

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    columns = [
        "Manager Name",
        "AMC",
        "Funds Currently Managed",
        "Total AUM Under Management (₹ Cr)",
        "Tenure in Fund (yrs)",
        "Co-Manager",
        "Previously Managed Funds",
        "Brief Bio",
        "Source URL",
        "Last Verified Date",
    ]
    with open(OUTPUT_PATH, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

    fund_count = sum(len(v) for v in by_mgr.values())
    print(
        f"[mgr-csv] {len(rows)} unique managers found across {fund_count} funds. "
        f"CSV saved to {OUTPUT_PATH}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
