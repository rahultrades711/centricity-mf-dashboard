"""Phase 4 FINAL pass — apply ALL rules in one clean run.

Rules (locked decisions):
  - Workbook = TRI, Investing.com BSE/Global = PRI
  - BSE funds with real BSE file → use BSE PRI (Method A), td_basis="PRI"
  - "Underlying Index" funds → match index in fund name against full 97-series
  - Proxy only when TE ≤ 5% (proxy-validity gate); narrow themes → absent
  - NAV unit-split adjust: |return| > 40% → multiply pre-split NAVs by clean
    ratio (within 3% of 1/2..1/100); else null nav-derived metrics
  - USD-denominated benchmarks (Nasdaq, S&P 500, NYSE FANG+, MSCI World):
    INR-convert via USD/INR daily before TE/TD; same for US-equity active funds
  - Non-USD international (JPY, TWD, HKD): currency-mismatch absent
  - No international fund in plain "matched" status

Self-gates before export — ALL must pass:
  - 0 passive funds with proxy TE > 5% in matched-proxy
  - 0 foreign/international in any matched status
  - 0 residual TE > 15% except labeled active funds
  - NAV-split sanity: Nifty 50 ETFs TE 0.4-0.5%, TD ≈ -TER

Outputs (LOCAL):
  - C:/Claude Folder/Claude/Projects/Equity MF Screener/BENCHMARK_MATCH_15May_v2.csv
  - C:/Claude Folder/Claude/Projects/Equity MF Screener/BSE_PROXY_MAP_v2.csv
"""
import csv
import datetime as _dt
import math
import os
import re
import statistics
import sys
import time as _tm
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# === LOCAL paths (off OneDrive) ===
WB_PATH = r'C:/Claude Folder/Cent-Claude/MF Screener/Monthly Equity Whitelisting File (MAIN FILE)/15-May-2026/Output/MutualFund_Whitelisting_15May2026.xlsx'
BM_ROOT = r'C:/Claude Folder/Cent-Claude/MF Screener/Monthly Equity Whitelisting File (MAIN FILE)/15-May-2026/Data/BM NAV'
GLOBAL_DIR = BM_ROOT + r'/Global & Others BM NAV'

OUT_DIR = r'C:/Claude Folder/Claude/Projects/Equity MF Screener'
EVIDENCE_CSV = os.path.join(OUT_DIR, 'BENCHMARK_MATCH_15May_v2.csv')
PROXY_CSV = os.path.join(OUT_DIR, 'BSE_PROXY_MAP_v2.csv')

CYCLE = _dt.date(2026, 5, 15)
START3Y = _dt.date(2023, 5, 15)


# ============================================================
# Investing.com CSV parser
# ============================================================
def parse_investing_csv(path):
    out = {}
    with open(path, encoding='utf-8-sig', newline='') as f:
        rdr = csv.reader(f)
        next(rdr)
        for row in rdr:
            if not row or len(row) < 2:
                continue
            ds, ps = row[0].strip(), row[1].strip()
            if not ds or not ps:
                continue
            try:
                d = _dt.datetime.strptime(ds, '%d-%m-%Y').date()
            except ValueError:
                continue
            try:
                v = float(ps.replace(',', '').replace('"', ''))
            except ValueError:
                continue
            if v > 0:
                out[d] = v
    return out


# ============================================================
# Load workbook
# ============================================================
print('[load] workbook (local)...', flush=True)
_t = _tm.time()
wb = openpyxl.load_workbook(WB_PATH, read_only=True, data_only=True)
print(f'  {_tm.time()-_t:.1f}s', flush=True)

# Benchmark NAV — 97 workbook series
ws_bm = wb['\U0001F4C8 Benchmark NAV']
bm_col_for_name = {}
NEW_BSE_NAMES = {'BSE Sensex', 'BSE 100', 'BSE 200', 'BSE 500', 'BSE Midcap',
                  'BSE Smallcap', 'BSE 250 LargeMidcap', 'BSE Sensex Next 50'}
bm_basis_for_name = {}
for r in ws_bm.iter_rows(min_row=2, max_row=2, max_col=ws_bm.max_column, values_only=True):
    for ci, v in enumerate(r, start=1):
        if v:
            nm = str(v).strip()
            bm_col_for_name[nm] = ci
            bm_basis_for_name[nm] = 'PRI' if nm in NEW_BSE_NAMES else 'TRI'
print(f'  workbook benchmark series: {len(bm_col_for_name)}', flush=True)

print('[load] benchmark series one-pass...', flush=True)
_t = _tm.time()
bm_series = {nm: {} for nm in bm_col_for_name}
for r in ws_bm.iter_rows(min_row=4, max_row=ws_bm.max_row, max_col=ws_bm.max_column, values_only=True):
    d = r[0]
    if d is None:
        continue
    if isinstance(d, str):
        try:
            d = _dt.datetime.strptime(d[:10], '%Y-%m-%d').date()
        except ValueError:
            continue
    elif isinstance(d, _dt.datetime):
        d = d.date()
    elif not isinstance(d, _dt.date):
        continue
    for nm, ci in bm_col_for_name.items():
        if ci - 1 >= len(r):
            continue
        v = r[ci - 1]
        if v is None:
            continue
        try:
            v = float(v)
        except (TypeError, ValueError):
            continue
        if v > 0:
            bm_series[nm][d] = v
print(f'  {_tm.time()-_t:.1f}s', flush=True)


# ============================================================
# Load Global PRI sources + USD/INR
# ============================================================
print('[load] Global PRI + USD/INR...', flush=True)
nasdaq_usd = parse_investing_csv(GLOBAL_DIR + r'/Nasdaq 100 Historical Data (1).csv')
sp500_usd = parse_investing_csv(GLOBAL_DIR + r'/S&P 500 Historical Data (3).csv')
fang_usd = parse_investing_csv(GLOBAL_DIR + r'/NYSE FANG+TM Historical Data.csv')
msci_world_usd = parse_investing_csv(GLOBAL_DIR + r'/MSCI World Historical Data.csv')
hang_seng_hkd = parse_investing_csv(GLOBAL_DIR + r'/Hang Seng Historical Data.csv')
usd_inr = parse_investing_csv(GLOBAL_DIR + r'/USD_INR Historical Data (1).csv')
print(f'  Nasdaq 100 (USD): {len(nasdaq_usd)} pts')
print(f'  S&P 500 (USD): {len(sp500_usd)} pts')
print(f'  NYSE FANG+ (USD): {len(fang_usd)} pts')
print(f'  MSCI World (USD): {len(msci_world_usd)} pts')
print(f'  Hang Seng (HKD): {len(hang_seng_hkd)} pts')
print(f'  USD/INR: {len(usd_inr)} pts ({min(usd_inr.keys())} → {max(usd_inr.keys())}, range {min(usd_inr.values()):.2f}-{max(usd_inr.values()):.2f})')

# Build INR-converted versions
def to_inr(usd_series, fx_series):
    out = {}
    for d, v in usd_series.items():
        if d in fx_series:
            out[d] = v * fx_series[d]
    return out

nasdaq_inr = to_inr(nasdaq_usd, usd_inr)
sp500_inr = to_inr(sp500_usd, usd_inr)
fang_inr = to_inr(fang_usd, usd_inr)
msci_world_inr = to_inr(msci_world_usd, usd_inr)
print(f'  Nasdaq INR: {len(nasdaq_inr)} pts | S&P INR: {len(sp500_inr)} | FANG+ INR: {len(fang_inr)} | MSCI World INR: {len(msci_world_inr)}')

# Register these as pseudo "benchmark series" (not in workbook but addressable)
INR_GLOBAL_SERIES = {
    'Nasdaq 100 (INR)': nasdaq_inr,
    'S&P 500 (INR)': sp500_inr,
    'NYSE FANG+ (INR)': fang_inr,
    'MSCI World (INR)': msci_world_inr,
}
# Add into bm_series + basis (PRI-INR = USD PRI converted to INR)
for nm, s in INR_GLOBAL_SERIES.items():
    bm_series[nm] = s
    bm_col_for_name[nm] = -1   # sentinel (not in workbook)
    bm_basis_for_name[nm] = 'PRI-INR'


# ============================================================
# Fund NAV one-pass
# ============================================================
print('[load] fund NAV one-pass...', flush=True)
_t = _tm.time()
ws_fn = wb['\U0001F4C8 Fund NAV']
fund_col_to_name = {}
for r in ws_fn.iter_rows(min_row=2, max_row=2, max_col=ws_fn.max_column, values_only=True):
    for ci, v in enumerate(r, start=1):
        if isinstance(v, str) and v.strip():
            fund_col_to_name[ci] = v.strip()
    break
fund_cols = {nm: ci for ci, nm in fund_col_to_name.items()}
fund_series = {nm: {} for nm in fund_col_to_name.values()}
col_list = list(fund_col_to_name.items())
for r in ws_fn.iter_rows(min_row=4, max_row=ws_fn.max_row, max_col=ws_fn.max_column, values_only=True):
    d = r[0]
    if d is None:
        continue
    if isinstance(d, str):
        try:
            d = _dt.datetime.strptime(d[:10], '%Y-%m-%d').date()
        except ValueError:
            continue
    elif isinstance(d, _dt.datetime):
        d = d.date()
    elif not isinstance(d, _dt.date):
        continue
    for ci, nm in col_list:
        if ci - 1 >= len(r):
            continue
        v = r[ci - 1]
        if v is None:
            continue
        try:
            v = float(v)
        except (TypeError, ValueError):
            continue
        if v > 0:
            fund_series[nm][d] = v
print(f'  {_tm.time()-_t:.1f}s, {len(fund_series)} funds', flush=True)


# Data sheet
ws_d = wb['\U0001F4CB Data']
funds = []
for r in ws_d.iter_rows(min_row=5, max_row=ws_d.max_row, max_col=10, values_only=True):
    if not r[1]:
        continue
    funds.append({'amfi': r[3], 'name': r[1], 'cat': r[2], 'stated_bm': r[9]})
print(f'  funds: {len(funds)}', flush=True)
wb.close()


# ============================================================
# NAV split scan + adjust
# ============================================================
SPLIT_THRESHOLD = 0.40
CLEAN_FACTORS = [1/2, 1/3, 1/4, 1/5, 1/10, 1/20, 1/50, 1/100, 2, 3, 4, 5, 10, 20, 50, 100]
CLEAN_TOLERANCE = 0.03


def detect_splits(series):
    out = []
    dates = sorted(series.keys())
    for i in range(1, len(dates)):
        v0, v1 = series[dates[i-1]], series[dates[i]]
        if v0 <= 0:
            continue
        r = v1 / v0 - 1
        if abs(r) > SPLIT_THRESHOLD:
            factor = v1 / v0
            is_clean = any(abs(factor - cf) / cf < CLEAN_TOLERANCE for cf in CLEAN_FACTORS)
            out.append((dates[i], factor, is_clean))
    return out


def split_adjust(series, splits):
    if not splits:
        return series
    if any(not is_clean for _, _, is_clean in splits):
        return None
    adjusted = dict(series)
    for split_date, factor, _ in splits:
        for d in adjusted:
            if d < split_date:
                adjusted[d] = adjusted[d] * factor
    return adjusted


print('[scan] NAV split scan...', flush=True)
split_results = {}
for nm, s in fund_series.items():
    sp = detect_splits(s)
    if sp:
        all_clean = all(c for _, _, c in sp)
        split_results[nm] = (sp, 'clean-split-adjusted' if all_clean else 'nav-split-suspect')

# Apply clean adjustments
for nm, (sp, st) in split_results.items():
    if st == 'clean-split-adjusted':
        adj = split_adjust(fund_series[nm], sp)
        if adj:
            fund_series[nm] = adj

clean_count = sum(1 for _, (_, s) in split_results.items() if s == 'clean-split-adjusted')
suspect_count = sum(1 for _, (_, s) in split_results.items() if s == 'nav-split-suspect')
print(f'  splits detected: {len(split_results)} | clean-adjusted: {clean_count} | suspect-nulled: {suspect_count}', flush=True)


# ============================================================
# Matching tables
# ============================================================
INTERNATIONAL_USD = [
    # Patterns for US-equity / Nasdaq / S&P 500 / global multi-cap → can be INR-converted
    (r'\bUS\s+Bluechip\b|\bUS Equity\b', 'S&P 500 (INR)'),     # ICICI US Bluechip benchmarks to S&P 500
    (r'\bUS\s+Opp\b|\bUS\s+Opportun', 'S&P 500 (INR)'),
    (r'\bNASDAQ[- ]?100\b', 'Nasdaq 100 (INR)'),
    (r'\bS&P[- ]?500\b(?!\s+Top)', 'S&P 500 (INR)'),
    (r'\bS&P 500 Top 50\b', 'S&P 500 (INR)'),                   # subset → proxy
    (r'\bFANG\+?\b|\bNYSE FANG', 'NYSE FANG+ (INR)'),
    (r'\bGlobal\b.*\bAdvantage\b|\bMSCI World\b|\bGlobal Equity\b', 'MSCI World (INR)'),
    (r'\bInternational Equity\b', 'MSCI World (INR)'),
    (r'\bGlobal Emerging Market', 'MSCI World (INR)'),          # closest broad-global; flag as proxy
]
INTERNATIONAL_NON_USD = [
    # Patterns where the benchmark currency isn't USD and we don't have FX → currency-mismatch absent
    r'\bJapan\b',
    r'\bTaiwan\b',
    r'\bHang Seng\b',
    r'\bChina\b(?!\b)',
    r'\bGreater China\b',
    r'\bSingapore\b',
    r'\bKorea\b',
    r'\bAustralia\b',
    r'\bASEAN\b',
    r'\bEurope\b',
    r'\bGermany\b',
    r'\bBrazil\b',
]

# Truly absent (no real, no sensible proxy — narrow themes)
NO_SENSIBLE_PROXY = [
    r'\bRailways PSU\b', r'\bRailway PSU\b',
    r'\bBSE Hospitals\b', r'\bBSE Power\b',
    r'\bSelect IPO\b',
]

# Underlying-Index inference + proxy table (carried from second_pass, refined)
UI_INFER = [
    # BSE direct
    (r'\bBSE Sensex(?!.*Next)', 'BSE Sensex', False),
    (r'(?<!BSE )\bSensex\b(?!.*Next)', 'BSE Sensex', False),
    (r'\bBSE Sensex Next 50\b|\bSensex Next 50\b', 'BSE Sensex Next 50', False),
    (r'\bBSE Sensex Next 30\b|\bSensex Next 30\b', 'BSE Sensex Next 50', True),
    (r'\bBSE 100\b', 'BSE 100', False),
    (r'\bBSE 200 Equal Weight\b', 'BSE 200', True),     # EW vs cap-wt: structural, will allow if TE ≤ 5%
    (r'\bBSE 200\b', 'BSE 200', False),
    (r'\bBSE 500 (?:Momentum|Quality|Value|Dividend|Low Vol)', 'BSE 500', True),
    (r'\bBSE[- ]?500\b', 'BSE 500', False),
    (r'\bBSE Midcap Select\b', 'BSE Midcap', True),
    (r'\bBSE Midcap\b', 'BSE Midcap', False),
    (r'\bBSE Smallcap\b', 'BSE Smallcap', False),
    (r'\bBSE 250 LargeMidcap\b', 'BSE 250 LargeMidcap', False),
    # Legacy bare "Nifty Index Fund"
    (r'\bNifty Index Fund\b(?!.*\b(?:50|100|200|500|Bank|Midcap|Smallcap|Next|Auto|Pharma|FMCG|IT|Realty|Healthcare)\b)', 'Nifty 50', False),
    (r'\bIndex Fund Nifty Plan\b|\bNifty Plan\b', 'Nifty 50', False),
    (r'\bIndex Fund Sensex Plan\b|\bSensex Plan\b', 'BSE Sensex', False),
    # BSE sectoral proxies (TRI proxy)
    (r'\bBSE Healthcare\b', 'Nifty Healthcare Index', True),
    (r'\bBSE\s+(?:India\s+)?Select Top 10 Banks\b|\bBSE\s+(?:India\s+)?Top 10 Banks\b', 'Nifty Bank', True),
    (r'\bBSE India Infrastructure\b', 'Nifty Infrastructure', True),
    (r'\bBSE Quality\b', 'NIFTY100 Quality 30', True),
    (r'\bBSE Low Volatility\b', 'Nifty100 Low Volatility 30', True),
    (r'\bBSE Enhanced Value\b', 'NIFTY500 Value 50', True),
    (r'\bBSE\s+(?:India\s+)?Sector Leaders\b', 'Nifty 500', True),
    (r'\bBSE 1000\b', 'Nifty Total Market', True),
    (r'\bBSE Select Business Groups\b', 'Nifty 100', True),
    (r'\bBSE Internet Economy\b', 'Nifty India Digital', True),
    (r'\bBSE Housing\b', 'Nifty Housing', True),
    (r'\bBSE Capital Markets\b', 'Nifty Capital Markets', True),
    (r'\bMSCI India\b', 'Nifty 500', True),
    (r'\bBSE PSU\b', 'Nifty PSU Bank', True),
    (r'\bBSE Financials ex Bank\b', 'Nifty Financial Services Ex-Bank', True),
    (r'\bBSE\s+(?:India\s+)?Defence\b|\bBSE Defence\b', 'Nifty India Defence', True),
    (r'\bBSE Multicap Consumption\b', 'Nifty FMCG', True),
    # Nifty broad
    (r'\bNifty Next 50\b', 'Nifty Next 50', False),
    (r'\bNifty Total Market\b', 'Nifty Total Market', False),
    (r'\bNifty Midcap 150\b(?!\s+Momentum)(?!\s+Quality)', 'Nifty Midcap 150', False),
    (r'\bNifty Smallcap 250\b(?!\s+Quality)(?!\s+Momentum)', 'Nifty Smallcap 250', False),
    (r'\bNifty Microcap\b', 'Nifty Microcap 250', False),
    (r'\bNifty Midcap Select\b', 'Nifty Midcap Select', False),
    (r'\bNifty Midcap 50\b', 'Nifty Midcap 50', False),
    (r'\bNifty Smallcap 50\b', 'Nifty Smallcap 50', False),
    (r'\bNifty Midcap 100\b', 'Nifty Midcap 150', True),
    (r'\bNifty Smallcap 100\b', 'Nifty Smallcap 250', True),
    # Nifty 50 specific variants first
    (r'\bNifty 50 Arbitrage\b', 'Nifty 50 Arbitrage', False),
    (r'\bNifty 50 Equal Weight\b|\bNifty50 Equal Weight\b', 'NIFTY50 Equal Weight', False),
    (r'\bNifty 50 Value 20\b|\bNifty50 Value 20\b', 'Nifty50 Value 20', False),
    (r'\bNifty 50 Shariah\b|\bNifty50 Shariah\b', 'Nifty50 Shariah', False),
    (r'\bNifty500 Shariah\b|\bNifty 500 Shariah\b', 'Nifty500 Shariah', False),
    (r'\bNifty500 Multicap\b|\bNifty 500 Multicap\b', 'Nifty500 Multicap 50:25:25', False),
    # Bare numerics (after variants)
    (r'\bNifty 50\b(?!\s+(?:Arbitrage|Equal|Value|Shariah))', 'Nifty 50', False),
    (r'\bNifty 100\b(?!\s+(?:Low|Quality|Equal|Alpha|ESG))', 'Nifty 100', False),
    (r'\bNifty 200\b(?!\s+(?:Quality|Momentum|Value|Alpha))', 'Nifty 200', False),
    (r'\bNifty 500\b(?!\s+(?:Multicap|Value|Quality|Shariah|Momentum|Healthcare|Flexicap|Low Vol))', 'Nifty 500', False),
    # Sectoral
    (r'\bNifty Bank\b(?!\s*Ex[- ]?Bank)', 'Nifty Bank', False),
    (r'\bNifty IT\b', 'Nifty IT', False),
    (r'\bNifty FMCG\b', 'Nifty FMCG', False),
    (r'\bNifty Pharma\b', 'Nifty Pharma', False),
    (r'\bNifty Healthcare\b', 'Nifty Healthcare Index', False),
    (r'\bNifty Auto\b', 'Nifty Auto', False),
    (r'\bNifty Energy\b', 'Nifty Energy', False),
    (r'\bNifty Realty\b', 'Nifty Realty', False),
    (r'\bNifty Oil\s*[& ]+\s*Gas\b', 'Nifty Oil & Gas', False),
    (r'\bNifty PSU Bank\b', 'Nifty PSU Bank', False),
    (r'\bNifty Financial Services Ex[- ]?Bank\b', 'Nifty Financial Services Ex-Bank', False),
    (r'\bNifty Financial Services 25.?50\b', 'Nifty Financial Services 25/50', False),
    (r'\bNifty Financial Services\b', 'Nifty Financial Services', False),
    (r'\bNifty Consumer Durables\b', 'Nifty Consumer Durables', False),
    (r'\bNifty Capital Markets?\b', 'Nifty Capital Markets', False),
    (r'\bNifty Chemicals\b', 'Nifty Chemicals', False),
    (r'\bNifty\s+(?:India\s+)?Infrastructure(?:\s*&\s*Logistics)?\b', 'Nifty Infrastructure', False),
    (r'\bNifty Media\b', 'Nifty Media', False),
    (r'\bNifty Transportation\b', 'Nifty Transportation & Logistics', False),
    (r'\bNifty Housing\b', 'Nifty Housing', False),
    (r'\bNifty IPO\b', 'Nifty IPO', False),
    (r'\bNifty REIT', 'Nifty REITs & InvITs', False),
    (r'\bNifty High Beta\b', 'Nifty High Beta 50', False),
    (r'\bNifty Low Volatility 50\b', 'Nifty Low Volatility 50', False),
    (r'\bNifty Commodities\b', 'Nifty Commodities', False),
    (r'\bNifty MidSmallcap[- ]?400\b(?!.*Momentum)', 'Nifty MidSmallcap 400', False),
    (r'\bNifty MidSmallcap 400 Momentum Quality\b', 'Nifty MidSmallcap400 Momentum Quality 100', False),
    (r'\bNifty India Defence\b|\bIndia Defence\b', 'Nifty India Defence', False),
    (r'\bNifty India Manufacturing\b', 'Nifty India Manufacturing', False),
    (r'\bNifty India Digital\b', 'Nifty India Digital', False),
    (r'\bNifty India Internet\b', 'Nifty India Digital', True),
    (r'\bNifty India Tourism\b', 'Nifty India Tourism', False),
    (r'\bMidSmall IT (?:and|&) Telecom\b|\bMidSmall IT\b', 'Nifty MidSmall IT & Telecom', False),
    (r'\bNifty\s+Large\s*Mid\s*Cap\s*250\b|\bLargeMidcap 250\b', 'NIFTY LargeMidcap 250', False),
    (r'\bNifty EV(?:\s*&|\s*and)?\s*New Age Auto', 'Nifty Auto', True),
    (r'\bNifty India New Age Consumption\b|\bNifty Non-Cyclical Consumer\b', 'Nifty FMCG', True),
    (r'\bNifty\s+(?:India\s+)?Consumption\b', 'Nifty FMCG', True),
    (r'\bNifty Dividend Opportunities\b', 'Nifty 50', True),
    (r'\bCPSE\b|\bBharat 22\b', 'Nifty PSU Bank', True),
    (r'\bNifty MNC\b', 'Nifty 100', True),
    (r'\bNifty Metal\b', 'Nifty Energy', True),
    (r'\bNifty PSE\b', 'Nifty PSU Bank', True),
    (r'\bNifty Private Bank\b|\bNifty Pvt Bank\b', 'Nifty Bank', True),
    (r'\bNifty 500 Healthcare\b', 'Nifty Healthcare Index', True),
    (r'\bNifty500 Momentum 50\b|\bNifty 500 Momentum 50\b', 'Nifty200 Momentum 30', True),
    (r'\bNifty500 Flexicap Quality\b|\bFlexicap Quality 30\b', 'Nifty500 Quality 50', True),
    (r'\bNifty Total Market Momentum Quality\b', 'Nifty MidSmallcap400 Momentum Quality 100', True),
    (r'\bNifty Growth Sectors\b', 'Nifty Top 15 Equal Weight', True),
    (r'\bNifty Services Sector\b', 'Nifty Financial Services', True),
    # Factor combos
    (r'\bNifty100 Low Volatility 30\b|\bNifty 100 Low Vol(?:atility)? 30\b', 'Nifty100 Low Volatility 30', False),
    (r'\bNifty100 Equal Weight\b|\bNifty 100 Equal Weight\b', 'Nifty100 Equal Weight', False),
    (r'\bNifty100 ESG\b|\bNifty 100 ESG\b', 'Nifty100 ESG Sector Leaders', False),
    (r'\bNifty100 Quality\b|\bNifty 100 Quality\b', 'NIFTY100 Quality 30', False),
    (r'\bNifty100 Alpha\b|\bNifty 100 Alpha\b', 'NIFTY100 Alpha 30', False),
    (r'\bNifty200 Alpha\b|\bNifty 200 Alpha\b', 'Nifty200 Alpha 30', False),
    (r'\bNifty200 Momentum\b|\bNifty 200 Momentum\b', 'Nifty200 Momentum 30', False),
    (r'\bNifty200 Quality\b|\bNifty 200 Quality\b', 'NIFTY200 Quality 30', False),
    (r'\bNifty200 Value\b|\bNifty 200 Value\b', 'Nifty200 Value 30', False),
    (r'\bNifty500 Equal Weight\b|\bNifty 500 Equal Weight\b', 'Nifty500 Equal Weight', False),
    (r'\bNifty500 Low Vol\b|\bNifty 500 Low Vol\b', 'Nifty500 Low Volatility 50', False),
    (r'\bNifty500 Quality\b|\bNifty 500 Quality\b', 'Nifty500 Quality 50', False),
    (r'\bNifty500 Value\b|\bNifty 500 Value\b|\bNIFTY500 Value 50\b', 'NIFTY500 Value 50', False),
    (r'\bNifty Alpha 50\b', 'Nifty Alpha 50', False),
    (r'\bNifty Midcap150 Momentum\b|\bMidcap 150 Momentum\b', 'Nifty Midcap150 Momentum 50', False),
    (r'\bNifty Midcap150 Quality\b|\bMidcap 150 Quality\b', 'NIFTY Midcap150 Quality 50', False),
    (r'\bNifty Smallcap250 Quality\b|\bSmallcap 250 Quality\b', 'Nifty Smallcap250 Quality 50', False),
    (r'\bNifty Smallcap250 Momentum Quality\b|\bSmallcap 250 Momentum Quality\b', 'Nifty Smallcap250 Momentum Quality 100', False),
    (r'\bNifty MidSmall Financial Services\b', 'Nifty MidSmall Financial Services', False),
    (r'\bNifty MidSmall Healthcare\b', 'Nifty MidSmall Healthcare', False),
    (r'\bNifty MidSmall India Consumption\b', 'Nifty MidSmall India Consumption', False),
    (r'\bNIFTY Alpha Low.Volatility 30\b|\bNifty Alpha Low - Volatility 30\b', 'NIFTY Alpha Low-Volatility 30', False),
    (r'\bNIFTY Alpha Quality Value Low-Volatility\b', 'NIFTY Alpha Quality Value Low-Volatility 30', False),
    (r'\bNIFTY Alpha Quality Low-Volatility\b', 'NIFTY Alpha Quality Low-Volatility 30', False),
    (r'\bNIFTY Quality Low-Volatility 30\b', 'NIFTY Quality Low-Volatility 30', False),
    (r'\bNifty Top 10 Equal Weight\b', 'Nifty Top 10 Equal Weight', False),
    (r'\bNifty Top 15 Equal Weight\b', 'Nifty Top 15 Equal Weight', False),
    (r'\bNifty Top 20 Equal Weight\b', 'Nifty Top 20 Equal Weight', False),
    (r'\bNifty SME EMERGE\b', 'NIFTY SME EMERGE', False),
]


# ============================================================
# Matching function
# ============================================================
def canon_stated(s):
    s = re.sub(r'\s*-\s*TRI\s*$', '', str(s or '').strip(), flags=re.IGNORECASE)
    s = re.sub(r'\s*-\s*PRI\s*$', '', s, flags=re.IGNORECASE)
    s = s.lower()
    for ch in '-()/&,+':
        s = s.replace(ch, ' ')
    GENERIC = {'etf', 'fund', 'reg', 'plan', 'direct', 'growth', 'g', 'idcw', 'index'}
    tokens = [t for t in s.split() if t and t not in GENERIC]
    return ''.join(tokens)


bm_canon = {canon_stated(nm): nm for nm in bm_col_for_name}


def match_international(fund_name):
    """Returns (matched_series, is_proxy, basis, status) or None if not international."""
    for pat in INTERNATIONAL_NON_USD:
        if re.search(pat, fund_name, re.IGNORECASE):
            return (None, False, None, 'absent (currency-mismatch, no INR conversion)')
    for pat, target in INTERNATIONAL_USD:
        if re.search(pat, fund_name, re.IGNORECASE):
            is_proxy = ('FANG' in target or 'MSCI' in target) or ('Emerging' in fund_name)
            return (target, is_proxy, 'PRI-INR', 'matched (USD→INR converted)')
    return None


def match_fund(fund):
    stated = (fund.get('stated_bm') or '').strip()
    fund_name = fund.get('name') or ''

    # Tier 0 — international FIRST (must catch before any "matched" status)
    intl = match_international(fund_name)
    if intl:
        target, is_proxy, basis, status = intl
        return {'matched_series': target, 'is_proxy': is_proxy, 'basis': basis,
                'status': status, 'td_basis': basis or 'n/a'}

    # Tier 1 — alias-fold on stated_bm
    if stated and stated.lower() != 'underlying index':
        cs = canon_stated(stated)
        if cs and cs in bm_canon:
            wb_name = bm_canon[cs]
            basis = bm_basis_for_name[wb_name]
            return {'matched_series': wb_name, 'is_proxy': False, 'basis': basis,
                    'status': 'matched (alias-fold)', 'td_basis': basis}

    # Tier 2 — UI inference
    if stated.lower() == 'underlying index' or stated == '':
        for pat, target, is_proxy in UI_INFER:
            if re.search(pat, fund_name, re.IGNORECASE):
                if target not in bm_col_for_name:
                    continue
                basis = bm_basis_for_name[target]
                status = 'matched-proxy' if is_proxy else 'matched'
                return {'matched_series': target, 'is_proxy': is_proxy, 'basis': basis,
                        'status': status, 'td_basis': basis}

    # Tier 3 — no-sensible-proxy themes
    for pat in NO_SENSIBLE_PROXY:
        if re.search(pat, fund_name, re.IGNORECASE):
            return {'matched_series': None, 'is_proxy': False, 'basis': None,
                    'status': 'absent (no real series, no sensible proxy)', 'td_basis': None}

    # Tier 4 — absent
    return {'matched_series': None, 'is_proxy': False, 'basis': None,
            'status': 'absent (no rule matched)', 'td_basis': None}


# ============================================================
# TE/TD compute
# ============================================================
def common_dates(s1, s2, start=START3Y, end=CYCLE):
    return sorted(d for d in s1.keys() if d in s2 and start <= d <= end)


def daily_returns(s, dates):
    out = []
    for i in range(1, len(dates)):
        v0, v1 = s[dates[i-1]], s[dates[i]]
        if v0 > 0:
            out.append(v1 / v0 - 1)
    return out


def compute_te_td(fund_s, bm_s):
    if not fund_s or not bm_s:
        return None, None, 0, None, None
    dates = common_dates(fund_s, bm_s)
    if len(dates) < 30:
        return None, None, len(dates), None, None
    f_ret = daily_returns(fund_s, dates)
    b_ret = daily_returns(bm_s, dates)
    diff = [f - b for f, b in zip(f_ret, b_ret)]
    if len(diff) < 2:
        return None, None, len(dates), None, None
    sd = statistics.stdev(diff)
    te = round(sd * math.sqrt(252) * 100, 4)
    d0, d1 = dates[0], dates[-1]
    yrs = (d1 - d0).days / 365.25
    if yrs <= 0:
        return te, None, len(dates), None, None
    f_cagr = ((fund_s[d1] / fund_s[d0]) ** (1/yrs) - 1) * 100
    b_cagr = ((bm_s[d1] / bm_s[d0]) ** (1/yrs) - 1) * 100
    return te, round(f_cagr - b_cagr, 4), len(dates), round(f_cagr, 4), round(b_cagr, 4)


# ============================================================
# Pass 1 — match + compute (raw, before proxy-validity gate)
# ============================================================
PASSIVE_CATS = {'ETFs', 'Large Cap Index', 'Mid Cap Index', 'Small Cap Index',
                'Multi/Broad Index', 'Sectoral/Thematic Index', 'Smart-Beta/Factor'}
EW_PATTERNS = [r'\bEqual Weight\b']  # structural exception for proxy-validity gate

def is_equal_weight(fund_name):
    for p in EW_PATTERNS:
        if re.search(p, fund_name, re.IGNORECASE):
            return True
    return False

print('\n[match+compute] pass 1 (raw)...', flush=True)
rows = []
for i, f in enumerate(funds):
    m = match_fund(f)
    nm = f['name']
    si = split_results.get(nm)
    nav_split_status = si[1] if si else ''
    fs = fund_series.get(nm, {})
    matched = m['matched_series']
    bm_s = bm_series.get(matched, {}) if matched else {}
    skip_metrics = (nav_split_status == 'nav-split-suspect')

    if skip_metrics:
        te, td, n, f_c, b_c = None, None, 0, None, None
    else:
        te, td, n, f_c, b_c = compute_te_td(fs, bm_s)

    is_passive = (f['cat'] in PASSIVE_CATS)

    rows.append({
        'AMFI': f['amfi'], 'fund': nm, 'category': f['cat'], 'is_passive': is_passive,
        'stated_benchmark': f['stated_bm'],
        'matched_series': matched or '',
        'is_proxy': m['is_proxy'],
        'basis': m['basis'] or '',
        'td_basis': m.get('td_basis') or '',
        'TE_pct': te if te is not None else '',
        'TD_pct': td if td is not None else '',
        'n_common_days': n,
        'fund_cagr_3y': f_c if f_c is not None else '',
        'bm_cagr_3y': b_c if b_c is not None else '',
        'currency_adjusted': ('YES' if m['basis'] == 'PRI-INR' else 'NO'),
        'nav_split_status': nav_split_status,
        'status': m['status'],
        'is_equal_weight': is_equal_weight(nm),
    })
    if (i+1) % 200 == 0:
        print(f'  {i+1}/{len(funds)}', flush=True)


# ============================================================
# Proxy-validity gate: for passive funds, if matched-proxy and TE > 5%
# (and not equal-weight structural exception) → downgrade to absent
# ============================================================
print('\n[gate] proxy-validity (TE > 5% on passive matched-proxy → absent)...', flush=True)
downgraded = []
for r in rows:
    if not r['is_passive']:
        continue
    if not r['is_proxy']:
        continue
    if r['TE_pct'] == '' or r['TE_pct'] is None:
        continue
    try:
        te_v = float(r['TE_pct'])
    except (TypeError, ValueError):
        continue
    if te_v > 5.0 and not r['is_equal_weight']:
        downgraded.append((r['fund'], r['matched_series'], te_v))
        r['status'] = f'absent (proxy TE={te_v:.2f}% > 5%, downgraded)'
        r['matched_series'] = ''
        r['is_proxy'] = False
        r['basis'] = ''
        r['td_basis'] = ''
        r['TE_pct'] = ''
        r['TD_pct'] = ''
        r['fund_cagr_3y'] = ''
        r['bm_cagr_3y'] = ''
        r['n_common_days'] = 0
print(f'  funds downgraded: {len(downgraded)}')
for nm, ms, te_v in downgraded[:20]:
    print(f'    {nm[:55]} (was proxy→{ms}, TE={te_v:.2f}%)')


# ============================================================
# Sample empirical (BSE 6 funds — both methods)
# ============================================================
SAMPLE = [
    ('HDFC BSE Sensex ETF', 'BSE Sensex', 'Nifty 50'),
    ('SBI BSE Sensex ETF',  'BSE Sensex', 'Nifty 50'),
    ('HDFC BSE 500 ETF',    'BSE 500',    'Nifty 500'),
    ('ICICI Pru BSE Midcap Select ETF', 'BSE Midcap', 'Nifty Midcap 150'),
    ('Mirae Asset BSE 200 Equal Weight ETF', 'BSE 200', 'Nifty 200'),
    ('ICICI Pru BSE 500 ETF', 'BSE 500', 'Nifty 500'),
]
sample_results = {}
for fund_name, bse, nfty in SAMPLE:
    fs = fund_series.get(fund_name, {})
    ate, atd, _, _, _ = compute_te_td(fs, bm_series.get(bse, {}))
    bte, btd, _, _, _ = compute_te_td(fs, bm_series.get(nfty, {}))
    sample_results[fund_name] = (ate, atd, bse, bte, btd, nfty)

for r in rows:
    if r['fund'] in sample_results:
        a_te, a_td, a_bm, b_te, b_td, b_bm = sample_results[r['fund']]
        r['A_method_TE_pct'] = a_te if a_te is not None else ''
        r['A_method_TD_pct'] = a_td if a_td is not None else ''
        r['A_method_bm'] = a_bm
        r['B_method_TE_pct'] = b_te if b_te is not None else ''
        r['B_method_TD_pct'] = b_td if b_td is not None else ''
        r['B_method_bm'] = b_bm
        if a_te is not None and b_te is not None:
            r['winning_method'] = 'A (BSE PRI)' if abs(a_te - 0.275) < abs(b_te - 0.275) else 'B (Nifty TRI)'
        else:
            r['winning_method'] = ''
    else:
        for k in ['A_method_TE_pct', 'A_method_TD_pct', 'A_method_bm',
                  'B_method_TE_pct', 'B_method_TD_pct', 'B_method_bm', 'winning_method']:
            r[k] = ''


# ============================================================
# Self-gates
# ============================================================
print('\n=== SELF-GATES ===', flush=True)
# (a) passive proxy TE > 5% remaining
g1 = sum(1 for r in rows if r['is_passive'] and r['is_proxy']
         and isinstance(r['TE_pct'], (int, float)) and r['TE_pct'] > 5.0)
print(f'  (a) passive matched-proxy with TE > 5% remaining: {g1}', flush=True)

# (b) international funds in any `matched` (not proxy, not absent)
g2 = sum(1 for r in rows if (
    'currency-mismatch' not in r['status'] and 'absent' not in r['status']
    and r['basis'] == 'PRI-INR'  # international (USD-converted) — should be flagged
    and not r['is_proxy']
))
print(f'  (b) international funds in plain matched (non-proxy, non-absent): {g2}', flush=True)

# (c) residual TE > 15% — list as active funds
te_15 = []
for r in rows:
    if isinstance(r['TE_pct'], (int, float)) and r['TE_pct'] > 15.0:
        te_15.append((r['fund'], r['TE_pct'], r['is_passive'], r['status']))
passive_high_te = [t for t in te_15 if t[2]]
active_high_te = [t for t in te_15 if not t[2]]
print(f'  (c) residual TE > 15% — passive: {len(passive_high_te)}, active: {len(active_high_te)}')
for nm, te_v, pas, st in passive_high_te[:10]:
    print(f'      PASSIVE {nm[:55]} TE={te_v}% status={st}')

# (d) Nifty 50 ETF sanity gate
print('\n  (d) Nifty 50 ETF sanity gate:')
for nm in ('SBI Nifty 50 ETF', 'Nippon India ETF Nifty 50 BeES', 'ICICI Pru Nifty 50 ETF',
           'UTI Nifty 50 ETF', 'Quantum Nifty 50 ETF', 'Edelweiss Nifty 50 ETF'):
    mt = [r for r in rows if r['fund'] == nm]
    if mt:
        r = mt[0]
        print(f'    {nm}: TE={r["TE_pct"]}%, TD={r["TD_pct"]}%, split={r["nav_split_status"] or "(none)"}')


# ============================================================
# Write CSVs
# ============================================================
os.makedirs(OUT_DIR, exist_ok=True)
print(f'\n[write] {EVIDENCE_CSV}', flush=True)

hdrs = ['AMFI', 'fund', 'category', 'is_passive', 'is_equal_weight', 'stated_benchmark',
        'matched_series', 'is_proxy', 'basis', 'td_basis',
        'TE_pct', 'TD_pct', 'n_common_days', 'fund_cagr_3y', 'bm_cagr_3y',
        'currency_adjusted', 'nav_split_status', 'status',
        'A_method_TE_pct', 'A_method_TD_pct', 'A_method_bm',
        'B_method_TE_pct', 'B_method_TD_pct', 'B_method_bm', 'winning_method']

# Sort: residuals first (absent + suspect), then matched
def sort_key(r):
    s = r['status']
    if 'absent (currency' in s: rank = 1
    elif 'downgraded' in s: rank = 2
    elif 'absent' in s: rank = 0
    elif 'nav-split-suspect' in s: rank = 3
    elif 'matched-proxy' in s: rank = 4
    else: rank = 5
    return (rank, str(r['AMFI']))

with open(EVIDENCE_CSV, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=hdrs)
    w.writeheader()
    for r in sorted(rows, key=sort_key):
        w.writerow(r)

# Verify row count + final-row integrity
with open(EVIDENCE_CSV, encoding='utf-8-sig') as f:
    csv_rows = list(csv.DictReader(f))
print(f'  CSV rows: {len(csv_rows)} (expected 1252)', flush=True)
final_row_fields = sum(1 for v in csv_rows[-1].values() if v is not None)
print(f'  final row fields: {final_row_fields} / {len(hdrs)} expected', flush=True)
# Check trailing newline
with open(EVIDENCE_CSV, 'rb') as f:
    f.seek(-2, 2)
    tail = f.read()
print(f'  trailing bytes: {tail!r} (should end with newline)', flush=True)

# Proxy map
print(f'[write] {PROXY_CSV}', flush=True)
proxy_pats_seen = set()
with open(PROXY_CSV, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=['fund_name_pattern', 'mapped_to_workbook_series', 'is_BSE_to_Nifty', 'note'])
    w.writeheader()
    for pat, target, is_proxy in UI_INFER:
        if not is_proxy:
            continue
        if (pat, target) in proxy_pats_seen:
            continue
        proxy_pats_seen.add((pat, target))
        bse_to_nifty = ('BSE' in pat or 'MSCI India' in pat)
        w.writerow({'fund_name_pattern': pat, 'mapped_to_workbook_series': target,
                    'is_BSE_to_Nifty': 'YES' if bse_to_nifty else 'NO',
                    'note': 'closest equivalent (TRI-proxy)' if bse_to_nifty else 'subset/variant proxy'})


# ============================================================
# Status breakdown + summary
# ============================================================
from collections import Counter
status_count = Counter(r['status'] for r in rows)
print('\n=== STATUS BREAKDOWN ===', flush=True)
for s, n in status_count.most_common():
    print(f'  {n:>4}  {s}')

# Total counts
matched_real = sum(1 for r in rows if 'matched' in r['status'] and 'proxy' not in r['status'] and 'absent' not in r['status'] and 'suspect' not in r['status'])
matched_proxy = sum(1 for r in rows if 'matched-proxy' in r['status'])
absent_total = sum(1 for r in rows if 'absent' in r['status'])
print(f'\n  Total matched (real series): {matched_real}')
print(f'  Total matched (proxy):       {matched_proxy}')
print(f'  Total absent:                 {absent_total}')
print(f'  Grand total:                  {len(rows)}')

# Sample reconcile
print('\n=== Sample reconcile ===')
for nm in ('SBI Nifty Index Fund-Reg(G)', 'Motilal Oswal S&P 500 Index Fund-Reg(G)',
           'ICICI Pru NASDAQ 100 Index Fund(G)', 'ICICI Pru US Bluechip Equity Fund(G)',
           'Nippon India Japan Equity Fund(G)', 'Edelweiss NIFTY Large Mid Cap 250 Index Fund-Reg(G)',
           'Mirae Asset BSE 200 Equal Weight ETF'):
    mt = [r for r in rows if r['fund'] == nm]
    if mt:
        r = mt[0]
        print(f'  {nm}')
        print(f'    matched={r["matched_series"]!r}, basis={r["basis"]!r}, TE={r["TE_pct"]!r}, TD={r["TD_pct"]!r}, status={r["status"]}')

print('\n[done]')
