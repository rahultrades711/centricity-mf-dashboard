"""
build_manager_history.py — Fix-List 8 Feature 1 (Step A)

Converts the Morningstar Fund Manager Details workbook into a per-fund
manager-history JSON keyed by AMFI Code (string).

INPUT
  ../../../Data/Fund Manager Data/Morningstar Fund manager Details as on
  30th April 2026.xlsx (override-able via CLI arg)

  Workbook layout — uniform across all 24 sheets:
    Row 1  : header  ['Group/Investment', 'Manager Tenure (Longest)',
                       'Manager History', 'ISIN', 'AMFI Code']
    Row 2  : blank
    Row 3  : the sheet's category title (e.g. "India OE Focused Fund")
    Row 4  : a 'Benchmark 1: …' marker row
    Row 5+ : per-fund data rows (col A scheme name, col B tenure-yrs,
             col C manager history string, col D ISIN, col E AMFI code)

  Manager History string is a ;-separated list of segments in either
  shape:
        [YYYY-MM-DD -- YYYY-MM-DD] Name        ← past manager
        [YYYY-MM-DD -- ]            Name       ← current manager

OUTPUT
  Dashboard-Repo/data/manager-history-YYYY-MM-DD.json
  {
      "as_of_date": "2026-04-30",
      "source": "Morningstar Fund manager Details as on 30th April 2026.xlsx",
      "funds": {
          "<AMFI>": {
              "managers": [
                  {"name": "...", "start": "YYYY-MM-DD",
                   "end": "YYYY-MM-DD" or null,
                   "is_current": bool,
                   "tenure_years": float},
                  ...   # sorted by start asc
              ]
          },
          ...
      }
  }

  Skips rows where AMFI Code is blank or non-numeric (catches header /
  benchmark / blank rows uniformly across sheets). Skips entries that fail
  the [start -- end] regex parse silently — they get logged once at the end
  of the run as "skipped N malformed segments."

USAGE
  py scripts/build_manager_history.py
  py scripts/build_manager_history.py <path-to-xlsx>
  py scripts/build_manager_history.py <path-to-xlsx> <output-json>

Designed-for-Change rules:
  • No hardcoded fund names / counts / categories — every value flows from
    the Excel. Adding a new sheet (e.g. "India OE Sector - Defence") needs
    zero code changes here; the iterator picks it up.
  • Output JSON is keyed by AMFI string for a clean cross-source merge in
    the dashboard (Designed-for-Change §1).
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Paths

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = (
    REPO_ROOT.parent.parent.parent
    / "Data"
    / "Fund Manager Data"
    / "Morningstar Fund manager Details as on 30th April 2026.xlsx"
)

# ---------------------------------------------------------------------------
# Manager-history string parsing

SEGMENT_RE = re.compile(
    r"\[\s*(?P<start>\d{4}-\d{2}-\d{2})\s*--\s*(?P<end>\d{4}-\d{2}-\d{2})?\s*\]\s*"
    r"(?P<name>.+?)\s*$"
)


def _parse_iso_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def _years_between(start: date, end: date) -> float:
    """365.25-day calendar years between two dates."""
    return round((end - start).days / 365.25, 2)


def parse_manager_history(history: str, today: date) -> tuple[list[dict], int]:
    """Parse a Morningstar Manager History string into a list of dicts.

    Returns (entries, skipped_count). Entries are sorted ascending by
    start date.  Skipped count is the number of ;-separated segments
    that didn't match the regex (logged but not failed).
    """
    if not history or not isinstance(history, str):
        return [], 0

    entries = []
    skipped = 0
    for raw in history.split(";"):
        seg = raw.strip()
        if not seg:
            continue
        m = SEGMENT_RE.match(seg)
        if not m:
            skipped += 1
            continue
        try:
            start = _parse_iso_date(m.group("start"))
        except ValueError:
            skipped += 1
            continue
        end_str = m.group("end")
        end = _parse_iso_date(end_str) if end_str else None
        is_current = end is None
        tenure_anchor = end if end else today
        tenure = _years_between(start, tenure_anchor)
        name = m.group("name").strip()
        if not name:
            skipped += 1
            continue
        entries.append(
            {
                "name": name,
                "start": start.isoformat(),
                "end": end.isoformat() if end else None,
                "is_current": is_current,
                "tenure_years": tenure,
            }
        )
    entries.sort(key=lambda e: e["start"])
    return entries, skipped


# ---------------------------------------------------------------------------
# Workbook iteration

HEADER_AMFI_COL = 5  # column E
HEADER_NAME_COL = 1  # column A
HEADER_HISTORY_COL = 3  # column C
DATA_START_ROW = 5  # rows 1-4 are header / blank / category title / benchmark


def _coerce_amfi(value) -> str | None:
    """Return an AMFI code as a string when the cell is numeric / numeric-string,
    else None.  Filters out the category title / benchmark / blank rows."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return None
        return str(int(value))
    if isinstance(value, str):
        s = value.strip()
        if s.isdigit():
            return s
    return None


def build_funds(xlsx_path: Path, today: date | None = None) -> dict:
    """Walk every sheet of the Morningstar workbook and emit a per-fund
    dict keyed by AMFI string."""
    today = today or date.today()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    funds: dict[str, dict] = {}
    total_skipped = 0
    duplicate_amfi = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # iter_rows for read-only is faster than indexed access
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx < DATA_START_ROW:
                continue
            if not row:
                continue
            amfi = _coerce_amfi(row[HEADER_AMFI_COL - 1] if len(row) >= HEADER_AMFI_COL else None)
            if amfi is None:
                continue
            scheme_name = row[HEADER_NAME_COL - 1] if len(row) >= HEADER_NAME_COL else None
            history = row[HEADER_HISTORY_COL - 1] if len(row) >= HEADER_HISTORY_COL else None
            entries, skipped = parse_manager_history(history, today=today)
            total_skipped += skipped
            # If a fund appears in multiple sheets (Morningstar's category
            # taxonomy isn't always 1-to-1 with SEBI), the first occurrence
            # wins — they should carry identical history strings anyway.
            if amfi in funds:
                duplicate_amfi.append(amfi)
                continue
            funds[amfi] = {
                "scheme_name": scheme_name,
                "category_sheet": sheetname,
                "managers": entries,
            }
    wb.close()
    return funds, total_skipped, duplicate_amfi


# ---------------------------------------------------------------------------
# Output

def emit(xlsx_path: Path, output_path: Path, today: date | None = None) -> None:
    today = today or date.today()
    print(f"reading: {xlsx_path}")
    funds, total_skipped, duplicate_amfi = build_funds(xlsx_path, today=today)
    payload = {
        "as_of_date": today.isoformat(),
        "source": xlsx_path.name,
        "funds": funds,
    }
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"wrote:   {output_path}")
    print(f"funds:   {len(funds)}")
    print(f"skipped malformed segments: {total_skipped}")
    if duplicate_amfi:
        print(f"duplicate AMFI codes (kept first): {len(duplicate_amfi)}")

    # Optional: cross-check against the latest screener JSON for a match-rate
    # canary — same pattern the screener / analytics converters use.
    screener_dir = REPO_ROOT / "data"
    screener_files = sorted(screener_dir.glob("screener-*.json"))
    if screener_files:
        latest = screener_files[-1]
        try:
            screener = json.loads(latest.read_text(encoding="utf-8"))
            screener_amfis = {str(f.get("scheme_code")) for f in screener.get("funds", [])}
            matched = len(set(funds) & screener_amfis)
            print(f"screener match: {matched}/{len(screener_amfis)} scheme codes "
                  f"({matched/len(screener_amfis)*100:.1f}%) against {latest.name}")
        except (OSError, json.JSONDecodeError):
            pass


def main(argv: list[str]) -> int:
    if len(argv) >= 2:
        xlsx_path = Path(argv[1]).expanduser().resolve()
    else:
        xlsx_path = DEFAULT_INPUT
    if not xlsx_path.exists():
        print(f"ERROR: input file not found: {xlsx_path}", file=sys.stderr)
        return 1

    if len(argv) >= 3:
        output_path = Path(argv[2]).expanduser().resolve()
    else:
        # Filename embeds the workbook's "as on date" — extract it from the
        # filename if possible, else use today.
        m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4})", xlsx_path.name)
        if m:
            try:
                day = int(m.group(1))
                mon_name = m.group(2)
                year = int(m.group(3))
                month_num = datetime.strptime(mon_name[:3], "%b").month
                as_of = date(year, month_num, day)
            except (ValueError, KeyError):
                as_of = date.today()
        else:
            as_of = date.today()
        output_path = REPO_ROOT / "data" / f"manager-history-{as_of.isoformat()}.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    emit(xlsx_path, output_path, today=date.today())
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
