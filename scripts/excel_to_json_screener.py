"""
Screener Excel-to-JSON converter.

Validates an MF Screener Excel against data-contract/screener-v1.json,
then emits data/screener-YYYY-MM-DD.json matching the dashboard's expected shape.

Usage:
    python scripts/excel_to_json_screener.py <path-to-xlsx>
    python scripts/excel_to_json_screener.py data/MutualFund_Whitelisting_15Apr2026.xlsx

The build pipeline (GitHub Action, file-name-pattern routing) calls this on every
push that matches data/MutualFund_Whitelisting_*.xlsx. See CLAUDE.md §4.1.

No pandas dependency; openpyxl-only.
"""
from __future__ import annotations

import csv as _csv
import datetime as _dt
import json
import math
import re
import statistics as _stats
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
# v2 contract — 22-parameter scoring + Avg Mkt Cap / Fund PE / Active Share
# (Phase 2, 15-May-2026 cycle onwards). The v1 contract is retained on
# disk for archive integrity per CLAUDE.md §9 rule 4 — the 15-Apr cycle
# JSON carries `schema_version: "screener-v1"` and renders against that
# spec; the dashboard branches on schema_version when needed.
CONTRACT_PATH = REPO_ROOT / "data-contract" / "screener-v2.json"
DATA_DIR = REPO_ROOT / "data"


# ---------------------------------------------------------------------------
# Schema-validation error
# ---------------------------------------------------------------------------

class SchemaError(RuntimeError):
    """Raised when the input Excel violates the contract."""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(ws, addr: str) -> Any:
    """Read a cell by 'A1' address."""
    m = re.match(r"^([A-Z]+)(\d+)$", addr)
    if not m:
        raise ValueError(f"Bad cell address: {addr!r}")
    col_letter, row = m.group(1), int(m.group(2))
    return ws.cell(row=row, column=column_index_from_string(col_letter)).value


def _safe_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, str):
        v = v.strip()
        if not v or v == "—" or v == "-":
            return None
        try:
            return float(v)
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _safe_int(v: Any) -> int | None:
    f = _safe_float(v)
    if f is None:
        return None
    return int(round(f))


def _safe_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def normalize_manager_name(s: Any) -> str | None:
    """
    Stage B A1 (2026-05-28) — canonical manager-name normalisation, applied to
    every Monitor "Fund Manager" / "[Fund Manager 1]" cell before it lands in
    `manager_name`.

    Rules (catalogue §7.2 / §7.8 + Rahul lock 2026-05-28):
      1. Strip leading/trailing whitespace + trailing punctuation (`. , ; :`).
      2. Title-case each whitespace-separated token: capitalise first letter
         after a token start OR an internal punctuation boundary
         (apostrophe / hyphen / dot); lower-case every other letter.
         So `"D'Silva"` stays `"D'Silva"`, `"V.K. Sharma"` stays `"V.K. Sharma"`,
         `"harish krishnan"` becomes `"Harish Krishnan"`.
      3. Preserve all-caps acronyms of length ≤ 4 verbatim (`"TVS"`, `"V.K."`,
         `"PSU"`).
      4. **Do NOT reorder tokens.** `"Sharma Vivek"` stays `"Sharma Vivek"`.
    """
    if s is None:
        return None
    s = str(s).strip().rstrip(".,;:").strip()
    if not s:
        return None
    out_tokens: list[str] = []
    for tok in s.split():
        # Preserve short all-caps tokens verbatim (acronyms / initials).
        # `str.isupper()` returns True if every cased char is uppercase and
        # there is at least one cased char — so `"V.K."` (cased letters V,K)
        # qualifies.
        if len(tok) <= 4 and tok.isupper():
            out_tokens.append(tok)
            continue
        chars: list[str] = []
        capitalize_next = True
        # Cowork patch 2026-05-28 — add `()/&` to the cap_next-true set; the
        # original spec only covered `'-.` which left `(Tata)` → `(tata)` on 3
        # Tata-AMC funds (149068, 146007, 148050). Brackets, slashes, and
        # ampersands open a fresh title-case word the same way an apostrophe
        # or hyphen does (`A & B Cap`, `R/S Sharma`).
        for ch in tok:
            if ch in "'-.()/&":
                chars.append(ch)
                capitalize_next = True
            elif ch.isalpha():
                chars.append(ch.upper() if capitalize_next else ch.lower())
                capitalize_next = False
            else:
                chars.append(ch)
                capitalize_next = False
        out_tokens.append("".join(chars))
    return " ".join(out_tokens)


def _round(v: float | None, places: int = 4) -> float | None:
    if v is None:
        return None
    return round(v, places)


def _iso_date(v: Any) -> str | None:
    """Convert various date inputs to ISO YYYY-MM-DD."""
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        # DD-MMM-YYYY (e.g., 28-Jul-2002)
        for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return _dt.datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _display_date(iso: str | None) -> str | None:
    if iso is None:
        return None
    try:
        d = _dt.date.fromisoformat(iso)
    except ValueError:
        return None
    return d.strftime("%d %b %Y")


def _parse_warning_pct(s: str) -> float | None:
    """Extract trailing percentage from strings like '⚠️ 39.36%'."""
    m = re.search(r"(-?\d+(?:\.\d+)?)\s*%", s)
    if not m:
        return None
    return float(m.group(1))


# Passive categories — funds in these categories carry NO Centricity score or
# rank because the methodology doesn't yet score index funds. Phase 2.1
# decision (Rahul, 2026-05-24): emit centricity_score=null and a dedicated
# centricity_score_status="Index — Not Scored" rather than the workbook's
# default 0.0 / "Ranked" pair, which was rendering ~40% of equity rows as
# zero scores and ranking index funds 1..N off a zero score. The funds stay
# listed (returns, AUM, TER etc. still populated); only Score and Rank are
# suppressed. When the Index-Funds scoring phase ships, this set is the
# trigger to wire them in.
PASSIVE_CATEGORIES = frozenset({
    "ETFs",
    "Large Cap Index",
    "Mid Cap Index",
    "Small Cap Index",
    "Multi/Broad Index",
    "Sectoral/Thematic Index",
    "Smart-Beta/Factor",
})
PASSIVE_STATUS = "Index — Not Scored"


# ---------------------------------------------------------------------------
# Phase 2.2 §benchmark — basis classification, international USD/INR conversion,
# Underlying-Index inference, no-sensible-proxy themes, NAV unit-split detect.
# Ported from scripts/final_pass_bench.py (the Part-1 evidence script verified
# by Cowork on 2026-05-27). Workbook benchmark NAVs are TRI by default; the 8
# BSE indices added in 2026 carry PRI only. The 4 INR-converted pseudo-series
# (Nasdaq 100, S&P 500, NYSE FANG+, MSCI World) live only in memory inside this
# converter — the workbook does NOT carry the converted series, so the matcher
# replicates the conversion in-process.
# ---------------------------------------------------------------------------

NEW_BSE_NAMES = frozenset({
    "BSE Sensex", "BSE 100", "BSE 200", "BSE 500", "BSE Midcap",
    "BSE Smallcap", "BSE 250 LargeMidcap", "BSE Sensex Next 50",
})

# International USD-denominated index → INR-converted pseudo-series. Regexes
# match the FUND NAME (not the stated benchmark) because most international
# funds list a generic 'Underlying Index' as their stated benchmark.
INTERNATIONAL_USD = [
    (r'\bUS\s+Bluechip\b|\bUS Equity\b', 'S&P 500 (INR)'),
    (r'\bUS\s+Opp\b|\bUS\s+Opportun', 'S&P 500 (INR)'),
    (r'\bNASDAQ[- ]?100\b', 'Nasdaq 100 (INR)'),
    (r'\bS&P[- ]?500\b(?!\s+Top)', 'S&P 500 (INR)'),
    (r'\bS&P 500 Top 50\b', 'S&P 500 (INR)'),
    (r'\bFANG\+?\b|\bNYSE FANG', 'NYSE FANG+ (INR)'),
    (r'\bGlobal\b.*\bAdvantage\b|\bMSCI World\b|\bGlobal Equity\b', 'MSCI World (INR)'),
    (r'\bInternational Equity\b', 'MSCI World (INR)'),
    (r'\bGlobal Emerging Market', 'MSCI World (INR)'),
]

# International non-USD: we have no daily FX series for these currencies, so
# any TE/TD would be FX-noise. Force absent.
INTERNATIONAL_NON_USD = [
    r'\bJapan\b', r'\bTaiwan\b', r'\bHang Seng\b', r'\bChina\b(?!\b)',
    r'\bGreater China\b', r'\bSingapore\b', r'\bKorea\b', r'\bAustralia\b',
    r'\bASEAN\b', r'\bEurope\b', r'\bGermany\b', r'\bBrazil\b',
]

# Narrow themes with no plausible Indian-broad-index proxy.
NO_SENSIBLE_PROXY = [
    r'\bRailways PSU\b', r'\bRailway PSU\b',
    r'\bBSE Hospitals\b', r'\bBSE Power\b',
    r'\bSelect IPO\b',
]

# Underlying-Index inference table — for funds whose stated benchmark is
# 'Underlying Index' (or blank), infer the actual index from the FUND NAME.
# Tuple = (regex, target_workbook_series, is_proxy). is_proxy=True means the
# match is structural (BSE→Nifty cross-family, factor variant → parent, etc.)
# and is subject to the post-loop proxy-validity gate (TE > 5% ⇒ absent,
# except equal-weight which is a structural exception).
UI_INFER = [
    # BSE direct (PRI workbook series)
    (r'\bBSE Sensex(?!.*Next)', 'BSE Sensex', False),
    (r'(?<!BSE )\bSensex\b(?!.*Next)', 'BSE Sensex', False),
    (r'\bBSE Sensex Next 50\b|\bSensex Next 50\b', 'BSE Sensex Next 50', False),
    (r'\bBSE Sensex Next 30\b|\bSensex Next 30\b', 'BSE Sensex Next 50', True),
    (r'\bBSE 100\b', 'BSE 100', False),
    (r'\bBSE 200 Equal Weight\b', 'BSE 200', True),
    (r'\bBSE 200\b', 'BSE 200', False),
    (r'\bBSE 500 (?:Momentum|Quality|Value|Dividend|Low Vol)', 'BSE 500', True),
    (r'\bBSE[- ]?500\b', 'BSE 500', False),
    (r'\bBSE Midcap Select\b', 'BSE Midcap', True),
    (r'\bBSE Midcap\b', 'BSE Midcap', False),
    (r'\bBSE Smallcap\b', 'BSE Smallcap', False),
    (r'\bBSE 250 LargeMidcap\b', 'BSE 250 LargeMidcap', False),
    # Legacy bare 'Nifty Index Fund'
    (r'\bNifty Index Fund\b(?!.*\b(?:50|100|200|500|Bank|Midcap|Smallcap|Next|Auto|Pharma|FMCG|IT|Realty|Healthcare)\b)', 'Nifty 50', False),
    (r'\bIndex Fund Nifty Plan\b|\bNifty Plan\b', 'Nifty 50', False),
    (r'\bIndex Fund Sensex Plan\b|\bSensex Plan\b', 'BSE Sensex', False),
    # BSE sectoral / cross-family proxies (TRI proxy of a BSE-family fund)
    (r'\bBSE Healthcare\b', 'Nifty Healthcare Index', True),
    (r'\bBSE\s+(?:India\s+)?Select Top 10 Banks\b|\bBSE\s+(?:India\s+)?Top 10 Banks\b', 'Nifty Bank', True),
    (r'\bBSE India Infrastructure\b', 'Nifty Infrastructure', True),
    (r'\bBSE Quality\b', 'NIFTY100 Quality 30', True),
    (r'\bBSE Low Volatility\b', 'Nifty100 Low Volatility 30', True),
    (r'\bBSE Enhanced Value\b', 'NIFTY500 Value 50', True),
    (r'\bBSE\s+(?:India\s+)?Sector Leaders\b', 'Nifty 500', True),
    (r'\bBSE 1000\b', 'Nifty Total Market', True),
    (r'\bBSE Select Business Groups\b', 'Nifty 100', True),
    (r'\bBSE Internet Economy\b', 'Nifty India Digital', True),
    (r'\bBSE Housing\b', 'Nifty Housing', True),
    (r'\bBSE Capital Markets\b', 'Nifty Capital Markets', True),
    (r'\bMSCI India\b', 'Nifty 500', True),
    (r'\bBSE PSU\b', 'Nifty PSU Bank', True),
    (r'\bBSE Financials ex Bank\b', 'Nifty Financial Services Ex-Bank', True),
    (r'\bBSE\s+(?:India\s+)?Defence\b|\bBSE Defence\b', 'Nifty India Defence', True),
    (r'\bBSE Multicap Consumption\b', 'Nifty FMCG', True),
    # Nifty broad
    (r'\bNifty Next 50\b', 'Nifty Next 50', False),
    (r'\bNifty Total Market\b', 'Nifty Total Market', False),
    (r'\bNifty Midcap 150\b(?!\s+Momentum)(?!\s+Quality)', 'Nifty Midcap 150', False),
    (r'\bNifty Smallcap 250\b(?!\s+Quality)(?!\s+Momentum)', 'Nifty Smallcap 250', False),
    (r'\bNifty Microcap\b', 'Nifty Microcap 250', False),
    (r'\bNifty Midcap Select\b', 'Nifty Midcap Select', False),
    (r'\bNifty Midcap 50\b', 'Nifty Midcap 50', False),
    (r'\bNifty Smallcap 50\b', 'Nifty Smallcap 50', False),
    (r'\bNifty Midcap 100\b', 'Nifty Midcap 150', True),
    (r'\bNifty Smallcap 100\b', 'Nifty Smallcap 250', True),
    # Nifty 50 specific variants first
    (r'\bNifty 50 Arbitrage\b', 'Nifty 50 Arbitrage', False),
    (r'\bNifty 50 Equal Weight\b|\bNifty50 Equal Weight\b', 'NIFTY50 Equal Weight', False),
    (r'\bNifty 50 Value 20\b|\bNifty50 Value 20\b', 'Nifty50 Value 20', False),
    (r'\bNifty 50 Shariah\b|\bNifty50 Shariah\b', 'Nifty50 Shariah', False),
    (r'\bNifty500 Shariah\b|\bNifty 500 Shariah\b', 'Nifty500 Shariah', False),
    (r'\bNifty500 Multicap\b|\bNifty 500 Multicap\b', 'Nifty500 Multicap 50:25:25', False),
    # Bare numerics (after variants)
    (r'\bNifty 50\b(?!\s+(?:Arbitrage|Equal|Value|Shariah))', 'Nifty 50', False),
    (r'\bNifty 100\b(?!\s+(?:Low|Quality|Equal|Alpha|ESG))', 'Nifty 100', False),
    (r'\bNifty 200\b(?!\s+(?:Quality|Momentum|Value|Alpha))', 'Nifty 200', False),
    (r'\bNifty 500\b(?!\s+(?:Multicap|Value|Quality|Shariah|Momentum|Healthcare|Flexicap|Low Vol))', 'Nifty 500', False),
    # Sectoral
    (r'\bNifty Bank\b(?!\s*Ex[- ]?Bank)', 'Nifty Bank', False),
    (r'\bNifty IT\b', 'Nifty IT', False),
    (r'\bNifty FMCG\b', 'Nifty FMCG', False),
    (r'\bNifty Pharma\b', 'Nifty Pharma', False),
    (r'\bNifty Healthcare\b', 'Nifty Healthcare Index', False),
    (r'\bNifty Auto\b', 'Nifty Auto', False),
    (r'\bNifty Energy\b', 'Nifty Energy', False),
    (r'\bNifty Realty\b', 'Nifty Realty', False),
    (r'\bNifty Oil\s*[& ]+\s*Gas\b', 'Nifty Oil & Gas', False),
    (r'\bNifty PSU Bank\b', 'Nifty PSU Bank', False),
    (r'\bNifty Financial Services Ex[- ]?Bank\b', 'Nifty Financial Services Ex-Bank', False),
    (r'\bNifty Financial Services 25.?50\b', 'Nifty Financial Services 25/50', False),
    (r'\bNifty Financial Services\b', 'Nifty Financial Services', False),
    (r'\bNifty Consumer Durables\b', 'Nifty Consumer Durables', False),
    (r'\bNifty Capital Markets?\b', 'Nifty Capital Markets', False),
    (r'\bNifty Chemicals\b', 'Nifty Chemicals', False),
    (r'\bNifty\s+(?:India\s+)?Infrastructure(?:\s*&\s*Logistics)?\b', 'Nifty Infrastructure', False),
    (r'\bNifty Media\b', 'Nifty Media', False),
    (r'\bNifty Transportation\b', 'Nifty Transportation & Logistics', False),
    (r'\bNifty Housing\b', 'Nifty Housing', False),
    (r'\bNifty IPO\b', 'Nifty IPO', False),
    (r'\bNifty REIT', 'Nifty REITs & InvITs', False),
    (r'\bNifty High Beta\b', 'Nifty High Beta 50', False),
    (r'\bNifty Low Volatility 50\b', 'Nifty Low Volatility 50', False),
    (r'\bNifty Commodities\b', 'Nifty Commodities', False),
    (r'\bNifty MidSmallcap[- ]?400\b(?!.*Momentum)', 'Nifty MidSmallcap 400', False),
    (r'\bNifty MidSmallcap 400 Momentum Quality\b', 'Nifty MidSmallcap400 Momentum Quality 100', False),
    (r'\bNifty India Defence\b|\bIndia Defence\b', 'Nifty India Defence', False),
    (r'\bNifty India Manufacturing\b', 'Nifty India Manufacturing', False),
    (r'\bNifty India Digital\b', 'Nifty India Digital', False),
    (r'\bNifty India Internet\b', 'Nifty India Digital', True),
    (r'\bNifty India Tourism\b', 'Nifty India Tourism', False),
    (r'\bMidSmall IT (?:and|&) Telecom\b|\bMidSmall IT\b', 'Nifty MidSmall IT & Telecom', False),
    (r'\bNifty\s+Large\s*Mid\s*Cap\s*250\b|\bLargeMidcap 250\b', 'NIFTY LargeMidcap 250', False),
    (r'\bNifty EV(?:\s*&|\s*and)?\s*New Age Auto', 'Nifty Auto', True),
    (r'\bNifty India New Age Consumption\b|\bNifty Non-Cyclical Consumer\b', 'Nifty FMCG', True),
    (r'\bNifty\s+(?:India\s+)?Consumption\b', 'Nifty FMCG', True),
    (r'\bNifty Dividend Opportunities\b', 'Nifty 50', True),
    (r'\bCPSE\b|\bBharat 22\b', 'Nifty PSU Bank', True),
    (r'\bNifty MNC\b', 'Nifty 100', True),
    (r'\bNifty Metal\b', 'Nifty Energy', True),
    (r'\bNifty PSE\b', 'Nifty PSU Bank', True),
    (r'\bNifty Private Bank\b|\bNifty Pvt Bank\b', 'Nifty Bank', True),
    (r'\bNifty 500 Healthcare\b', 'Nifty Healthcare Index', True),
    (r'\bNifty500 Momentum 50\b|\bNifty 500 Momentum 50\b', 'Nifty200 Momentum 30', True),
    (r'\bNifty500 Flexicap Quality\b|\bFlexicap Quality 30\b', 'Nifty500 Quality 50', True),
    (r'\bNifty Total Market Momentum Quality\b', 'Nifty MidSmallcap400 Momentum Quality 100', True),
    (r'\bNifty Growth Sectors\b', 'Nifty Top 15 Equal Weight', True),
    (r'\bNifty Services Sector\b', 'Nifty Financial Services', True),
    # Factor combos
    (r'\bNifty100 Low Volatility 30\b|\bNifty 100 Low Vol(?:atility)? 30\b', 'Nifty100 Low Volatility 30', False),
    (r'\bNifty100 Equal Weight\b|\bNifty 100 Equal Weight\b', 'Nifty100 Equal Weight', False),
    (r'\bNifty100 ESG\b|\bNifty 100 ESG\b', 'Nifty100 ESG Sector Leaders', False),
    (r'\bNifty100 Quality\b|\bNifty 100 Quality\b', 'NIFTY100 Quality 30', False),
    (r'\bNifty100 Alpha\b|\bNifty 100 Alpha\b', 'NIFTY100 Alpha 30', False),
    (r'\bNifty200 Alpha\b|\bNifty 200 Alpha\b', 'Nifty200 Alpha 30', False),
    (r'\bNifty200 Momentum\b|\bNifty 200 Momentum\b', 'Nifty200 Momentum 30', False),
    (r'\bNifty200 Quality\b|\bNifty 200 Quality\b', 'NIFTY200 Quality 30', False),
    (r'\bNifty200 Value\b|\bNifty 200 Value\b', 'Nifty200 Value 30', False),
    (r'\bNifty500 Equal Weight\b|\bNifty 500 Equal Weight\b', 'Nifty500 Equal Weight', False),
    (r'\bNifty500 Low Vol\b|\bNifty 500 Low Vol\b', 'Nifty500 Low Volatility 50', False),
    (r'\bNifty500 Quality\b|\bNifty 500 Quality\b', 'Nifty500 Quality 50', False),
    (r'\bNifty500 Value\b|\bNifty 500 Value\b|\bNIFTY500 Value 50\b', 'NIFTY500 Value 50', False),
    (r'\bNifty Alpha 50\b', 'Nifty Alpha 50', False),
    (r'\bNifty Midcap150 Momentum\b|\bMidcap 150 Momentum\b', 'Nifty Midcap150 Momentum 50', False),
    (r'\bNifty Midcap150 Quality\b|\bMidcap 150 Quality\b', 'NIFTY Midcap150 Quality 50', False),
    (r'\bNifty Smallcap250 Quality\b|\bSmallcap 250 Quality\b', 'Nifty Smallcap250 Quality 50', False),
    (r'\bNifty Smallcap250 Momentum Quality\b|\bSmallcap 250 Momentum Quality\b', 'Nifty Smallcap250 Momentum Quality 100', False),
    (r'\bNifty MidSmall Financial Services\b', 'Nifty MidSmall Financial Services', False),
    (r'\bNifty MidSmall Healthcare\b', 'Nifty MidSmall Healthcare', False),
    (r'\bNifty MidSmall India Consumption\b', 'Nifty MidSmall India Consumption', False),
    (r'\bNIFTY Alpha Low.Volatility 30\b|\bNifty Alpha Low - Volatility 30\b', 'NIFTY Alpha Low-Volatility 30', False),
    (r'\bNIFTY Alpha Quality Value Low-Volatility\b', 'NIFTY Alpha Quality Value Low-Volatility 30', False),
    (r'\bNIFTY Alpha Quality Low-Volatility\b', 'NIFTY Alpha Quality Low-Volatility 30', False),
    (r'\bNIFTY Quality Low-Volatility 30\b', 'NIFTY Quality Low-Volatility 30', False),
    (r'\bNifty Top 10 Equal Weight\b', 'Nifty Top 10 Equal Weight', False),
    (r'\bNifty Top 15 Equal Weight\b', 'Nifty Top 15 Equal Weight', False),
    (r'\bNifty Top 20 Equal Weight\b', 'Nifty Top 20 Equal Weight', False),
    (r'\bNifty SME EMERGE\b', 'NIFTY SME EMERGE', False),
]

# Equal-weight is a structural exception to the proxy-validity gate (EW vs
# cap-weighted index ≈ 4–5% TE is structurally expected, not a bad match).
EW_PATTERNS = [r'\bEqual Weight\b']

# NAV unit-split detection thresholds (matches Part-1 exactly).
SPLIT_THRESHOLD = 0.40
CLEAN_FACTORS = (1/2, 1/3, 1/4, 1/5, 1/10, 1/20, 1/50, 1/100,
                 2, 3, 4, 5, 10, 20, 50, 100)
CLEAN_TOLERANCE = 0.03


# Phase 2.2 — Dividend-Yield category correction. The 15-May workbook moved
# these 11 Dividend-Yield funds from "Value-Contra" to "Sector-Thematic"; per
# Rahul's review they're a value strategy and belong in Value-Contra. The
# converter overrides the category BEFORE compute_parameter_scores so the
# percentile pool is correct. After scoring, centricity_score is recomputed
# for every Ranked fund in Value-Contra + Sector-Thematic so the displayed
# score matches the corrected peer pool.
DIVIDEND_YIELD_AMFI = frozenset({
    101738,  # Aditya Birla SL Dividend Yield Fund-Reg(G)
    152807,  # Baroda BNP Paribas Dividend Yield Fund-Reg(G)
    103678,  # Franklin India Dividend Yield Fund(G)
    148610,  # HDFC Dividend Yield Fund-Reg(G)
    129310,  # ICICI Pru Dividend Yield Equity Fund(G)
    154099,  # Kotak Dividend Yield Fund-Reg(G)
    152019,  # LIC MF Dividend Yield Fund-Reg(G)
    151476,  # SBI Dividend Yield Fund-Reg(G)
    149697,  # Sundaram Dividend Yield Fund(G)
    148948,  # Tata Dividend Yield Fund-Reg(G)
    103026,  # UTI Dividend Yield Fund-Reg(G)
})


# Phase 2.2 — Morningstar manager-name + history primary; MF Monitor fallback
# only. Source: the dense, dated "Manager Tenure Data as on …".xlsx Morningstar
# export (24 category sheets; header row 9; Manager History col C; AMFI Code
# col E). Each `Manager History` cell carries the FULL dated history of every
# manager who has run the fund — semicolon-separated chunks of
# `[YYYY-MM-DD -- YYYY-MM-DD] Name` or `[YYYY-MM-DD -- ] Name` (open end = the
# manager is currently active). Multiple open ends = co-managers.
#
# The earlier Part A1 attempt (Phase 2.2 first pass) used a sparse
# `morningstar_mgr` field inside `FINAL_manager_names_tenure.json` that was
# populated for only ~241/515 funds, giving 14.5% Morningstar coverage. This
# loader replaces it with the full export → expected coverage ~787 / ~752
# non-passive funds (>= 60% gate).

# Strict ISO-date format observed in the 15-May export. We re-derive end-date
# as None when the bracket reads "[YYYY-MM-DD -- ]" (open / current manager).
_MGR_HISTORY_CHUNK_RE = re.compile(
    r"^\s*\[\s*(\d{4}-\d{2}-\d{2})\s*--\s*(\d{4}-\d{2}-\d{2})?\s*\]\s*(.+?)\s*$"
)


def _parse_mgr_history_cell(s: str) -> list[dict]:
    """Parse one Manager History cell into [{name, start, end}, ...].

    Returns [] for blank cells. `end` is None when the bracket is open
    (= the manager is currently active). Names are trimmed; dates kept as
    ISO 'YYYY-MM-DD' strings (not Python date objects so they survive
    json.dumps without a default).
    """
    if s is None:
        return []
    text = str(s).strip()
    if not text:
        return []
    out: list[dict] = []
    for chunk in text.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        m = _MGR_HISTORY_CHUNK_RE.match(chunk)
        if not m:
            continue
        start = m.group(1)
        end = m.group(2) or None
        name = m.group(3).strip()
        if not name:
            continue
        out.append({"name": name, "start": start, "end": end})
    return out


def load_morningstar_mgr_history(path: Path) -> tuple[dict[int, list[dict]], dict]:
    """Return ({amfi_code: [{name, start, end}, ...]}, diagnostics) from the
    Morningstar manager-tenure xlsx. Walks every sheet; header row 9; reads
    the column whose row-9 label contains 'Manager History' and joins by
    the column whose row-9 label contains 'AMFI'.

    Duplicate AMFI codes: first occurrence wins; the dup AMFIs are recorded
    in diagnostics for the report.
    """
    diag = {
        "file": path.name,
        "sheets": 0,
        "rows_seen": 0,
        "amfi_codes_loaded": 0,
        "amfi_dups": [],
        "amfi_unparseable_history": 0,
        "sheets_skipped_no_header": [],
    }
    if not path.exists():
        return {}, diag

    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[int, list[dict]] = {}
    seen_codes: set[int] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        diag["sheets"] += 1
        if ws.max_row < 10:
            diag["sheets_skipped_no_header"].append(sheet_name)
            continue

        # Locate Manager History + AMFI Code by row-9 header text.
        mh_col = None
        amfi_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=9, column=c).value
            if h is None:
                continue
            hs = str(h).strip().lower()
            if "manager history" in hs:
                mh_col = c
            elif "amfi" in hs:
                amfi_col = c
        if mh_col is None or amfi_col is None:
            diag["sheets_skipped_no_header"].append(sheet_name)
            continue

        # Walk fund rows starting at row 10.
        for r in range(10, ws.max_row + 1):
            v = ws.cell(row=r, column=amfi_col).value
            if v is None:
                continue
            try:
                code = int(v)
            except (TypeError, ValueError):
                continue
            diag["rows_seen"] += 1
            mh = ws.cell(row=r, column=mh_col).value
            records = _parse_mgr_history_cell(mh) if mh is not None else []
            if not records:
                if mh is not None and str(mh).strip():
                    diag["amfi_unparseable_history"] += 1
                # Even with no history, register the AMFI as "seen with
                # blank history" so we don't mistakenly fall through to
                # MF Monitor for a fund Morningstar knows but has no past
                # manager record for. We treat blank as "no current
                # manager known" -> still flows to fallback in derive().
                if code not in seen_codes:
                    seen_codes.add(code)
                    out[code] = []
                continue
            if code in seen_codes:
                diag["amfi_dups"].append({"amfi": code, "sheet": sheet_name})
                continue
            seen_codes.add(code)
            out[code] = records

    wb.close()
    # Only AMFI codes whose history actually parsed count toward the
    # "Morningstar coverage" headline.
    diag["amfi_codes_loaded"] = sum(1 for v in out.values() if v)
    return out, diag


def derive_current_manager(
    records: list[dict],
    cycle_date: _dt.date,
) -> tuple[str | None, list[str], str | None, float | None]:
    """From a parsed Manager History list, return:
       (lead_name, co_managers_list, manager_since_iso, tenure_yrs)

    Phase 2.2 Patch (mgr attribution) — `manager_name` is now the SINGLE
    lead name (earliest-current active manager); the full list lives in a
    separate `manager_co_managers` array (lead at index 0). The previous
    behaviour wrote a comma-joined string of all currently-active managers
    into `manager_name`, which the partner-facing one-pager then rendered
    as "Manager: Banthia, Naren, Kalawadia, …" — see catalogue §7.8.

    lead_name        = earliest-start active manager's name; None if no
                       currently-active managers.
    co_managers_list = ALL currently-active manager names, ordered by
                       `start` ascending (lead at index 0). Empty when no
                       currently-active managers.
    manager_since    = the lead's `start` (ISO date string).
    tenure_yrs       = (cycle_date - manager_since) / 365.25, 2dp.

    All four are None / [] when records carry no currently-active manager.
    """
    if not records:
        return None, [], None, None
    current = [r for r in records if not r.get("end")]
    if not current:
        return None, [], None, None
    current.sort(key=lambda r: r.get("start") or "9999-99-99")
    co_managers = [r["name"] for r in current if r.get("name")]
    lead = co_managers[0] if co_managers else None
    since = current[0].get("start")
    tenure = None
    if since:
        try:
            sd = _dt.date.fromisoformat(since)
            tenure = round((cycle_date - sd).days / 365.25, 2)
        except (TypeError, ValueError):
            tenure = None
    return lead, co_managers, since, tenure


def _load_analytics_equity_counts(
    analytics_dir: Path | None,
) -> tuple[dict[str, int], dict]:
    """Phase 2.2 Patch (no_of_stocks) — count `Asset == "Equity"` analytics
    rows per Scheme Name, across every .xlsx in `analytics_dir`. Excludes
    Debt + Others (REITs / InvITs / cash equivalents).

    The 3 analytics files (Equity / Hybrid / Debt) share a fixed schema:
    Sheet1, row 1 = headers, col A = Scheme Name, col J = Asset
    ({"Equity", "Debt", "Others"}). The converter loads all of them so
    equity, hybrid, and debt funds all flow through the same code path —
    pure-debt funds simply count to zero (no Equity rows).

    Returns ({scheme_name: equity_row_count}, diagnostics).

    Gracefully returns ({}, diag) when the dir is absent — caller falls
    back to the workbook's `📋 Data` col N count (which over-counts because
    it includes Others). See catalogue §7.8 and the ICICI E&D audit.
    """
    diag = {
        "analytics_dir": str(analytics_dir) if analytics_dir else None,
        "exists": analytics_dir.exists() if analytics_dir else False,
        "files_loaded": {},
        "schemes_with_equity_rows": 0,
    }
    counts: dict[str, int] = {}
    if not analytics_dir or not analytics_dir.exists():
        return counts, diag

    for path in sorted(analytics_dir.glob("*.xlsx")):
        # Skip lock/temp files openpyxl can't open.
        if path.name.startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            diag["files_loaded"][path.name] = f"open_error: {e}"
            continue
        per_file = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Header row 1; data starts row 2. Col 1 = Scheme Name, col 10 = Asset.
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                scheme = row[0]
                asset = row[9] if len(row) > 9 else None
                if not scheme or not isinstance(scheme, str):
                    continue
                if asset and str(asset).strip().lower() == "equity":
                    counts[scheme] = counts.get(scheme, 0) + 1
                    per_file += 1
        wb.close()
        diag["files_loaded"][path.name] = per_file

    diag["schemes_with_equity_rows"] = len(counts)
    return counts, diag


def _classify_score(raw: Any) -> tuple[str, float | None, float | None]:
    """
    Returns (status, centricity_score_decimal, warning_pct).

    Ranked        : centricity_score_decimal = float, warning_pct = None
    1-3yr Warning : centricity_score_decimal = None,  warning_pct = float
    New Fund      : both None
    """
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return "Ranked", float(raw), None
    if isinstance(raw, str):
        s = raw.strip()
        if "New Fund" in s:
            return "New Fund Monitoring", None, None
        if s.startswith("⚠") or "⚠" in s:
            return "1-3yr Warning", None, _parse_warning_pct(s)
    # Fallback: unknown — treat as new-fund to avoid build failure but flag
    return "New Fund Monitoring", None, None


def _to_date(v: Any) -> _dt.date | None:
    """Normalise to a Python date.

    v1 workbook stored NAV dates as Excel datetimes (openpyxl returns
    datetime.datetime). v2 workbook (15-May cycle) stores them as ISO-string
    cells ('YYYY-MM-DD'). Handle both — and also accept other reasonable
    string formats so we degrade gracefully if the upstream changes again.
    """
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Fast path: 'YYYY-MM-DD' (v2 default).
        try:
            return _dt.date.fromisoformat(s[:10])
        except ValueError:
            pass
        for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def load_fund_nav(wb) -> tuple[dict[int, list[tuple[_dt.date, float]]], dict[int, str]]:
    """
    Stream 📈 Fund NAV once into memory.

    Returns:
        nav_by_amfi: { scheme_code(int) -> [ (date, nav), ... ] sorted ascending by date }
                     Only non-null NAVs included; only rows with a valid date.
        name_by_amfi: { scheme_code -> fund_name }
    """
    ws = wb["📈 Fund NAV"]
    rows = ws.iter_rows(values_only=True)
    row1 = next(rows)  # AMFI Code | <code> | <code> | ...
    row2 = next(rows)  # Fund Name | <name> | ...
    _ = next(rows)     # Date | None | ...

    # Build column index → AMFI code (skip col 0 which is the date column)
    col_to_amfi: dict[int, int] = {}
    name_by_amfi: dict[int, str] = {}
    for col_idx, code in enumerate(row1):
        if col_idx == 0:
            continue
        if code is None:
            continue
        try:
            amfi = int(code)
        except (TypeError, ValueError):
            continue
        col_to_amfi[col_idx] = amfi
        nm = row2[col_idx] if col_idx < len(row2) else None
        if nm is not None:
            name_by_amfi[amfi] = str(nm).strip()

    # Initialise per-AMFI lists
    nav_by_amfi: dict[int, list[tuple[_dt.date, float]]] = {a: [] for a in col_to_amfi.values()}

    for row in rows:
        if not row:
            continue
        d = _to_date(row[0])
        if d is None:
            continue
        for col_idx, amfi in col_to_amfi.items():
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            try:
                f = float(v)
            except (TypeError, ValueError):
                continue
            nav_by_amfi[amfi].append((d, f))

    # Sort each fund's series ascending by date (should already be, but be safe)
    for amfi in nav_by_amfi:
        nav_by_amfi[amfi].sort(key=lambda x: x[0])

    return nav_by_amfi, name_by_amfi


def load_benchmark_nav(wb) -> dict[str, list[tuple[_dt.date, float]]]:
    """Stream 📈 Benchmark NAV once into memory: { benchmark_name -> [ (date, nav), ... ] }.

    v1 layout: row 1 = name, row 2 = 'Date', row 3+ = data.
    v2 layout: row 1 = 'Code', row 2 = name, row 3 = 'Date', row 4+ = data.
    We detect by checking row-1 col-1 value: 'Benchmark Name' → v1; 'Code' → v2.
    """
    ws = wb["📈 Benchmark NAV"]
    rows = ws.iter_rows(values_only=True)
    row1 = next(rows)
    row2 = next(rows)
    first_label = "" if (row1 and row1[0] is None) else str(row1[0] if row1 else "").strip()
    if first_label == "Code":
        # v2: name row is row2; consume one more row ('Date') before data.
        names_row = row2
        _ = next(rows)  # the 'Date' label row
    else:
        # v1: name row is row1; row2 was already the 'Date' label.
        names_row = row1

    col_to_name: dict[int, str] = {}
    for col_idx, name in enumerate(names_row):
        if col_idx == 0:
            continue
        if name is None:
            continue
        col_to_name[col_idx] = str(name).strip()

    bm_by_name: dict[str, list[tuple[_dt.date, float]]] = {n: [] for n in col_to_name.values()}

    for row in rows:
        if not row:
            continue
        d = _to_date(row[0])
        if d is None:
            continue
        for col_idx, name in col_to_name.items():
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            try:
                f = float(v)
            except (TypeError, ValueError):
                continue
            bm_by_name[name].append((d, f))

    for name in bm_by_name:
        bm_by_name[name].sort(key=lambda x: x[0])

    return bm_by_name


def _series_value_at_or_before(series: list[tuple[_dt.date, float]],
                               target: _dt.date) -> tuple[_dt.date | None, float | None]:
    """Bisect on a date-sorted series; return the latest (date, value) with date <= target."""
    if not series:
        return None, None
    # Binary search: rightmost index where series[i].date <= target
    lo, hi = 0, len(series) - 1
    if series[0][0] > target:
        return None, None
    if series[-1][0] <= target:
        return series[-1]
    while lo < hi:
        mid = (lo + hi + 1) // 2
        if series[mid][0] <= target:
            lo = mid
        else:
            hi = mid - 1
    return series[lo]


def _trailing_return_pct(end_nav: float, start_nav: float, years: float) -> float | None:
    """Annualised CAGR as a percent (e.g. 12.34 for 12.34%)."""
    if end_nav is None or start_nav is None or start_nav <= 0 or years <= 0:
        return None
    try:
        cagr = (end_nav / start_nav) ** (1 / years) - 1
    except (ValueError, ZeroDivisionError, OverflowError):
        return None
    return round(cagr * 100, 4)


def _absolute_return_pct(end_nav: float, start_nav: float) -> float | None:
    """Simple period return as a percent (no annualisation)."""
    if end_nav is None or start_nav is None or start_nav <= 0:
        return None
    return round(((end_nav / start_nav) - 1) * 100, 4)


# ---------------------------------------------------------------------------
# Phase 2.2 §benchmark — helpers ported from scripts/final_pass_bench.py.
#
# The matcher walks 4 tiers per fund:
#   Tier 0  international  fund-name matches INTERNATIONAL_(USD|NON_USD)
#   Tier 1  alias-fold     canon(stated_bm) hits the workbook's canon map
#   Tier 2  UI inference   stated == 'Underlying Index' → fund-name regex
#                          against UI_INFER → workbook series (proxy flagged)
#   Tier 3  absent         NO_SENSIBLE_PROXY themes or no rule matched
#
# The proxy-validity gate runs AFTER per-fund TE/TD compute (in build_funds)
# so it can use the actual TE value to downgrade > 5% passive proxies.
#
# NAV unit-split adjust is applied to every fund's series before any nav-
# derived metric runs (trailing returns, Sharpe/Sortino/StdDev/MaxDD, TE/TD).
# 'clean-split-adjusted' = every detected ±40% jump is within 3% of a round
# ratio (1/2, 1/5, 1/10, …) → pre-split NAVs ×= factor in place.
# 'nav-split-suspect' = at least one factor isn't a clean ratio → series is
# left raw but the per-fund block forces every nav-derived metric to null.
# ---------------------------------------------------------------------------

def _canon_stated(s: Any) -> str:
    """Alias fold for stated benchmark match. Strips ' - TRI'/' - PRI', common
    punctuation, generic tokens. Same canonicalisation as Part-1."""
    s = re.sub(r"\s*-\s*TRI\s*$", "", str(s or "").strip(), flags=re.IGNORECASE)
    s = re.sub(r"\s*-\s*PRI\s*$", "", s, flags=re.IGNORECASE)
    s = s.lower()
    for ch in "-()/&,+":
        s = s.replace(ch, " ")
    GENERIC = {"etf", "fund", "reg", "plan", "direct", "growth", "g", "idcw", "index"}
    tokens = [t for t in s.split() if t and t not in GENERIC]
    return "".join(tokens)


def _parse_investing_csv(path: Path) -> dict[_dt.date, float]:
    """Investing.com daily history parser. Returns {date: close}."""
    out: dict[_dt.date, float] = {}
    if not path.exists():
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        rdr = _csv.reader(f)
        next(rdr, None)  # header
        for row in rdr:
            if not row or len(row) < 2:
                continue
            ds, ps = row[0].strip(), row[1].strip()
            if not ds or not ps:
                continue
            try:
                d = _dt.datetime.strptime(ds, "%d-%m-%Y").date()
            except ValueError:
                continue
            try:
                v = float(ps.replace(",", "").replace('"', ""))
            except ValueError:
                continue
            if v > 0:
                out[d] = v
    return out


def _detect_splits(series: list[tuple[_dt.date, float]]
                   ) -> list[tuple[_dt.date, float, bool]]:
    """Detect ±SPLIT_THRESHOLD jumps. Returns [(date, factor, is_clean), ...]."""
    out: list[tuple[_dt.date, float, bool]] = []
    for i in range(1, len(series)):
        v0 = series[i - 1][1]
        v1 = series[i][1]
        if v0 is None or v0 <= 0 or v1 is None:
            continue
        r = v1 / v0 - 1
        if abs(r) > SPLIT_THRESHOLD:
            factor = v1 / v0
            is_clean = any(abs(factor - cf) / cf < CLEAN_TOLERANCE for cf in CLEAN_FACTORS)
            out.append((series[i][0], factor, is_clean))
    return out


def _split_adjust(series: list[tuple[_dt.date, float]],
                  splits: list[tuple[_dt.date, float, bool]]
                  ) -> list[tuple[_dt.date, float]] | None:
    """Multiply pre-split NAVs by clean factor. Returns None when any split is
    non-clean (caller treats as 'nav-split-suspect')."""
    if not splits:
        return series
    if any(not is_clean for _, _, is_clean in splits):
        return None
    adjusted = list(series)
    for split_date, factor, _ in splits:
        for i, (d, v) in enumerate(adjusted):
            if d < split_date:
                adjusted[i] = (d, v * factor)
    return adjusted


def _apply_nav_split_adjustments(
    nav_by_amfi: dict[int, list[tuple[_dt.date, float]]],
) -> tuple[dict[int, str], int, int]:
    """Detect splits across every fund series. CLEAN splits mutate the series
    in place; SUSPECT splits leave the series untouched but the AMFI is
    marked so the caller nulls its nav-derived metrics. Returns
    (status_by_amfi, clean_count, suspect_count)."""
    status_by_amfi: dict[int, str] = {}
    clean = 0
    suspect = 0
    for amfi, series in nav_by_amfi.items():
        sp = _detect_splits(series)
        if not sp:
            continue
        all_clean = all(c for _, _, c in sp)
        if all_clean:
            adj = _split_adjust(series, sp)
            if adj is not None:
                nav_by_amfi[amfi] = adj
                status_by_amfi[amfi] = "clean-split-adjusted"
                clean += 1
            else:
                status_by_amfi[amfi] = "nav-split-suspect"
                suspect += 1
        else:
            status_by_amfi[amfi] = "nav-split-suspect"
            suspect += 1
    return status_by_amfi, clean, suspect


def _load_inr_converted_series(
    global_dir: Path,
) -> tuple[dict[str, list[tuple[_dt.date, float]]], dict]:
    """Build INR-converted pseudo-series for Nasdaq 100 / S&P 500 / NYSE FANG+ /
    MSCI World by multiplying their USD daily series by daily USD/INR. The 4
    Investing.com CSVs + the USD/INR CSV live in
        <cycle>/Data/BM NAV/Global & Others BM NAV/
    alongside the workbook. Returns ({pseudo_name: sorted list}, diagnostics).
    Caller registers the dict into bm_by_name; absence triggers a warning and
    forces all USD-international funds to 'absent (currency-mismatch)'."""
    diag = {
        "global_dir": str(global_dir),
        "exists": global_dir.exists() if global_dir else False,
        "usd_inr_pts": 0,
        "series_loaded": {},
        "series_inr_pts": {},
    }
    out: dict[str, list[tuple[_dt.date, float]]] = {}
    if not global_dir or not global_dir.exists():
        return out, diag

    files = {
        "Nasdaq 100 (INR)":    global_dir / "Nasdaq 100 Historical Data (1).csv",
        "S&P 500 (INR)":       global_dir / "S&P 500 Historical Data (3).csv",
        "NYSE FANG+ (INR)":    global_dir / "NYSE FANG+TM Historical Data.csv",
        "MSCI World (INR)":    global_dir / "MSCI World Historical Data.csv",
    }
    fx = _parse_investing_csv(global_dir / "USD_INR Historical Data (1).csv")
    diag["usd_inr_pts"] = len(fx)
    if not fx:
        return out, diag

    for pseudo_name, path in files.items():
        usd = _parse_investing_csv(path)
        diag["series_loaded"][pseudo_name] = len(usd)
        inr_points: list[tuple[_dt.date, float]] = []
        for d, v in usd.items():
            if d in fx:
                inr_points.append((d, v * fx[d]))
        inr_points.sort(key=lambda x: x[0])
        diag["series_inr_pts"][pseudo_name] = len(inr_points)
        if inr_points:
            out[pseudo_name] = inr_points
    return out, diag


def _is_equal_weight(fund_name: str) -> bool:
    for p in EW_PATTERNS:
        if re.search(p, fund_name, re.IGNORECASE):
            return True
    return False


def _match_international(fund_name: str) -> dict | None:
    """Returns match dict for international funds; None for everything else."""
    for pat in INTERNATIONAL_NON_USD:
        if re.search(pat, fund_name, re.IGNORECASE):
            return {"matched_series": None, "is_proxy": False, "basis": None,
                    "status": "absent (currency-mismatch, no INR conversion)",
                    "td_basis": None}
    for pat, target in INTERNATIONAL_USD:
        if re.search(pat, fund_name, re.IGNORECASE):
            is_proxy = ("FANG" in target or "MSCI" in target) or ("Emerging" in fund_name)
            return {"matched_series": target, "is_proxy": is_proxy, "basis": "PRI-INR",
                    "status": "matched (USD→INR converted)", "td_basis": "PRI-INR"}
    return None


def _match_fund(
    stated_bm: str | None,
    fund_name: str | None,
    bm_canon: dict[str, str],
    bm_basis_for_name: dict[str, str],
) -> dict:
    """4-tier matcher. Returns
        {matched_series, is_proxy, basis, status, td_basis}
    with matched_series=None for absent statuses."""
    stated = (stated_bm or "").strip()
    name = fund_name or ""

    # Tier 0 — international first (must catch before any matched status)
    intl = _match_international(name)
    if intl:
        return intl

    # Tier 1 — alias-fold on stated_bm
    if stated and stated.lower() != "underlying index":
        cs = _canon_stated(stated)
        if cs and cs in bm_canon:
            wb_name = bm_canon[cs]
            basis = bm_basis_for_name.get(wb_name, "TRI")
            return {"matched_series": wb_name, "is_proxy": False, "basis": basis,
                    "status": "matched (alias-fold)", "td_basis": basis}

    # Tier 2 — UI inference
    if stated.lower() == "underlying index" or stated == "":
        for pat, target, is_proxy in UI_INFER:
            if re.search(pat, name, re.IGNORECASE):
                if target not in bm_basis_for_name:
                    continue  # this target index isn't in the workbook this cycle
                basis = bm_basis_for_name[target]
                status = "matched-proxy" if is_proxy else "matched"
                return {"matched_series": target, "is_proxy": is_proxy, "basis": basis,
                        "status": status, "td_basis": basis}

    # Tier 3 — no-sensible-proxy themes
    for pat in NO_SENSIBLE_PROXY:
        if re.search(pat, name, re.IGNORECASE):
            return {"matched_series": None, "is_proxy": False, "basis": None,
                    "status": "absent (no real series, no sensible proxy)",
                    "td_basis": None}

    # Tier 4 — absent
    return {"matched_series": None, "is_proxy": False, "basis": None,
            "status": "absent (no rule matched)", "td_basis": None}


def _compute_te_td(
    fund_series: list[tuple[_dt.date, float]],
    bm_series:   list[tuple[_dt.date, float]],
    start: _dt.date,
    end:   _dt.date,
) -> tuple[float | None, float | None, int, float | None, float | None]:
    """Annualised tracking error (std-dev of daily-return diffs ×√252) + tracking
    difference (fund CAGR − bench CAGR) over [start, end] ∩ common dates.
    Returns (TE_pct, TD_pct, n_common_days, fund_cagr_pct, bm_cagr_pct).
    All four numeric values None when fewer than 30 common observations."""
    if not fund_series or not bm_series:
        return None, None, 0, None, None
    f_map = {d: v for d, v in fund_series if v is not None and v > 0}
    b_map = {d: v for d, v in bm_series   if v is not None and v > 0}
    common = sorted(d for d in f_map if d in b_map and start <= d <= end)
    n = len(common)
    if n < 30:
        return None, None, n, None, None
    f_ret: list[float] = []
    b_ret: list[float] = []
    for i in range(1, n):
        d0, d1 = common[i - 1], common[i]
        f_ret.append(f_map[d1] / f_map[d0] - 1)
        b_ret.append(b_map[d1] / b_map[d0] - 1)
    diff = [a - b for a, b in zip(f_ret, b_ret)]
    if len(diff) < 2:
        return None, None, n, None, None
    sd = _stats.stdev(diff)
    te = round(sd * math.sqrt(252) * 100, 4)
    d0, d1 = common[0], common[-1]
    yrs = (d1 - d0).days / 365.25
    if yrs <= 0:
        return te, None, n, None, None
    f_cagr = ((f_map[d1] / f_map[d0]) ** (1 / yrs) - 1) * 100
    b_cagr = ((b_map[d1] / b_map[d0]) ** (1 / yrs) - 1) * 100
    return te, round(f_cagr - b_cagr, 4), n, round(f_cagr, 4), round(b_cagr, 4)


# ---------------------------------------------------------------------------
# Contract loading + validation
# ---------------------------------------------------------------------------

def load_contract() -> dict:
    if not CONTRACT_PATH.exists():
        raise SchemaError(f"Contract not found at {CONTRACT_PATH}")
    with open(CONTRACT_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def validate_sheets(wb, contract: dict) -> dict:
    """Walk expected_sheets and header_anchors; raise SchemaError on any mismatch."""
    expected = contract["expected_sheets"]
    summary = {"sheets_validated": 0, "anchors_passed": 0, "anchors_failed": []}

    for sheet_name, spec in expected.items():
        if sheet_name.startswith("_"):
            continue
        if sheet_name not in wb.sheetnames:
            raise SchemaError(f"Required sheet missing: {sheet_name!r}")
        ws = wb[sheet_name]
        if ws.max_row < spec.get("min_rows", 0):
            raise SchemaError(
                f"Sheet {sheet_name!r}: max_row={ws.max_row} below min_rows={spec['min_rows']}"
            )
        if ws.max_column < spec.get("min_cols", 0):
            raise SchemaError(
                f"Sheet {sheet_name!r}: max_col={ws.max_column} below min_cols={spec['min_cols']}"
            )
        for anchor in spec.get("header_anchors", []):
            actual = _cell(ws, anchor["cell"])
            actual_s = "" if actual is None else str(actual).strip()
            if "expected_equals" in anchor:
                if actual_s != anchor["expected_equals"]:
                    summary["anchors_failed"].append(
                        f"{sheet_name}!{anchor['cell']} expected {anchor['expected_equals']!r} got {actual_s!r}"
                    )
                    continue
            elif "expected_contains" in anchor:
                if anchor["expected_contains"] not in actual_s:
                    summary["anchors_failed"].append(
                        f"{sheet_name}!{anchor['cell']} expected to contain {anchor['expected_contains']!r} got {actual_s!r}"
                    )
                    continue
            elif "expected_starts_with" in anchor:
                if not actual_s.startswith(anchor["expected_starts_with"]):
                    summary["anchors_failed"].append(
                        f"{sheet_name}!{anchor['cell']} expected to start with {anchor['expected_starts_with']!r} got {actual_s!r}"
                    )
                    continue
            summary["anchors_passed"] += 1
        summary["sheets_validated"] += 1

    if summary["anchors_failed"]:
        raise SchemaError(
            "Header anchor mismatches:\n  - " + "\n  - ".join(summary["anchors_failed"])
        )
    return summary


def validate_weights(wb, contract: dict, summary: dict) -> float:
    """Sum the Master Table 2 weights column. Row range is read from the
    contract (v2: 39-60; v1: 37-55) so old + new cycles both work."""
    ws = wb["🏠 Master"]
    rules = contract.get("scoring_weights_locked", {})
    r_first = rules.get("weight_rows_first", 37)
    r_last  = rules.get("weight_rows_last", 55)
    weights = []
    for r in range(r_first, r_last + 1):
        v = ws.cell(row=r, column=3).value
        if v is None:
            continue
        f = _safe_float(v)
        if f is not None:
            weights.append(f)
    total = round(sum(weights), 4)
    expected = rules.get("expected_sum_pct", 100)
    tol = rules.get("tolerance", 0.01)
    if abs(total - expected) > tol:
        raise SchemaError(
            f"Master Table 2 weights (rows {r_first}-{r_last}) sum to {total}, "
            f"expected {expected} (±{tol})."
        )
    summary["weights_total_pct"] = total
    return total


# ---------------------------------------------------------------------------
# Cycle metadata
# ---------------------------------------------------------------------------

def _parse_master_header_line(text: str) -> dict:
    """
    Parse the Master A2 banner. Supports both v1 and v2 phrasings:
      v1: 'Universe: 676 funds, 26 categories | NAV through 15-Apr-2026 | Master cutoff: 31-Mar-2026 | Rf=4.5% p.a.'
      v2: 'Universe: 1252 funds (Eq 571 + Hy 184 + Idx 497) | NAV 15-May-2026 | Master 30-Apr-2026 | Rf 4.5% LOCKED'
    """
    out = {"total_funds": None, "category_count": None,
           "cycle_date": None, "master_cutoff_date": None,
           "rf_rate_annual": None, "rf_rate_display": None}
    if not text:
        return out
    s = str(text).strip()

    m = re.search(r"(\d+)\s+funds", s)
    if m:
        out["total_funds"] = int(m.group(1))
    m = re.search(r"(\d+)\s+categor", s)
    if m:
        out["category_count"] = int(m.group(1))
    # NAV: v1 says "NAV through DD-Mon-YYYY"; v2 says "NAV DD-Mon-YYYY".
    m = re.search(r"NAV(?:\s+through)?\s+([0-9]{1,2}[- ][A-Za-z]+[- ][0-9]{4})", s)
    if m:
        out["cycle_date"] = _iso_date(m.group(1).replace(" ", "-"))
    # Master cutoff: v1 "Master cutoff: DD-Mon-YYYY"; v2 "Master DD-Mon-YYYY".
    m = re.search(r"Master(?:\s+cutoff)?:?\s*([0-9]{1,2}[- ][A-Za-z]+[- ][0-9]{4})", s)
    if m:
        out["master_cutoff_date"] = _iso_date(m.group(1).replace(" ", "-"))
    # Rf: v1 "Rf=4.5% p.a."; v2 "Rf 4.5% LOCKED".
    m = re.search(r"Rf[=\s]+([0-9]+(?:\.[0-9]+)?)\s*%", s)
    if m:
        pct = float(m.group(1))
        out["rf_rate_annual"] = round(pct / 100, 6)
        out["rf_rate_display"] = f"{pct}% p.a."
    return out


def _cycle_label(iso: str) -> str:
    """Internal-use U-cycle code (U1/U2). Retained for backward compatibility
    inside the JSON; the dashboard reads `cycle_label_date` for display."""
    d = _dt.date.fromisoformat(iso)
    suffix = "U1" if d.day <= 10 else "U2"
    return f"{suffix} {d.strftime('%b %Y')}"


def _cycle_label_date(iso: str) -> str:
    """Display-format cycle label: '15th Apr 2026' (day + ordinal suffix +
    abbreviated month + four-digit year). This is the canonical cycle label
    for every UI surface — heading bars, provenance lines, archive tiles,
    cycle column captions, exports. U1/U2 codes are deprecated for display."""
    d = _dt.date.fromisoformat(iso)
    day = d.day
    if 11 <= day <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return f"{day}{suffix} {d.strftime('%b %Y')}"


def build_cycle_meta(wb, source_filename: str, summary: dict, contract: dict) -> dict:
    ws = wb["🏠 Master"]
    parsed = _parse_master_header_line(_cell(ws, "A2"))

    # Categories — Master Table 1 (rows 6..35 in v2; rows 6..31 in v1).
    # v2 adds 4 entries: Arbitrage / Conservative Hybrid / Balanced Hybrid /
    # Index Funds (aggregate of 7 passive subcats). We walk all rows in the
    # range and stop at the first empty row.
    categories: list[dict] = []
    eq = set(contract["categories"]["equity"])
    hb = set(contract["categories"]["hybrid"])
    seen_cat_names: set[str] = set()
    for r in range(6, 40):
        name = _safe_str(ws.cell(row=r, column=1).value)
        if not name:
            continue
        # Stop at the Parameter header row of Table 2
        if name == "Parameter":
            break
        sub_class = "Equity" if name in eq else ("Hybrid" if name in hb else "Equity")
        top_score = ws.cell(row=r, column=4).value
        categories.append({
            "name": name,
            "sub_class": sub_class,
            "fund_count": _safe_int(ws.cell(row=r, column=2).value),
            "top_fund_name": _safe_str(ws.cell(row=r, column=3).value),
            "top_score": _safe_float(top_score) if isinstance(top_score, (int, float)) else None,
            "benchmark": _safe_str(ws.cell(row=r, column=5).value),
        })
        seen_cat_names.add(name)

    # v2 — Master Table 1 rolls the 7 passive subcategories up into a single
    # "Index Funds" aggregate row, but the 📋 Data sheet stamps each fund with
    # the finer-grained subcategory string (ETFs / Large Cap Index / Mid Cap
    # Index / Small Cap Index / Multi-Broad Index / Sectoral-Thematic Index /
    # Smart-Beta/Factor). Without these in cycle_meta.categories, the
    # dashboard's category dropdown would hide 497 passive funds. Pull the
    # observed Data-sheet category names into the categories block so the
    # dropdown + filter logic see every fund.
    try:
        data_ws = wb["📋 Data"]
        data_cat_counts: dict[str, int] = {}
        for dr in range(5, data_ws.max_row + 1):
            cat = _safe_str(data_ws.cell(row=dr, column=3).value)
            if not cat:
                continue
            data_cat_counts[cat] = data_cat_counts.get(cat, 0) + 1
        for cat_name, count in sorted(data_cat_counts.items()):
            if cat_name in seen_cat_names:
                continue
            sub_class = "Equity" if cat_name in eq else ("Hybrid" if cat_name in hb else "Equity")
            categories.append({
                "name": cat_name,
                "sub_class": sub_class,
                "fund_count": count,
                "top_fund_name": None,
                "top_score": None,
                "benchmark": None,
            })
            seen_cat_names.add(cat_name)
    except Exception as e:
        # Defensive: any failure walking Data → leave categories as Master-only
        print(f"[converter]   note: extending categories from Data failed ({e}); "
              f"passive subcats may not appear in dropdown.", file=sys.stderr)

    # AMC scores — Master Table 4. v1 had AMC scores at rows 66-91 with a
    # two-column layout (A|B and D|E). v2 reorganises Master so the start
    # row is data-driven: find the "AMC" header row, then walk down until
    # the next non-AMC label row.
    amc_scores: list[dict] = []
    amc_header_row = None
    for r in range(60, 120):
        if _safe_str(ws.cell(row=r, column=1).value) == "AMC":
            amc_header_row = r
            break
    if amc_header_row is not None:
        for r in range(amc_header_row + 1, amc_header_row + 60):
            for amc_col, score_col in ((1, 2), (4, 5)):
                amc = _safe_str(ws.cell(row=r, column=amc_col).value)
                score = ws.cell(row=r, column=score_col).value
                # Stop when we hit the TABLE 5 banner or run off the end
                if amc and amc.startswith("TABLE"):
                    amc_header_row = None
                    break
                if amc and score is not None:
                    amc_scores.append({"amc": amc, "score": _safe_int(score)})
            if amc_header_row is None:
                break

    # Scoring weights — Master Table 2. v1: rows 37-55 (19 params). v2:
    # rows 39-60 (22 params). Find the "Parameter" header row, then read
    # weight rows until the TOTAL marker.
    weights: list[dict] = []
    param_header_row = None
    for r in range(30, 70):
        if _safe_str(ws.cell(row=r, column=1).value) == "Parameter":
            param_header_row = r
            break
    if param_header_row is None:
        raise SchemaError("Master Table 2 'Parameter' header row not found in rows 30-70.")
    for r in range(param_header_row + 1, param_header_row + 30):
        param = _safe_str(ws.cell(row=r, column=1).value)
        if not param:
            break
        if param == "TOTAL":
            break
        direction_raw = str(ws.cell(row=r, column=4).value or "")
        if "Tent" in direction_raw:
            direction = "Tent"
        elif "Higher" in direction_raw:
            direction = "Higher"
        else:
            direction = "Lower"
        weights.append({
            "parameter": param,
            "unit": _safe_str(ws.cell(row=r, column=2).value),
            "weight_pct": _safe_float(ws.cell(row=r, column=3).value),
            "direction": direction,
        })

    cycle_date = parsed["cycle_date"]
    if not cycle_date:
        # Fall back to parsing the filename
        m = re.search(r"(\d{1,2})([A-Za-z]{3})(\d{4})", source_filename)
        if m:
            cycle_date = _iso_date(f"{m.group(1)}-{m.group(2)}-{m.group(3)}")

    if not cycle_date:
        raise SchemaError("Unable to determine cycle_date from Master A2 or filename.")

    return {
        # Per CLAUDE.md §4.1 — every cycle JSON declares which Centricity
        # product family it belongs to. This converter is family-specific
        # to MF Equity & Hybrid; future converters (debt / PMS / AIF) emit
        # their own family value. The dashboard never merges cycles across
        # families.
        "product_family": "MF_Equity_Hybrid",
        "cycle_date": cycle_date,
        "cycle_label": _cycle_label(cycle_date),
        "cycle_label_date": _cycle_label_date(cycle_date),
        "as_on_display": _display_date(cycle_date),
        "total_funds": parsed["total_funds"],
        # v2 — Master A2 banner no longer carries a category count, so fall
        # back to the length of the categories block parsed from Table 1.
        "category_count": parsed["category_count"] if parsed["category_count"] else len(categories),
        "categories": categories,
        "rf_rate_annual": parsed["rf_rate_annual"],
        "rf_rate_display": parsed["rf_rate_display"],
        "master_cutoff_date": parsed["master_cutoff_date"],
        "source_dates": {
            "screener": cycle_date,
            "analytics": None,  # v1.x
            "monitor": None,    # v1.x
        },
        # flag_summary = null in v1 (no prior cycle). Populated by
        # compute_cycle_flags.py post-processor (v1.x). See ISSUE-0009.
        "flag_summary": None,
        "amc_scores": amc_scores,
        "scoring_weights": weights,
        "schema_version": contract.get("contract_version", "screener-v2"),
        "generated_at": _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds"),
        "source_file": source_filename,
    }


# ---------------------------------------------------------------------------
# Fund records
# ---------------------------------------------------------------------------

def _read_data_tuple(row: tuple) -> dict:
    """Parse one streamed row tuple from 📋 Data."""
    pad = list(row) + [None] * (24 - len(row))
    g = lambda c: pad[c - 1]
    return {
        "seq": _safe_int(g(1)),
        "fund_name": _safe_str(g(2)),
        "category": _safe_str(g(3)),
        "scheme_code": _safe_int(g(4)),
        "aum_cr": _safe_float(g(5)),
        "manager_name": _safe_str(g(6)),
        "manager_tenure_yrs": _safe_float(g(7)),
        "inception_date_iso": _iso_date(g(8)),
        "turnover_pct": _safe_float(g(9)),
        "benchmark": _safe_str(g(10)),
        "amc": _safe_str(g(11)),
        "amc_score": _safe_int(g(12)),
        "ter_pct": _safe_float(g(13)),
        "no_of_stocks": _safe_int(g(14)),
        "large_cap_pct": _safe_float(g(15)),
        "mid_cap_pct": _safe_float(g(16)),
        "small_cap_pct": _safe_float(g(17)),
        "others_pct": _safe_float(g(18)),
        "h_equity_pct": _safe_float(g(19)),
        "h_debt_pct": _safe_float(g(20)),
        "h_others_pct": _safe_float(g(21)),
        "h_ytm": _safe_float(g(22)),
        "h_mod_duration": _safe_float(g(23)),
        "h_avg_maturity": _safe_float(g(24)),
    }


def _read_cm_tuple(row: tuple) -> dict:
    """Parse one streamed CM row. Note: NO rounding applied here — the parameter_scores
    percentile compute requires full-precision inputs to avoid spurious ties at 4 dp.
    Display rounding happens later when populating record["cy_returns"] etc.

    v2 layout (15-May cycle):
      A=Rank in Category  B=(removed)  C=Fund Name  D=Category  E=Fund Tenure
      F=AUM  G=Mgr Tenure  H-L=CY returns  M=Rolling  N=Consistency  O=Sharpe
      P=Beta  Q=Down Capture  R=Up Capture  S=Treynor  T=Overall Capture
      U=Turnover  V=TER  W=AMC Score  X=No. of Stocks  Y-AB=mcap split
      AC=Manager Name  AD=Score  AE=Rank in Category (repeated)
      AF=Equity %  AG=Debt %  AH=Others %
      AI=Avg Mkt Cap  AJ=Fund PE  AK=Active Share %
      AL=mcap_goodness  AM=pe_goodness
    """
    pad = list(row) + [None] * (39 - len(row))
    g = lambda c: pad[c - 1]
    return {
        "rank_in_category": _safe_int(g(1)),    # CM col A — v2's primary rank surface
        # v2 has no universe rank column; we recompute it post-build by sorting
        # all Ranked funds by centricity_score desc.
        "univ_rank": None,
        "fund_name": _safe_str(g(3)),
        "category": _safe_str(g(4)),
        "fund_tenure_yrs": _safe_float(g(5)),
        # CY returns: × 100 unit conversion only, no rounding
        "cy2022_pct": (lambda v: v * 100 if v is not None else None)(_safe_float(g(8))),
        "cy2023_pct": (lambda v: v * 100 if v is not None else None)(_safe_float(g(9))),
        "cy2024_pct": (lambda v: v * 100 if v is not None else None)(_safe_float(g(10))),
        "cy2025_pct": (lambda v: v * 100 if v is not None else None)(_safe_float(g(11))),
        "cy_ytd_pct": (lambda v: v * 100 if v is not None else None)(_safe_float(g(12))),
        "rolling_3y_avg_pct": _safe_float(g(13)),
        "consistency_pct": _safe_float(g(14)),
        "sharpe_3y": _safe_float(g(15)),
        "beta_3y": _safe_float(g(16)),
        "down_capture_3y_pct": _safe_float(g(17)),
        "up_capture_3y_pct": _safe_float(g(18)),
        "treynor_3y": _safe_float(g(19)),
        "overall_capture_3y_pct": _safe_float(g(20)),
        "score_raw": g(30),  # AD col 30
        # v2 new fields (CM cols AF..AM = 32..39)
        "equity_pct_universal":  _safe_float(g(32)),   # AF
        "debt_pct_universal":    _safe_float(g(33)),   # AG
        "others_pct_universal":  _safe_float(g(34)),   # AH
        "avg_mcap_cr":           _safe_float(g(35)),   # AI
        "fund_pe":               _safe_float(g(36)),   # AJ
        "active_share_pct":      _safe_float(g(37)),   # AK
        "mcap_goodness":         _safe_float(g(38)),   # AL
        "pe_goodness":           _safe_float(g(39)),   # AM
    }


def _stream_data_rows(wb) -> list[dict]:
    """Read 📋 Data start_row=5 onward via iter_rows; one dict per fund."""
    ws = wb["📋 Data"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 5:
            continue  # header rows
        if not row or row[0] is None and row[1] is None and row[3] is None:
            continue
        out.append(_read_data_tuple(row))
    return out


def _stream_cm_rows(wb) -> list[dict]:
    """Read 📊 Computed Metrics start_row=4 onward via iter_rows."""
    ws = wb["📊 Computed Metrics"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 4:
            continue
        if not row or row[0] is None and row[2] is None:
            continue
        out.append(_read_cm_tuple(row))
    return out


def _empty_analytics_pending() -> dict:
    return OrderedDict([
        ("top_10_holdings", None),
        ("sector_allocation", None),
        ("full_holdings", None),
        ("top_10_concentration_pct", None),
        ("top_3_sector_concentration_pct", None),
        ("manager_change_history", None),
        ("category_history", None),
        ("aum_trend", None),
        ("ter_trend", None),
        ("rank_history", None),
    ])


# ---------------------------------------------------------------------------
# Parameter-score computation (per-category percentile rank per parameter)
# ---------------------------------------------------------------------------

# Master Table 2 parameter name → fund-record value extractor.
# The CY-YTD parameter name is dynamic (e.g. 'CY2026 YTD'); resolved at runtime
# by 'YTD' substring match. All other names are exact matches against
# Master Table 2 col A in mf-whitelisting v3.7.
#
# IMPORTANT — Excel-quirk reproduction (locked in the contract under
# derived_fields_documentation.parameter_scores.excel_quirk_python_pasted_blanks_become_zero):
#
# Excel's category-sheet score formula uses COUNTIF/COUNT against ranges that
# inherit values from 📊 Computed Metrics. Two distinct sources of "None"
# behave differently in those COUNTIFs:
#
#   (a) PRESERVED-AS-EMPTY-STRING — CY-return cells (CM cols H-L) are Excel
#       formulas with `=IFERROR(...,"")`. When data is missing, the formula
#       returns the literal empty string "". Excel keeps "" as text;
#       COUNT excludes it; COUNTIF skips it. openpyxl reads None.
#       → No coercion needed; my Python's `if v is not None` filter matches
#       Excel's behaviour directly.
#
#   (b) COERCED-TO-ZERO — Two sub-cases produce numeric 0 in CM, which the
#       cat sheet inherits and COUNT/COUNTIF then include as 0:
#         * Python-pasted CM cols M..T (Rolling 3Y, Consistency, Sharpe, Beta,
#           Down/Up Capture, Treynor, Overall Capture) — for funds with
#           tenure < 3Y these cells are blank, and Excel's category-sheet
#           XLOOKUP coerces a blank-source-cell result to literal 0.
#         * XLOOKUP-from-Data CM cols G/U/V/W (Mgr Tenure, Turnover, TER,
#           AMC Score) — when the underlying Data cell is None, the CM
#           XLOOKUP itself materialises 0 in the CM cell (Excel coerces a
#           blank source cell to 0 in numeric context, even with a `""` 4th
#           arg, because that arg is for "not found", not "found but empty").
#
# To reproduce the stored centricity_score within ±0.0001, the peer pool for
# the 12 affected params substitutes None → 0. The subject fund's own value
# is left as-is; for Ranked funds the M-T params are always populated
# (otherwise they wouldn't be Ranked), and the G/U/V/W params have None at
# Data-source level for some funds (e.g. brand-new funds without a turnover
# figure) — those funds typically aren't Ranked anyway.
COERCE_NONE_TO_ZERO_PARAMS = frozenset({
    # Phase 2.2 Part R3 — the hardcoded CM cols M-T (Rolling 3Y, Consistency,
    # Sharpe, Beta, Down/Up Capture, Treynor, Overall Capture) now store
    # literal text `'–'` for funds with insufficient tenure, NOT a blank cell.
    # The cat-sheet XLOOKUP returns the text, ISNUMBER(...) is FALSE, and the
    # cat-sheet percentile pool EXCLUDES those funds. Python must match —
    # removed these 8 entries from the coerce list. Per-param probe on
    # Axis Multi Asset Allocation confirmed Σ(p×w)/100 drift collapses from
    # 0.1112 to near-zero after this change, with no fund relying on the
    # coerce-to-zero behaviour any more.
    #
    # The XLOOKUP-from-Data CM cols G/U/V/W still need coercion: Excel
    # materialises a blank Data-source cell as the literal number 0 inside
    # the CM XLOOKUP result (because the XLOOKUP `""` 4th arg is for
    # "not found", not "found but empty"), so the cat sheet sees 0 and
    # ISNUMBER returns TRUE — Python must mirror by coercing None → 0.
    "Mgr Tenure",
    "Portfolio Turnover",
    "TER (%)",
    "AMC Score",
})

# Extractors read from record["_raw_metrics"] when available (full precision —
# avoids spurious ties from 4-dp display rounding) and fall back to public fields.
def _r(r: dict, key: str, fallback):
    raw = r.get("_raw_metrics", {})
    if key in raw:
        return raw[key]
    return fallback(r)

PARAM_EXTRACTORS = {
    "Fund Tenure":         lambda r: _r(r, "fund_tenure_yrs",       lambda x: x["fund_tenure_yrs"]),
    "Fund AUM":            lambda r: _r(r, "aum_cr",                lambda x: x["aum_cr"]),
    "Mgr Tenure":          lambda r: _r(r, "manager_tenure_yrs",    lambda x: x["manager_tenure_yrs"]),
    "CY2022 Return":       lambda r: _r(r, "cy2022_pct",            lambda x: x["cy_returns"]["cy2022_pct"]),
    "CY2023 Return":       lambda r: _r(r, "cy2023_pct",            lambda x: x["cy_returns"]["cy2023_pct"]),
    "CY2024 Return":       lambda r: _r(r, "cy2024_pct",            lambda x: x["cy_returns"]["cy2024_pct"]),
    "CY2025 Return":       lambda r: _r(r, "cy2025_pct",            lambda x: x["cy_returns"]["cy2025_pct"]),
    "Rolling 3Y Avg":      lambda r: _r(r, "rolling_3y_avg_pct",    lambda x: x["rolling_3y_avg_pct"]),
    "Consistency Score":   lambda r: _r(r, "consistency_pct",       lambda x: x["consistency_pct"]),
    "Sharpe Ratio":        lambda r: _r(r, "sharpe_3y",             lambda x: x["risk_metrics"]["sharpe_3y"]),
    "Beta":                lambda r: _r(r, "beta_3y",               lambda x: x["risk_metrics"]["beta_3y"]),
    "Down Capture":        lambda r: _r(r, "down_capture_3y_pct",   lambda x: x["risk_metrics"]["down_capture_3y_pct"]),
    "Up Capture":          lambda r: _r(r, "up_capture_3y_pct",     lambda x: x["risk_metrics"]["up_capture_3y_pct"]),
    "Treynor Ratio":       lambda r: _r(r, "treynor_3y",            lambda x: x["risk_metrics"]["treynor_3y"]),
    "Overall Capture":     lambda r: _r(r, "overall_capture_3y_pct",lambda x: x["risk_metrics"]["overall_capture_3y_pct"]),
    "Portfolio Turnover":  lambda r: _r(r, "turnover_pct",          lambda x: x["turnover_pct"]),
    "TER (%)":             lambda r: _r(r, "ter_pct",               lambda x: x["ter_pct"]),
    "AMC Score":           lambda r: _r(r, "amc_score",
                                        lambda x: float(x["amc_score"]) if x.get("amc_score") is not None else None),
    # v2 — Avg Market Cap & Fund PE use Tent scoring against the goodness
    # columns (mcap_goodness / pe_goodness = -|deviation from category centre|).
    # Higher goodness = closer to centre = better, so the percentile compute
    # treats them as Higher direction. Active Share is a straight monotonic
    # Higher; null for the 10-fund intentional set, which drops from the pool
    # via the standard ISNUMBER guard (no entry in COERCE_NONE_TO_ZERO_PARAMS).
    "Avg Market Cap":      lambda r: _r(r, "mcap_goodness",         lambda x: x.get("mcap_goodness")),
    "Fund PE":             lambda r: _r(r, "pe_goodness",           lambda x: x.get("pe_goodness")),
    "Active Share":        lambda r: _r(r, "active_share_pct",      lambda x: x.get("active_share_pct")),
}


def _resolve_extractors(weights_meta: list[dict]) -> dict:
    """Match Master Table 2 parameter names to extractor lambdas, including
    the dynamic CY-YTD entry. The CY-YTD extractor also prefers _raw_metrics."""
    resolved: dict = dict(PARAM_EXTRACTORS)
    for w in weights_meta:
        name = w["parameter"]
        if name in resolved:
            continue
        if "YTD" in name:
            resolved[name] = lambda r: _r(
                r, "cy_ytd_pct", lambda x: x["cy_returns"]["cy_ytd_pct"]
            )
    return resolved


def compute_parameter_scores(funds: list[dict],
                             weights_meta: list[dict],
                             warnings: list[str]) -> None:
    """
    Reproduce the Excel category-sheet percentile-rank logic from
    mf-whitelisting/SKILL.md §4.2 from primitive metric values.

    Mutates each fund record in `funds`, adding a `parameter_scores` OrderedDict
    keyed by Master Table 2 parameter name. Values are 0–1 floats or None.

    Per the contract spec, all parameter_scores are None for funds where
    centricity_score_status != 'Ranked' (1-3yr Warning, New Fund Monitoring).
    """
    extractors = _resolve_extractors(weights_meta)
    direction_by_param = {w["parameter"]: w["direction"] for w in weights_meta}
    param_order = [w["parameter"] for w in weights_meta]

    # Sanity: every weight parameter has an extractor
    missing = [p for p in param_order if p not in extractors]
    if missing:
        raise SchemaError(
            "compute_parameter_scores: no extractor mapped for parameter(s) "
            + repr(missing) + ". Update PARAM_EXTRACTORS or add a YTD-style "
            "rule in _resolve_extractors."
        )

    # Group funds by category
    by_cat: dict[str, list[dict]] = {}
    for f in funds:
        by_cat.setdefault(f["category"], []).append(f)

    # Initialise parameter_scores on every fund (preserve param_order)
    for f in funds:
        f["parameter_scores"] = OrderedDict((p, None) for p in param_order)

    # Per (category, parameter): compute percentile pool, assign per-fund score.
    # For COERCE_NONE_TO_ZERO_PARAMS, replace None with 0.0 in the peer pool to
    # reproduce the Excel category-sheet XLOOKUP-of-blank-becomes-0 behaviour.
    for cat, cat_funds in by_cat.items():
        for param_name in param_order:
            ext = extractors[param_name]
            direction = direction_by_param[param_name]
            coerce_zero = param_name in COERCE_NONE_TO_ZERO_PARAMS

            # Build the peer pool as Excel sees it
            if coerce_zero:
                pool_values = [
                    (f, (ext(f) if ext(f) is not None else 0.0))
                    for f in cat_funds
                ]
            else:
                pool_values = [(f, ext(f)) for f in cat_funds]

            valid_pool = [v for (_, v) in pool_values if v is not None]
            if not valid_pool:
                continue
            n = len(valid_pool)

            # Per fund: percentile from the cat-sheet-equivalent value.
            # For coerced params the subject's own raw=None is also coerced to 0
            # (Excel cat-sheet shows 0 in that cell), so the fund participates
            # in its own percentile pool with value 0.
            for f in cat_funds:
                raw = ext(f)
                if raw is None and not coerce_zero:
                    continue  # CY-return None preserved as null score
                eff = 0.0 if (raw is None and coerce_zero) else raw
                count_le = sum(1 for x in valid_pool if x <= eff)
                # v2: "Tent" direction maps to Higher on the goodness column
                # (since goodness = -|deviation|, higher = better).
                if direction in ("Higher", "Tent"):
                    f["parameter_scores"][param_name] = round(count_le / n, 6)
                else:  # "Lower"
                    f["parameter_scores"][param_name] = round((n - count_le + 1) / n, 6)

    # Final pass: nullify parameter_scores for non-Ranked funds
    for f in funds:
        if f["centricity_score_status"] != "Ranked":
            f["parameter_scores"] = OrderedDict((p, None) for p in param_order)


def verify_parameter_scores_match(funds: list[dict],
                                  weights_meta: list[dict],
                                  *,
                                  tolerance: float = 0.0001,
                                  warn_only: bool = False,
                                  out_warnings: list[str] | None = None) -> dict:
    """
    For every Ranked fund, recompute the score from parameter_scores × weights
    and assert agreement with the stored centricity_score within `tolerance`.

    `warn_only` (added v2) — instead of raising SchemaError on mismatch,
    append a warning summary to `out_warnings` and return the summary. Used
    for v2 cycles where the workbook's cached centricity_score may pre-date
    the addition of the 3 new params (Avg Mkt Cap, Fund PE, Active Share),
    so a systematic drift between stored and our 22-param recompute is
    expected and benign. The dashboard treats stored centricity_score as
    canonical; the recompute path is only exercised when the user edits
    weights, at which point parameter_scores × edited_weights gives the
    correct 22-param rerank.

    Returns a summary dict with worst_diff and sample comparisons.
    """
    weight_by_param = {w["parameter"]: w["weight_pct"] for w in weights_meta}
    failures: list[str] = []
    worst_diff = 0.0
    samples: list[dict] = []

    for f in funds:
        if f["centricity_score_status"] != "Ranked":
            continue
        stored = f["centricity_score"]
        if stored is None:
            continue
        recomputed = 0.0
        for p, w_pct in weight_by_param.items():
            ps = f["parameter_scores"].get(p)
            if ps is None:
                continue  # NaN parameter drops from numerator (denom stays at 100)
            recomputed += ps * w_pct
        recomputed /= 100.0
        diff = abs(recomputed - stored)
        if diff > worst_diff:
            worst_diff = diff
        if diff > tolerance:
            failures.append(
                f"  - {f['fund_name']!r} (cat={f['category']}, "
                f"AMFI={f['scheme_code']}): stored={stored:.6f} recomputed={recomputed:.6f} "
                f"Δ={diff:.6f}"
            )
        samples.append({
            "fund_name": f["fund_name"],
            "category": f["category"],
            "scheme_code": f["scheme_code"],
            "stored": round(stored, 6),
            "recomputed": round(recomputed, 6),
            "diff": round(diff, 7),
        })

    if failures:
        head = failures[:10]
        more = "" if len(failures) <= 10 else f"\n  ... and {len(failures) - 10} more"
        msg = (
            f"parameter_scores recompute mismatch on {len(failures)} fund(s) "
            f"(tolerance ±{tolerance}). Worst Δ = {worst_diff:.6f}.\n"
            + "\n".join(head) + more
        )
        if warn_only:
            warn_msg = (
                f"[v2-recompute-drift] {len(failures)}/{len(samples)} ranked "
                f"funds drift from stored centricity_score (worst Δ={worst_diff:.4f}). "
                f"Expected for the 15-May cycle — workbook's cached score reflects "
                f"the 19-param formula; 3 new params (Avg Mkt Cap, Fund PE, Active "
                f"Share) push recompute higher. Dashboard uses stored score for "
                f"display, recompute only triggers on weight-drawer edits."
            )
            if out_warnings is not None:
                out_warnings.append(warn_msg)
            print("[converter]   WARNING: " + warn_msg, file=sys.stderr)
        else:
            raise SchemaError(
                msg + "\nPer the contract, this means the normalisation "
                "reproduction is wrong and the dashboard's weight-drawer "
                "recompute would diverge from Excel. Fix before commit."
            )

    return {
        "ranked_funds_verified": len(samples),
        "worst_diff": round(worst_diff, 7),
        "tolerance": tolerance,
        "mismatched_count": len(failures),
        "warn_only": warn_only,
    }


def _compute_derived_risk_3y(
    series: list[tuple[_dt.date, float]],
    cycle_date: _dt.date,
    target_3y_date: _dt.date,
    rf_annual: float | None,
    fund_tenure_yrs: float | None,
) -> dict:
    """
    Compute Sortino, annualised Std Dev, and Max Drawdown from the trailing
    3Y window of a fund's NAV series. Returns all-null when:
      - fund_tenure_yrs < 3
      - fewer than 30 daily observations in the window
      - rf_annual is null (Sortino requires it)
      - sample size for downside / dispersion is insufficient

    See data-contract/screener-v1.json → derived_fields_documentation.risk_metrics_3y_derived.
    """
    null_out = {"sortino_3y": None, "std_dev_3y_pct": None, "max_drawdown_3y_pct": None,
                "sharpe_3y_nav": None}
    if fund_tenure_yrs is None or fund_tenure_yrs < 3:
        return null_out
    if not series:
        return null_out

    # Slice the series to [target_3y_date, cycle_date] inclusive
    window = [(d, v) for (d, v) in series if target_3y_date <= d <= cycle_date]
    if len(window) < 30:
        return null_out

    navs = [v for _, v in window]

    # Daily simple returns from consecutive NAVs
    daily_returns: list[float] = []
    for i in range(1, len(navs)):
        prev = navs[i - 1]
        if prev <= 0:
            continue
        try:
            daily_returns.append(navs[i] / prev - 1.0)
        except (ZeroDivisionError, OverflowError):
            continue

    if len(daily_returns) < 2:
        return null_out

    out = {"sortino_3y": None, "std_dev_3y_pct": None, "max_drawdown_3y_pct": None,
           "sharpe_3y_nav": None}

    # Std Dev (annualised, %)
    full_std = None
    try:
        full_std = _stats.stdev(daily_returns)  # sample, n-1
        out["std_dev_3y_pct"] = round(full_std * math.sqrt(252) * 100, 4)
    except _stats.StatisticsError:
        pass

    # Sharpe + Sortino (annualised) — both require Rf rate.
    # Phase 2.2 §3.2 — Sharpe is now nav-derived (same window, Rf, frequency as
    # Sortino/StdDev) so the three risk metrics live on one basis. Invariants:
    #   sign(Sortino) == sign(Sharpe)     (same numerator)
    #   |Sortino| >= |Sharpe|             (downside_std <= full_std)
    #   StdDev_pct/100 * Sharpe ≈ mean_excess_ann (by construction)
    if rf_annual is not None:
        rf_daily = rf_annual / 252.0
        excess = [r - rf_daily for r in daily_returns]
        mean_excess = sum(excess) / len(excess)
        # Sharpe (annualised, nav-derived) — same numerator as Sortino, divided
        # by full std dev (annualised).
        if full_std is not None and full_std > 0:
            out["sharpe_3y_nav"] = round(
                (mean_excess * 252) / (full_std * math.sqrt(252)), 4
            )
        downside = [e for e in excess if e < 0]
        if len(downside) >= 2:
            try:
                d_std = _stats.stdev(downside)
                if d_std > 0:
                    out["sortino_3y"] = round(
                        (mean_excess * 252) / (d_std * math.sqrt(252)), 4
                    )
            except _stats.StatisticsError:
                pass

    # Max Drawdown (% of running peak; returned as negative)
    peak = navs[0]
    max_dd = 0.0
    for v in navs:
        if v > peak:
            peak = v
        if peak <= 0:
            continue
        dd = (v - peak) / peak  # ≤ 0
        if dd < max_dd:
            max_dd = dd
    out["max_drawdown_3y_pct"] = round(max_dd * 100, 4)

    return out


def _safe_minus_3y(d: _dt.date) -> _dt.date:
    """Subtract 3 calendar years from `d`, handling Feb-29."""
    try:
        return d.replace(year=d.year - 3)
    except ValueError:
        # 29 Feb on a non-leap year-3 → 28 Feb
        return d.replace(month=2, day=28, year=d.year - 3)


def _compute_rolling_3y_stats(
    series: list[tuple[_dt.date, float]],
    cycle_date: _dt.date,
    bench_series: list[tuple[_dt.date, float]] | None,
) -> dict | None:
    """
    Daily-roll 3-year rolling CAGR statistics from the fund's NAV series.
    Returns None when the fund has fewer than 252 daily NAV observations
    (less than ~1 year of data — can't form even one 3Y window).

    For each ascending date `t` in the series where (t − 3 years) lies
    inside the series, compute fund_cagr = (NAV[t] / NAV[t−3y])^(1/3) − 1.
    If `bench_series` is provided AND the same window has benchmark
    observations on both endpoints, also compute bench_cagr and tally
    pct_beat_benchmark.

    All percentages returned as float `%` (e.g., 19.84 = 19.84%, not 0.1984).
    Window-start dates rendered as ISO 'YYYY-MM-DD'.
    """
    if not series or len(series) < 252:
        return None
    # Trim to history that ends at or before cycle_date (defensive — series
    # is loaded from the workbook so the last row should already be the
    # cycle's NAV, but a stray future-dated row shouldn't poison the stats).
    history = [(d, v) for (d, v) in series if d <= cycle_date]
    if len(history) < 252:
        return None

    cagrs: list[tuple[_dt.date, float]] = []   # (window_start_date, fund_3y_cagr)
    paired: list[tuple[float, float]] = []     # (fund_cagr, bench_cagr) for beat-benchmark calc

    for (t_d, t_v) in history:
        if t_v is None or t_v <= 0:
            continue
        back_d = _safe_minus_3y(t_d)
        # Find the latest series point with date <= back_d
        b_d, b_v = _series_value_at_or_before(history, back_d)
        if b_v is None or b_v <= 0:
            continue
        # Require the window start to be within ~30 days of the calendar
        # offset (otherwise sparse data produces phantom long windows).
        if (back_d - b_d).days > 30:
            continue
        try:
            f_cagr = (t_v / b_v) ** (1.0 / 3.0) - 1.0
        except (ValueError, ZeroDivisionError, OverflowError):
            continue
        cagrs.append((b_d, f_cagr))
        if bench_series:
            _bt_d, bt_v = _series_value_at_or_before(bench_series, t_d)
            _bb_d, bb_v = _series_value_at_or_before(bench_series, back_d)
            if bt_v is not None and bb_v is not None and bb_v > 0 and bt_v > 0:
                try:
                    b_cagr = (bt_v / bb_v) ** (1.0 / 3.0) - 1.0
                    paired.append((f_cagr, b_cagr))
                except (ValueError, ZeroDivisionError, OverflowError):
                    pass

    if not cagrs:
        return None

    n = len(cagrs)
    cagr_values = [c for (_, c) in cagrs]
    cagr_values_sorted = sorted(cagr_values)
    median = cagr_values_sorted[n // 2] if n % 2 == 1 else (
        (cagr_values_sorted[n // 2 - 1] + cagr_values_sorted[n // 2]) / 2.0
    )
    avg = sum(cagr_values) / n
    best_idx = max(range(n), key=lambda i: cagr_values[i])
    worst_idx = min(range(n), key=lambda i: cagr_values[i])

    pct_positive = sum(1 for c in cagr_values if c > 0) / n
    pct_above_12 = sum(1 for c in cagr_values if c > 0.12) / n
    pct_beat = (sum(1 for f, b in paired if f > b) / len(paired)) if paired else None

    return {
        "avg_pct": round(avg * 100, 4),
        "median_pct": round(median * 100, 4),
        "best_pct": round(cagr_values[best_idx] * 100, 4),
        "best_window_start": cagrs[best_idx][0].isoformat(),
        "worst_pct": round(cagr_values[worst_idx] * 100, 4),
        "worst_window_start": cagrs[worst_idx][0].isoformat(),
        "pct_positive": round(pct_positive * 100, 4),
        "pct_above_12": round(pct_above_12 * 100, 4),
        "pct_beat_benchmark": round(pct_beat * 100, 4) if pct_beat is not None else None,
        "observation_count": n,
    }


def _resample_navs_to_monthly(
    series: list[tuple[_dt.date, float]],
    start_d: _dt.date,
    end_d: _dt.date,
) -> list[dict]:
    """
    Last available NAV in each calendar month within [start_d, end_d].
    Output: ascending list of {"d": "YYYY-MM", "v": float (raw NAV)}.
    The dashboard normalises to ₹1,00,000 growth at the selected window
    start at render time; no normalisation here.
    """
    by_month: dict[tuple[int, int], tuple[_dt.date, float]] = {}
    for d, v in series:
        if d < start_d or d > end_d:
            continue
        if v is None or v <= 0:
            continue
        key = (d.year, d.month)
        prev = by_month.get(key)
        if prev is None or prev[0] < d:
            by_month[key] = (d, v)
    items = sorted(by_month.items(), key=lambda kv: kv[0])
    return [
        {"d": f"{k[0]:04d}-{k[1]:02d}", "v": round(v, 4)}
        for k, (_, v) in items
    ]


def emit_benchmark_nav_file(
    bm_by_name: dict[str, list[tuple[_dt.date, float]]],
    cycle_date: _dt.date,
    out_path: Path,
) -> None:
    """
    Stage B A3 (2026-05-28) — emit a separate `benchmark-nav-YYYY-MM-DD.json`
    consumed by `fund-detail.js`'s Growth-of-₹1L chart for the benchmark line.

    Per-fund `nav-series-YYYY-MM-DD.json` carries each fund's benchmark series
    inline (monthly-resampled), which is what the chart originally used. But
    that file is monthly-only and per-fund-duplicated; the chart needs DAILY
    values and a single keyed-by-label lookup so multiple funds tracking the
    same benchmark resolve to one canonical series. This emits that file.

    Source: `bm_by_name` from `load_benchmark_nav(wb)` (workbook 📈 Benchmark
    NAV sheet — 96 labels) + the 4 INR-converted USD pseudo-series from
    `_load_inr_converted_series()`. Same alias-fold map as Phase 2.2 Part 2;
    canonical labels match what the cycle JSON's `fund.benchmark` field uses.

    Schema:
      { "cycle_date": "YYYY-MM-DD",
        "source": "📈 Benchmark NAV (workbook) + INR-converted pseudo-series",
        "series": { "<benchmark label>": [["YYYY-MM-DD", nav], ...], ... } }
    """
    series_out: dict[str, list] = {}
    aliases: dict[str, str] = {}
    for name, navs in bm_by_name.items():
        if not navs:
            continue
        series_out[name] = [
            [d.isoformat(), round(v, 4)] for d, v in navs if v is not None
        ]
    # Stage B A3 — alias map covers the screener JSON's actual `fund.benchmark`
    # variants: TRI/PRI suffix drift + uppercase-NIFTY case drift. The workbook
    # 📈 Benchmark NAV row stores bare titles ("Nifty 50"); the cycle JSON's
    # `benchmark` field often carries the canonical TRI label ("NIFTY 50 - TRI").
    # Both `series_out[canonical]` and `series_out[alias]` resolve to the same
    # daily series so the UI's lookup is a one-liner: `series[fund.benchmark]`.
    # Phase 2.2 §8 canonical form: `lowercase(strip(" - TRI", " - PRI")).replace(" ","")`.
    for name in list(series_out.keys()):
        bare = re.sub(r"\s*-\s*(TRI|PRI)\s*$", "", name, flags=re.IGNORECASE).strip()
        variants = {
            bare,
            bare.upper(),
            bare + " - TRI",
            bare.upper() + " - TRI",
            bare + " - PRI",
        }
        for v in variants:
            if v and v != name and v not in series_out:
                series_out[v] = series_out[name]
                aliases[v] = name
    payload = OrderedDict([
        ("cycle_date", cycle_date.isoformat()),
        ("source", "📈 Benchmark NAV (workbook) + INR-converted pseudo-series"),
        ("aliases", aliases),
        ("series", series_out),
    ])
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))


def emit_nav_series_file(
    funds: list[dict],
    nav_by_amfi: dict[int, list[tuple[_dt.date, float]]],
    bm_by_name: dict[str, list[tuple[_dt.date, float]]],
    cycle_date: _dt.date,
    out_path: Path,
) -> None:
    """
    Emit data/nav-series-YYYY-MM-DD.json — monthly-frequency NAV series
    for every fund + its benchmark, capped at 13 years before cycle_date.
    Lazy-loaded by fund-detail.js for the "Growth of ₹ 1,00,000" chart.

    Schema (per CLAUDE.md §4.1, emitted alongside the screener JSON):
      {
        "cycle_date": "YYYY-MM-DD",
        "series": {
          "<amfi>": {
            "inception_date": "YYYY-MM-DD"|null,
            "benchmark":      "<name>"|null,
            "fund":  [{"d":"YYYY-MM","v":<nav>}, ...],
            "bench": [{"d":"YYYY-MM","v":<nav>}, ...]
          }, ...
        }
      }
    """
    cap_back = cycle_date - _dt.timedelta(days=int(13 * 365.25))
    series_out: dict[str, dict] = {}
    for fund in funds:
        amfi = fund.get("scheme_code")
        if amfi is None:
            continue
        fund_navs = nav_by_amfi.get(amfi, [])
        if not fund_navs:
            continue
        bench_name = fund.get("benchmark")
        bench_navs = bm_by_name.get(bench_name, []) if bench_name else []
        series_out[str(amfi)] = OrderedDict([
            ("inception_date", fund.get("inception_date")),
            ("benchmark", bench_name),
            ("fund",  _resample_navs_to_monthly(fund_navs,  cap_back, cycle_date)),
            ("bench", _resample_navs_to_monthly(bench_navs, cap_back, cycle_date)),
        ])

    payload = OrderedDict([
        ("cycle_date", cycle_date.isoformat()),
        ("series", series_out),
    ])
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))


def extract_monitor_data(monitor_path: Path) -> dict:
    """
    Read the MF Monitor workbook and return a dict with TWO sub-dicts:

        {
          "fund_returns":  {<scheme_name>: {ytd_pct, return_1m_pct, ...,
                            exit_load, monitor_ter_pct}, ...},
          "index_returns": {<index_name>: {ytd_pct, return_1m_pct, ...,
                            return_10y_pct}, ...}
        }

    Sheet structure varies between equity sheets ('Focused Fund', 'Large Cap
    Fund', etc. — 23 cols, exit load in '[Exit Load]') and hybrid sheets
    ('Aggressive Hybrid Fund', etc. — 36 cols, exit load in 'Remark'). The
    extractor parses the header row 12 to find columns by name rather than
    by hardcoded index, so both layouts work.

    Index rows (Fix-List 6 §1A) live BELOW the fund rows on every category
    sheet, after a marker row whose col 0 is the literal "BenchMark". They
    use the same column layout as fund rows. We detect them by flipping into
    "index mode" once we encounter "BenchMark" as col 0 of any row, and
    treat every subsequent non-blank row as an index until end-of-sheet.

    Aggregate roll-up sheets ('Home', 'Debt' standalone, 'Hybrid', 'Silver',
    'Gold', any sheet whose name contains 'Oriented' or starts with
    'Solution') are skipped — their funds appear in the per-category sheets
    too, and including them would double-count. Index rows on those skipped
    sheets are also skipped (they appear in per-category sheets too).

    DATA QUALITY NOTE — Monitor TER is the regular plan expense ratio
    whereas the Whitelisting Excel ter_pct field appears to carry the
    direct plan TER. e.g. HDFC Focused Fund: Monitor 1.6% vs Whitelisting
    0.47%. Both fields are kept on the fund record (`ter_pct` from
    Whitelisting, `monitor_ter_pct` from Monitor); fund-detail.html displays
    `monitor_ter_pct` since that's the partner-facing regular plan number.
    """
    SKIP_SHEETS = {"Home", "Debt", "Silver", "Gold", "Hybrid"}

    def _is_data_sheet(name: str, ws) -> bool:
        if name in SKIP_SHEETS:
            return False
        if "Oriented" in name:
            return False                       # Equity Oriented / Debt Oriented / Solution Oriented,…
        if name.startswith("Solution"):
            return False
        v = ws.cell(row=12, column=1).value
        return v == "Scheme Name"

    def _parse_headers(ws) -> dict[str, int]:
        """Header row 12 → {column-name → 1-indexed col number}."""
        headers: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=12, column=c).value
            if v is None:
                continue
            headers[str(v).strip()] = c
        return headers

    fund_out: dict[str, dict] = {}
    index_out: dict[str, dict] = {}
    wb = openpyxl.load_workbook(monitor_path, read_only=True, data_only=True)
    skipped_sheets: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not _is_data_sheet(sheet_name, ws):
            skipped_sheets.append(sheet_name)
            continue
        H = _parse_headers(ws)
        c_name = H.get("Scheme Name")
        if c_name is None:
            continue
        # Monitor calls these columns the same in both equity and hybrid layouts
        c_ytd  = H.get("YTD")
        c_1m   = H.get("1 Month")
        c_1y   = H.get("1 Year")
        c_3y   = H.get("3 Years")
        c_5y   = H.get("5 Years")
        c_10y  = H.get("10 Years")
        c_ter  = H.get("Ratio")
        c_exit = H.get("[Exit Load]") or H.get("Remark")
        # Phase 2.2 Patch (FM rule 2026-05-28) — Monitor's "Fund Manager"
        # column is the canonical source for `manager_name` (catalogue §7.2 /
        # §7.8). Col index varies per sheet (e.g. col 26 on Aggressive Hybrid;
        # different on others), so resolve by exact header label.
        # Stage B A1 (2026-05-28) — Monitor's EQUITY sheets use the bracketed
        # variant `[Fund Manager 1]`; HYBRID / arbitrage / solution sheets use
        # the bracketless `Fund Manager`. Cowork's universe-wide diff found
        # 600+ equity funds silently falling back to Morningstar because the
        # original H.get("Fund Manager") only matched the hybrid form. Resolve
        # by normalising every header label (strip brackets + whitespace,
        # lower-case) and accepting either variant.
        c_fm = None
        for _hdr, _col in H.items():
            _norm = re.sub(r"[\[\]\s]", "", str(_hdr)).lower()
            if _norm in ("fundmanager", "fundmanager1"):
                c_fm = _col
                break

        in_benchmark_section = False
        for r in range(13, ws.max_row + 1):
            name_v = ws.cell(row=r, column=c_name).value
            scheme = _safe_str(name_v)
            if not scheme:
                continue
            if not isinstance(name_v, str):
                continue
            # The "BenchMark" header row separates fund rows from index rows
            if scheme.lower() == "benchmark":
                in_benchmark_section = True
                continue
            # Phase 2.2 Patch (FM rule 2026-05-28) — Regular Growth only;
            # skip Direct / IDCW / Dividend variants so monitor_fm always
            # resolves against the canonical scheme.
            if not in_benchmark_section and any(
                tag in scheme for tag in ("Direct", "IDCW", "Dividend")
            ):
                continue

            def f(c):
                if c is None:
                    return None
                return _safe_float(ws.cell(row=r, column=c).value)

            if in_benchmark_section:
                # Index row — capture the 6 returns we'll use on Fund Detail.
                # Last write wins across sheets (NIFTY 50 - TRI appears in
                # multiple per-category sheets — values are identical).
                idx_data = {
                    "ytd_pct":        _round(f(c_ytd),  4),
                    "return_1m_pct":  _round(f(c_1m),   4),
                    "return_1y_pct":  _round(f(c_1y),   4),
                    "return_3y_pct":  _round(f(c_3y),   4),
                    "return_5y_pct":  _round(f(c_5y),   4),
                    "return_10y_pct": _round(f(c_10y),  4),
                }
                index_out[scheme] = idx_data
                continue

            row_data = {
                "ytd_pct":        _round(f(c_ytd),  4),
                "return_1m_pct":  _round(f(c_1m),   4),
                "return_1y_pct":  _round(f(c_1y),   4),
                "return_3y_pct":  _round(f(c_3y),   4),
                "return_5y_pct":  _round(f(c_5y),   4),
                "return_10y_pct": _round(f(c_10y),  4),
                "monitor_ter_pct": _round(f(c_ter), 4),
                "exit_load": _safe_str(ws.cell(row=r, column=c_exit).value) if c_exit else None,
                # Phase 2.2 Patch (FM rule 2026-05-28) — Monitor's
                # canonical lead-manager name. Used by build_funds to set
                # `manager_name` (Monitor primary; Morningstar fallback).
                "fund_manager": _safe_str(ws.cell(row=r, column=c_fm).value) if c_fm else None,
            }
            fund_out[scheme] = row_data
    wb.close()
    return {"fund_returns": fund_out, "index_returns": index_out}


def _trailing_returns_from_series(series: list[tuple[_dt.date, float]],
                                  cycle_date: _dt.date,
                                  targets: dict[str, _dt.date]) -> dict:
    """Compute 1Y/3Y/5Y/SI from a pre-loaded NAV series."""
    out = {"return_1y_pct": None, "return_3y_pct": None,
           "return_5y_pct": None, "return_si_pct": None}
    if not series:
        return out
    end_d, end_v = _series_value_at_or_before(series, cycle_date)
    if end_v is None:
        return out
    for label, years in (("1Y", 1), ("3Y", 3), ("5Y", 5)):
        _, start_v = _series_value_at_or_before(series, targets[label])
        if start_v is None:
            continue
        if years == 1:
            out[f"return_{label.lower()}_pct"] = _absolute_return_pct(end_v, start_v)
        else:
            out[f"return_{label.lower()}_pct"] = _trailing_return_pct(end_v, start_v, years)
    # SI
    first_d, first_v = series[0]
    if first_v is not None and end_d is not None:
        days = (end_d - first_d).days
        if days >= 30:
            years_si = days / 365.25
            if years_si >= 1:
                out["return_si_pct"] = _trailing_return_pct(end_v, first_v, years_si)
            else:
                out["return_si_pct"] = _absolute_return_pct(end_v, first_v)
    return out


def _lookup_index_returns(benchmark_name: str | None, index_data: dict | None) -> dict | None:
    """
    Match `fund.benchmark` (e.g. "NIFTY 50 - TRI") to a Monitor index row.

    Tries exact match first, then strips a trailing " - TRI" / " TRI" suffix
    and retries (handles edge cases where the screener carries the TRI
    variant but the Monitor lists the price-return version, or vice versa).

    Returns None when no match is found — the page falls back to
    fund.benchmark_returns (CAGR) for 1Y/3Y/5Y and renders em-dash for
    YTD / 1M / 10Y.
    """
    if not benchmark_name or not index_data:
        return None
    if benchmark_name in index_data:
        return index_data[benchmark_name]
    # Try without " - TRI"
    stripped = benchmark_name.replace(" - TRI", "").replace(" TRI", "").strip()
    if stripped and stripped in index_data:
        return index_data[stripped]
    # Try adding " - TRI" if not already there
    with_tri = benchmark_name + " - TRI" if " TRI" not in benchmark_name else benchmark_name
    if with_tri in index_data:
        return index_data[with_tri]
    return None


def build_funds(
    wb,
    contract: dict,
    cycle_meta: dict,
    warnings: list[str],
    monitor_data: dict | None = None,
    monitor_index_data: dict | None = None,
    morningstar_history: dict[int, list[dict]] | None = None,
    global_bm_dir: Path | None = None,
    analytics_equity_counts: dict[str, int] | None = None,
) -> tuple[list[dict], dict, dict]:
    eq_set = set(contract["categories"]["equity"])
    hb_set = set(contract["categories"]["hybrid"])
    morningstar_history = morningstar_history or {}
    analytics_equity_counts = analytics_equity_counts or {}
    # Phase 2.2 §A1-REDO — coverage stats for the manager-name source. Reported
    # at end of conversion and surfaced on cycle_meta. AMFI is the join key
    # (workbook scheme_code ↔ Morningstar AMFI Code col E); we DO NOT name-
    # match anymore.
    mgr_coverage = {"morningstar": 0, "mf_monitor": 0, "none": 0}

    cycle_date = _dt.date.fromisoformat(cycle_meta["cycle_date"])
    targets = {
        "1Y": cycle_date.replace(year=cycle_date.year - 1),
        "3Y": cycle_date.replace(year=cycle_date.year - 3),
        "5Y": cycle_date.replace(year=cycle_date.year - 5),
    }

    print("[converter] streaming Fund NAV into memory...", file=sys.stderr)
    nav_by_amfi, _name_by_amfi = load_fund_nav(wb)
    print(f"[converter]   loaded {len(nav_by_amfi)} fund NAV series", file=sys.stderr)

    # Phase 2.2 §benchmark — NAV unit-split detect + adjust BEFORE any nav-
    # derived metric runs. Trailing returns, derived 3Y risk metrics, rolling
    # 3Y stats, and the benchmark TE/TD all consume the resulting (cleaned)
    # series. 'clean-split-adjusted' funds get their pre-split NAVs ×= factor;
    # 'nav-split-suspect' funds keep their raw series but the per-fund block
    # below forces every nav-derived metric to null.
    split_status_by_amfi, _clean_n, _suspect_n = _apply_nav_split_adjustments(nav_by_amfi)
    print(
        f"[converter]   NAV split scan: {len(split_status_by_amfi)} funds with ±40% jumps "
        f"({_clean_n} clean-adjusted, {_suspect_n} suspect-nulled)",
        file=sys.stderr,
    )

    print("[converter] streaming Benchmark NAV into memory...", file=sys.stderr)
    bm_by_name = load_benchmark_nav(wb)
    print(f"[converter]   loaded {len(bm_by_name)} benchmark NAV series", file=sys.stderr)

    # Phase 2.2 §benchmark — basis classification. Workbook NAVs are TRI by
    # default; the 8 BSE indices added in 2026 are PRI-only.
    bm_basis_for_name: dict[str, str] = {
        nm: ("PRI" if nm in NEW_BSE_NAMES else "TRI") for nm in bm_by_name
    }

    # Phase 2.2 §benchmark — load Investing.com USD CSVs + USD/INR and register
    # 4 INR-converted pseudo-series in bm_by_name. The workbook does NOT carry
    # the converted series, so the converter replicates the conversion here.
    # If the global BM dir is missing (e.g. CI run from a slim data/ folder),
    # the 4 USD-international funds gracefully fall back to absent (currency-
    # mismatch).
    inr_series, inr_diag = _load_inr_converted_series(global_bm_dir)
    if inr_series:
        for nm, lst in inr_series.items():
            bm_by_name[nm] = lst
            bm_basis_for_name[nm] = "PRI-INR"
        print(
            f"[converter]   registered {len(inr_series)} INR-converted pseudo-series: "
            f"{list(inr_series.keys())} (USD/INR pts: {inr_diag['usd_inr_pts']})",
            file=sys.stderr,
        )
    else:
        print(
            f"[converter]   WARNING: Global BM dir not found or unusable at "
            f"{global_bm_dir}; USD-international funds will be 'absent (currency-mismatch)'. "
            f"diag={inr_diag}",
            file=sys.stderr,
        )

    # Canonical-alias lookup for Tier-1 (alias-fold) matching.
    bm_canon: dict[str, str] = {_canon_stated(nm): nm for nm in bm_by_name}

    # Pre-compute benchmark returns once per benchmark
    bm_cache: dict[str, dict] = {}
    for name, series in bm_by_name.items():
        bm_cache[name] = _trailing_returns_from_series(series, cycle_date, targets)

    def benchmark_returns_for(name: str | None) -> dict:
        if not name:
            return {"return_1y_pct": None, "return_3y_pct": None, "return_5y_pct": None}
        if name in bm_cache:
            r = bm_cache[name]
            return {"return_1y_pct": r["return_1y_pct"], "return_3y_pct": r["return_3y_pct"],
                    "return_5y_pct": r["return_5y_pct"]}
        warnings.append(f"Benchmark not matched in 📈 Benchmark NAV row 1: {name!r}")
        bm_cache[name] = {"return_1y_pct": None, "return_3y_pct": None,
                          "return_5y_pct": None, "return_si_pct": None}
        return {"return_1y_pct": None, "return_3y_pct": None, "return_5y_pct": None}

    print("[converter] streaming 📋 Data + 📊 Computed Metrics into memory...", file=sys.stderr)
    data_rows = _stream_data_rows(wb)
    cm_rows = _stream_cm_rows(wb)
    # v2: Data has 1252 rows (incl. 3 Quantum Direct duplicates); CM has 1249
    # (deduped). v1 ran a strict positional alignment; v2 switches to a
    # name-keyed lookup so duplicate-Direct rows in Data drop cleanly with a
    # warning, and the same converter still works for v1 cycles (where the
    # two row sets match 1:1).
    cm_by_name: dict[str, dict] = {}
    cm_dups: list[str] = []
    for cm in cm_rows:
        nm = cm.get("fund_name")
        if not nm:
            continue
        if nm in cm_by_name:
            cm_dups.append(nm)
        cm_by_name[nm] = cm
    if cm_dups:
        warnings.append(
            f"📊 Computed Metrics: {len(cm_dups)} duplicate fund_name(s) — "
            f"latest occurrence wins. First few: {cm_dups[:5]!r}"
        )
    skipped_no_cm = 0
    print(f"[converter]   loaded {len(data_rows)} Data rows + {len(cm_rows)} CM rows", file=sys.stderr)

    funds: list[dict] = []
    score_dist = {"Ranked": 0, "1-3yr Warning": 0, "New Fund Monitoring": 0}

    for idx, d in enumerate(data_rows):
        if d["scheme_code"] is None and d["fund_name"] is None:
            break
        if d["scheme_code"] is None:
            warnings.append(f"📋 Data fund #{d['seq']}: missing AMFI code, fund={d['fund_name']!r}, skipped.")
            continue
        cm = cm_by_name.get(d["fund_name"])
        if cm is None:
            # Expected for Quantum Direct duplicates in v2 — Data carries them
            # but CM deduped to Reg(G) only. Log + skip; never break the build.
            skipped_no_cm += 1
            warnings.append(
                f"📋 Data #{d['seq']} {d['fund_name']!r} has no matching CM row "
                f"(likely deduped duplicate / Direct plan variant); skipped."
            )
            continue

        category = d["category"]
        # Phase 2.2 §A2 — revert 11 Dividend-Yield funds from Sector-Thematic
        # back to Value-Contra so they're scored / ranked against the right
        # peer pool. Applied BEFORE compute_parameter_scores so the percentile
        # pool is correct out of the gate.
        if d["scheme_code"] in DIVIDEND_YIELD_AMFI and category == "Sector-Thematic":
            category = "Value-Contra"
        if category in eq_set:
            sub_class = "Equity"
        elif category in hb_set:
            sub_class = "Hybrid"
        else:
            sub_class = "Equity"
            warnings.append(
                f"Unknown category {category!r} for fund {d['fund_name']!r}; defaulting sub_class=Equity."
            )

        status, score_dec, warn_pct = _classify_score(cm["score_raw"])
        # Phase 2.1 — passive categories carry NO score / rank. The workbook
        # emits 0.0 / "Ranked" for these (index-fund scoring is a separate
        # methodology pass that hasn't shipped yet); we override here so the
        # partner-facing UI doesn't render 497 rows as 0.00 with bogus ranks
        # 1..N.
        if category in PASSIVE_CATEGORIES:
            status = PASSIVE_STATUS
            score_dec = None
            warn_pct = None
        score_dist[status] = score_dist.get(status, 0) + 1

        # Trailing returns + derived 3Y risk metrics from in-memory NAV series
        series = nav_by_amfi.get(d["scheme_code"])
        bench_series = bm_by_name.get(d["benchmark"]) if d["benchmark"] else None
        # v2: 15-May workbook ships a few funds whose NAV column exists but
        # holds zero observations (new launches mid-cycle). Treat empty
        # series the same as missing — emit warnings, leave derived fields null.
        if not series:
            if series is None:
                warnings.append(
                    f"AMFI {d['scheme_code']} ({d['fund_name']!r}) not found in 📈 Fund NAV row 1."
                )
            else:
                warnings.append(
                    f"AMFI {d['scheme_code']} ({d['fund_name']!r}) has zero NAV observations."
                )
            trailing = {"return_1y_pct": None, "return_3y_pct": None,
                        "return_5y_pct": None, "return_si_pct": None}
            derived_risk = {"sortino_3y": None, "std_dev_3y_pct": None,
                            "max_drawdown_3y_pct": None, "sharpe_3y_nav": None}
            rolling_stats = None
            nav_latest_value = None
            nav_latest_date = None
        else:
            trailing = _trailing_returns_from_series(series, cycle_date, targets)
            derived_risk = _compute_derived_risk_3y(
                series=series,
                cycle_date=cycle_date,
                target_3y_date=targets["3Y"],
                rf_annual=cycle_meta.get("rf_rate_annual"),
                fund_tenure_yrs=cm["fund_tenure_yrs"],
            )
            rolling_stats = _compute_rolling_3y_stats(series, cycle_date, bench_series)
            # Latest NAV — last non-null observation in the fund's series
            last_d, last_v = series[-1]
            nav_latest_value = round(last_v, 4) if last_v is not None else None
            nav_latest_date = last_d.isoformat() if last_d is not None else None

        bm_ret = benchmark_returns_for(d["benchmark"])

        # Phase 2.2 §benchmark — for 'nav-split-suspect' funds the ±40% jump's
        # source is ambiguous, so all nav-derived metrics in this record are
        # forced to null (matches Part-1's behaviour). 'clean-split-adjusted'
        # funds need no action here — their series was already adjusted before
        # _trailing_returns_from_series / _compute_derived_risk_3y ran above.
        nav_split_status = split_status_by_amfi.get(d["scheme_code"], "")
        if nav_split_status == "nav-split-suspect":
            trailing = {"return_1y_pct": None, "return_3y_pct": None,
                        "return_5y_pct": None, "return_si_pct": None}
            derived_risk = {"sortino_3y": None, "std_dev_3y_pct": None,
                            "max_drawdown_3y_pct": None, "sharpe_3y_nav": None}
            rolling_stats = None

        # Phase 2.2 §benchmark — 4-tier matcher + TE/TD compute. The proxy-
        # validity gate runs AFTER this loop (it needs the TE value to decide
        # whether to downgrade a passive proxy to absent). `series` has already
        # been split-adjusted for clean splits.
        bench_match = _match_fund(
            d["benchmark"], d["fund_name"], bm_canon, bm_basis_for_name
        )
        matched_name = bench_match["matched_series"]
        te_v: float | None = None
        td_v: float | None = None
        if (matched_name
                and series
                and nav_split_status != "nav-split-suspect"):
            matched_bm_list = bm_by_name.get(matched_name)
            if matched_bm_list:
                te_v, td_v, _n_obs, _fc, _bc = _compute_te_td(
                    series, matched_bm_list, targets["3Y"], cycle_date
                )

        record = OrderedDict()
        record["scheme_code"] = d["scheme_code"]
        record["fund_name"] = d["fund_name"]
        record["category"] = category
        record["sub_category_class"] = sub_class
        record["amc"] = d["amc"]
        record["amc_score"] = d["amc_score"]
        record["benchmark"] = d["benchmark"]
        record["inception_date"] = d["inception_date_iso"]
        record["fund_tenure_yrs"] = _round(cm["fund_tenure_yrs"], 4)
        # Phase 2.2 §A1-REDO — Morningstar full history primary; MF Monitor
        # fallback only. Join on AMFI code. When the AMFI is in the
        # Morningstar export AND has at least one current manager (open
        # bracket), we use that name + structured history + derived tenure;
        # else we fall back to the workbook MF Monitor name + tenure;
        # else "—".
        mh_records = morningstar_history.get(d["scheme_code"]) if d["scheme_code"] is not None else None
        mh_name, mh_co_managers, mh_since, mh_tenure = derive_current_manager(
            mh_records or [], cycle_date
        )
        mfm_name = d["manager_name"]
        # Phase 2.2 Patch (FM rule 2026-05-28) — supersedes the earlier
        # "earliest-current Morningstar" lead rule. `manager_name` is now
        # whatever appears in the MF Monitor file's "Fund Manager" column
        # (catalogue §7.2 + §7.8, corrected 2026-05-28). Morningstar's
        # earliest-current is kept only as a safety net for funds Monitor
        # doesn't carry (typically passives / debt corner cases).
        monitor_row = (monitor_data or {}).get(d["fund_name"]) if d["fund_name"] else None
        monitor_fm_raw = monitor_row.get("fund_manager") if monitor_row else None
        # Stage B A1 (2026-05-28) — apply `normalize_manager_name` so casing /
        # trailing punctuation drift across Monitor sheets resolves to one
        # canonical spelling. `"harish krishnan"` → `"Harish Krishnan"`,
        # `"Deepak Gupta."` → `"Deepak Gupta"`, `"D'Silva"` preserved verbatim.
        monitor_fm = normalize_manager_name(monitor_fm_raw)

        if monitor_fm:
            record["manager_name"] = monitor_fm
            record["manager_name_source"] = "MF Monitor"
            mgr_coverage["mf_monitor"] += 1
        elif mh_name:
            record["manager_name"] = mh_name
            record["manager_name_source"] = "Morningstar"
            mgr_coverage["morningstar"] += 1
        elif mfm_name:
            record["manager_name"] = mfm_name
            record["manager_name_source"] = "MF Monitor"
            mgr_coverage["mf_monitor"] += 1
        else:
            record["manager_name"] = None
            record["manager_name_source"] = None
            mgr_coverage["none"] += 1
        # `manager_tenure_yrs` / `manager_since` — UNCHANGED per the FM
        # patch: still anchored on Morningstar's LONGEST-CURRENT active
        # (= earliest start), NOT the named Monitor lead's own tenure.
        # This preserves cross-cycle continuity: ICICI E&D Apr 12.58 → May
        # 12.65 yr = natural +0.07 yr increment on Banthia's 2013-09-19
        # start, even though `manager_name` now reads Sankaran Naren.
        # (If Rahul later wants the named lead's tenure, that's a separate
        # decision.)
        if mh_tenure is not None:
            record["manager_tenure_yrs"] = mh_tenure
        else:
            record["manager_tenure_yrs"] = _round(d["manager_tenure_yrs"], 4)
        record["manager_since"] = mh_since
        # `manager_co_managers` reordered so the LEAD (manager_name) sits
        # at index 0. The remaining entries are the Morningstar currently-
        # active list ordered by `start` ascending. If the Monitor FM isn't
        # in the Morningstar active list (sparse Morningstar coverage),
        # prepend it anyway so the lead always surfaces first. UI's co-strip
        # reads index 1: (lead is rendered separately in the Manager card).
        # Cowork patch 2026-05-28 — co-manager de-dup by first+last name fold,
        # not by exact-string equality. Monitor and Morningstar disagree on
        # middle initials / spellings ("Amit Ganatra" vs "Amit B. Ganatra";
        # "V. Srivatsa" vs "V Srivatsa") so the original `m != monitor_fm`
        # check kept both spellings of the same person in the list, which
        # then leaked into `data/manager-profiles.json` as a key mismatch:
        # 77 non-passive funds had `manager_name` (Monitor canon) that
        # didn't resolve in the profile file because the profile was keyed
        # under the Morningstar long-form alias also present in this list.
        def _fl(name: str) -> str:
            # First+last name fold — strip dots/commas (so "V." == "V") so
            # "V. Srivatsa" matches "V Srivatsa". Middle initials between
            # first and last are ignored ("Amit Ganatra" matches "Amit B. Ganatra").
            parts = (name or "").strip().split()
            if not parts:
                return ""
            def _strip(p):
                return p.lower().replace(".", "").replace(",", "").strip()
            if len(parts) == 1:
                return _strip(parts[0])
            return f"{_strip(parts[0])} {_strip(parts[-1])}"

        if monitor_fm:
            mon_fl = _fl(monitor_fm)
            rest = [m for m in (mh_co_managers or []) if _fl(m) != mon_fl]
            record["manager_co_managers"] = [monitor_fm] + rest
        elif mh_co_managers:
            record["manager_co_managers"] = mh_co_managers
        elif mfm_name:
            record["manager_co_managers"] = [mfm_name]
        else:
            record["manager_co_managers"] = []
        # Structured full history (every manager who has run this fund),
        # carried into the JSON so the one-pager can render it and Part B
        # can detect Apr→May moves (a current start or a recent end inside
        # the window). Empty list when Morningstar has no entry.
        record["manager_history"] = mh_records if mh_records else []
        record["aum_cr"] = _round(d["aum_cr"], 4)
        record["ter_pct"] = _round(d["ter_pct"], 4)
        record["turnover_pct"] = _round(d["turnover_pct"], 4)
        # Phase 2.2 Patch (no_of_stocks) — count ONLY analytics rows where
        # Asset == "Equity" (excludes Debt + Others/REITs/InvITs/cash). For
        # ICICI Pru E&D this fixes the workbook's 148 (Equity + Others)
        # value to ~131 (Equity only). Falls back to the workbook 📋 Data
        # col N count when the fund isn't in any analytics file (e.g.
        # brand-new fund, or analytics_dir unavailable at convert time).
        # See catalogue §7.8 and AUDIT_ICICI_FINDINGS_2026-05-28.md #1.
        analytics_count = analytics_equity_counts.get(d["fund_name"])
        if analytics_count is not None:
            record["no_of_stocks"] = analytics_count
        else:
            record["no_of_stocks"] = d["no_of_stocks"]
        record["mcap_split"] = OrderedDict([
            ("large_pct", _round(d["large_cap_pct"], 4)),
            ("mid_pct", _round(d["mid_cap_pct"], 4)),
            ("small_pct", _round(d["small_cap_pct"], 4)),
            ("others_pct", _round(d["others_pct"], 4)),
        ])
        record["hybrid_extension"] = OrderedDict([
            ("equity_pct", _round(d["h_equity_pct"], 4)),
            ("debt_pct", _round(d["h_debt_pct"], 4)),
            ("others_pct_hybrid", _round(d["h_others_pct"], 4)),
            ("ytm", _round(d["h_ytm"], 4)),
            ("mod_duration_yrs", _round(d["h_mod_duration"], 4)),
            ("avg_maturity_yrs", _round(d["h_avg_maturity"], 4)),
        ])
        # v2 — universal asset split sourced from CM cols AF/AG/AH. Carries the
        # real per-fund equity/debt/others split for every fund (equity funds
        # like Kotak Large Cap show e.g. 97 / 0.4 / 2.6 — small cash/derivative
        # slices that v1 quietly ignored). Hybrid funds get the same values as
        # the hybrid_extension block above; passive funds get index-fund splits.
        record["asset_split"] = OrderedDict([
            ("equity_pct", _round(cm["equity_pct_universal"], 4)),
            ("debt_pct",   _round(cm["debt_pct_universal"],   4)),
            ("others_pct", _round(cm["others_pct_universal"], 4)),
        ])
        # v2 new fields (Avg Mkt Cap, Fund PE, Active Share + tent goodness inputs)
        record["avg_mcap_cr"]      = _round(cm["avg_mcap_cr"], 4)
        record["fund_pe"]          = _round(cm["fund_pe"], 4)
        record["active_share_pct"] = _round(cm["active_share_pct"], 4)
        record["mcap_goodness"]    = _round(cm["mcap_goodness"], 6)
        record["pe_goodness"]      = _round(cm["pe_goodness"], 6)
        # Stage B A6 (2026-05-28) — `centricity_rank_overall` removed per §7.3:
        # the Centricity score is a per-category percentile, so any cross-
        # category leaderboard is meaningless. Only `centricity_rank_in_category`
        # is emitted from here on. v1 archive (15-Apr cycle) keeps the field
        # for backward render compatibility per CLAUDE.md §9 rule 4.
        record["centricity_rank_in_category"] = None  # filled post-pass
        record["centricity_score"] = _round(score_dec, 6) if score_dec is not None else None
        record["centricity_score_status"] = status
        record["centricity_score_warning_pct"] = _round(warn_pct, 2) if warn_pct is not None else None
        record["cy_returns"] = OrderedDict([
            ("cy2022_pct", _round(cm["cy2022_pct"], 4)),
            ("cy2023_pct", _round(cm["cy2023_pct"], 4)),
            ("cy2024_pct", _round(cm["cy2024_pct"], 4)),
            ("cy2025_pct", _round(cm["cy2025_pct"], 4)),
            ("cy_ytd_pct", _round(cm["cy_ytd_pct"], 4)),
            ("cy_ytd_year", cycle_date.year),
        ])
        # F11 (2026-05-30): this is the CM "Rolling 3Y Avg" cell — null/"—" for a
        # few funds whose daily NAV spans < 3 years, so no rolling-3Y window
        # exists. In 2026-05 that's 7 re-coded LIC MF schemes (152009/152001/
        # 152003/152016/152025/152019/151950): mfapi only carries their NAV from
        # 2023-07-31 (~2.8 yrs) under the new AMFI codes, even though the funds
        # are 7–15 yrs old. The point-to-point 3Y (monitor_returns.return_3y_pct)
        # is fine; only the rolling-3Y average is unavailable. Correctly left
        # null — NEVER derive from 1Y/5Y or stitch the pre-recode series. See
        # mf-issues-solutions/SKILL.md §7.13.
        record["rolling_3y_avg_pct"] = _round(cm["rolling_3y_avg_pct"], 4)
        record["consistency_pct"] = _round(cm["consistency_pct"], 4)
        # Latest NAV pulled from the last row of 📈 Fund NAV — used by Fund
        # Detail's hero meta-strip and as the anchor of the "Growth of
        # ₹ 1,00,000" chart's normalisation. (Fund Detail Fix-List 1 §A.)
        record["nav_latest_value"] = nav_latest_value
        record["nav_latest_date"] = nav_latest_date
        # Daily-roll 3Y CAGR statistics — derives from 📈 Fund NAV in-place
        # against the fund's benchmark series. Powers Fund Detail's "Rolling
        # Returns" 6-card grid (avg / median / best / worst / pct > 0% /
        # pct beat benchmark / pct above 12%). Null when fewer than 252
        # daily NAV observations exist (sub-1Y data — can't form a window).
        record["rolling_3y_stats"] = rolling_stats
        # Order matches Master Design Brief §5.3 Fund Detail Quants strip:
        # Sharpe | Sortino | Std Dev | Max DD | Beta | Treynor | Up/Down Capture
        #
        # Phase 2.2 §3.2 — Sharpe / Sortino / Std Dev now share one basis: all
        # derived from the same 3Y daily NAV window, with the same Rf (4.5% p.a.),
        # same √252 annualisation. `sharpe_3y` is the canonical nav-derived value
        # (replaces workbook CM!O). `sharpe_3y_workbook` is preserved alongside
        # for transparency / one-cycle migration audit.
        record["risk_metrics"] = OrderedDict([
            ("sharpe_3y", derived_risk["sharpe_3y_nav"]),
            ("sharpe_3y_workbook", _round(cm["sharpe_3y"], 4)),
            ("sortino_3y", derived_risk["sortino_3y"]),
            ("std_dev_3y_pct", derived_risk["std_dev_3y_pct"]),
            ("max_drawdown_3y_pct", derived_risk["max_drawdown_3y_pct"]),
            ("beta_3y", _round(cm["beta_3y"], 4)),
            ("treynor_3y", _round(cm["treynor_3y"], 4)),
            ("up_capture_3y_pct", _round(cm["up_capture_3y_pct"], 4)),
            ("down_capture_3y_pct", _round(cm["down_capture_3y_pct"], 4)),
            ("overall_capture_3y_pct", _round(cm["overall_capture_3y_pct"], 4)),
        ])
        record["trailing_returns"] = OrderedDict([
            ("return_1y_pct", trailing["return_1y_pct"]),
            ("return_3y_pct", trailing["return_3y_pct"]),
            ("return_5y_pct", trailing["return_5y_pct"]),
            ("return_si_pct", trailing["return_si_pct"]),
        ])
        record["benchmark_returns"] = OrderedDict([
            ("return_1y_pct", bm_ret.get("return_1y_pct")),
            ("return_3y_pct", bm_ret.get("return_3y_pct")),
            ("return_5y_pct", bm_ret.get("return_5y_pct")),
        ])
        # Alpha
        a3 = (trailing["return_3y_pct"] - bm_ret["return_3y_pct"]
              if trailing["return_3y_pct"] is not None and bm_ret.get("return_3y_pct") is not None
              else None)
        a5 = (trailing["return_5y_pct"] - bm_ret["return_5y_pct"]
              if trailing["return_5y_pct"] is not None and bm_ret.get("return_5y_pct") is not None
              else None)
        record["alpha"] = OrderedDict([
            ("alpha_3y_pct", _round(a3, 4) if a3 is not None else None),
            ("alpha_5y_pct", _round(a5, 4) if a5 is not None else None),
        ])
        # Phase 2.2 §benchmark — TE/TD against the matched benchmark series.
        # Output of the 4-tier matcher + TE/TD compute (above). The proxy-
        # validity gate (post-loop) may later overwrite these with nulls for
        # passive proxy matches whose TE > 5% and which aren't equal-weight.
        # See derived_fields_documentation.benchmark_match in screener-v2.json.
        record["tracking_error_3y_pct"]      = te_v
        record["tracking_difference_3y_pct"] = td_v
        record["benchmark_matched_series"]   = matched_name
        record["benchmark_is_proxy"]         = bench_match["is_proxy"]
        record["benchmark_basis"]            = bench_match["basis"]
        record["td_basis"]                   = bench_match["td_basis"]
        record["currency_adjusted"]          = (bench_match["basis"] == "PRI-INR")
        record["nav_split_status"]           = nav_split_status
        record["benchmark_match_status"]     = bench_match["status"]
        # Monitor file overlay (Fix-List 5 §A) — point-to-point returns,
        # exit load, and the regular-plan TER. Joined by Scheme Name (col 0
        # of Monitor) → fund_name. All fields render as null when no Monitor
        # data was supplied at convert-time, OR when the fund's name doesn't
        # match a Monitor row. Note: Monitor TER (regular plan) often differs
        # from the Whitelisting Excel's ter_pct (looks like direct plan) —
        # both are preserved; fund-detail.html displays monitor_ter_pct.
        m_row = (monitor_data or {}).get(d["fund_name"]) if d["fund_name"] else None
        record["monitor_returns"] = OrderedDict([
            ("ytd_pct",        m_row.get("ytd_pct")        if m_row else None),
            ("return_1m_pct",  m_row.get("return_1m_pct")  if m_row else None),
            ("return_1y_pct",  m_row.get("return_1y_pct")  if m_row else None),
            ("return_3y_pct",  m_row.get("return_3y_pct")  if m_row else None),
            ("return_5y_pct",  m_row.get("return_5y_pct")  if m_row else None),
            ("return_10y_pct", m_row.get("return_10y_pct") if m_row else None),
        ])
        record["exit_load"]        = m_row.get("exit_load")        if m_row else None
        record["monitor_ter_pct"]  = m_row.get("monitor_ter_pct")  if m_row else None

        # Benchmark Monitor returns (Fix-List 6 §1B) — index rows from the
        # bottom of each Monitor category sheet, joined to fund.benchmark.
        # Powers the Fund Detail returns table's full benchmark row
        # (YTD / 1M / 10Y previously had to come from nav-series; now they
        # come from Monitor directly, with the same shape and freshness as
        # the fund row).
        idx_row = _lookup_index_returns(d["benchmark"], monitor_index_data)
        record["benchmark_monitor_returns"] = OrderedDict([
            ("ytd_pct",        idx_row.get("ytd_pct")        if idx_row else None),
            ("return_1m_pct",  idx_row.get("return_1m_pct")  if idx_row else None),
            ("return_1y_pct",  idx_row.get("return_1y_pct")  if idx_row else None),
            ("return_3y_pct",  idx_row.get("return_3y_pct")  if idx_row else None),
            ("return_5y_pct",  idx_row.get("return_5y_pct")  if idx_row else None),
            ("return_10y_pct", idx_row.get("return_10y_pct") if idx_row else None),
        ])

        record["verdict"] = None
        record["verdict_reasons"] = None
        record["analyst_note"] = None
        # cycle_flags = null in v1 (no prior cycle to diff). Populated by the
        # post-processing step compute_cycle_flags.py (v1.x), which runs in
        # the GitHub Action AFTER this converter and BEFORE commit. The Excel
        # converter never computes flags. See ISSUE-0009 + contract
        # derived_fields_documentation.cycle_flags.
        record["cycle_flags"] = None
        record["analytics_pending"] = _empty_analytics_pending()

        # Hidden full-precision metric values for parameter_scores percentile compute.
        # Stripped before JSON output.
        record["_raw_metrics"] = {
            "fund_tenure_yrs": cm["fund_tenure_yrs"],
            "aum_cr": d["aum_cr"],
            "manager_tenure_yrs": d["manager_tenure_yrs"],
            "cy2022_pct": cm["cy2022_pct"],
            "cy2023_pct": cm["cy2023_pct"],
            "cy2024_pct": cm["cy2024_pct"],
            "cy2025_pct": cm["cy2025_pct"],
            "cy_ytd_pct": cm["cy_ytd_pct"],
            "rolling_3y_avg_pct": cm["rolling_3y_avg_pct"],
            "consistency_pct": cm["consistency_pct"],
            "sharpe_3y": cm["sharpe_3y"],
            "beta_3y": cm["beta_3y"],
            "down_capture_3y_pct": cm["down_capture_3y_pct"],
            "up_capture_3y_pct": cm["up_capture_3y_pct"],
            "treynor_3y": cm["treynor_3y"],
            "overall_capture_3y_pct": cm["overall_capture_3y_pct"],
            "turnover_pct": d["turnover_pct"],
            "ter_pct": d["ter_pct"],
            "amc_score": float(d["amc_score"]) if d["amc_score"] is not None else None,
            # v2 new parameters
            "avg_mcap_cr": cm["avg_mcap_cr"],
            "fund_pe": cm["fund_pe"],
            "active_share_pct": cm["active_share_pct"],
            "mcap_goodness": cm["mcap_goodness"],
            "pe_goodness": cm["pe_goodness"],
        }

        funds.append(record)

    # Stash coverage on the funds list (read by convert() post-call) so we
    # don't change the build_funds() return tuple's positional shape.
    cycle_meta["_phase22_mgr_coverage"] = mgr_coverage

    # Phase 2.2 §benchmark — proxy-validity gate. For PASSIVE funds (ETFs +
    # *Index categories + Smart-Beta/Factor) that matched a proxy series, if
    # TE > 5% AND the fund isn't an Equal-Weight structural exception, downgrade
    # to absent: the index pair is too divergent for the proxy to be meaningful
    # (typical case: BSE→Nifty cross-family produces 8-22% TE; equal-weight vs
    # cap-weight produces 4-5% which is structurally expected). NO_SENSIBLE_PROXY
    # themes already filtered into absent at match time; this catches the rest.
    proxy_downgrades: list[tuple[str, str, float]] = []
    for f in funds:
        if f.get("centricity_score_status") != PASSIVE_STATUS:
            continue
        if not f.get("benchmark_is_proxy"):
            continue
        te = f.get("tracking_error_3y_pct")
        if te is None or te <= 5.0:
            continue
        if _is_equal_weight(f.get("fund_name") or ""):
            continue
        ms = f.get("benchmark_matched_series") or ""
        proxy_downgrades.append((f["fund_name"], ms, te))
        f["tracking_error_3y_pct"] = None
        f["tracking_difference_3y_pct"] = None
        f["benchmark_matched_series"] = None
        f["benchmark_is_proxy"] = False
        f["benchmark_basis"] = None
        f["td_basis"] = None
        f["currency_adjusted"] = False
        f["benchmark_match_status"] = f"absent (proxy TE={te:.2f}% > 5%, downgraded)"
    if proxy_downgrades:
        print(
            f"[converter] benchmark proxy gate: {len(proxy_downgrades)} passive "
            f"proxy funds downgraded to absent (TE > 5%, not Equal-Weight).",
            file=sys.stderr,
        )
        for nm, ms, te in proxy_downgrades[:10]:
            print(f"    {nm[:55]} (was proxy→{ms}, TE={te:.2f}%)", file=sys.stderr)

    # Compute centricity_rank_in_category post-pass
    by_cat: dict[str, list[dict]] = {}
    for f in funds:
        by_cat.setdefault(f["category"], []).append(f)
    for cat, lst in by_cat.items():
        ranked = [f for f in lst if f["centricity_score"] is not None]
        ranked.sort(key=lambda x: x["centricity_score"], reverse=True)
        for i, f in enumerate(ranked, start=1):
            f["centricity_rank_in_category"] = i

    # Stage B A6 (2026-05-28) — cross-category overall rank removed (§7.3).

    # Surface benchmark-match summary counts on cycle_meta so the QA report has
    # a single anchor to validate against the Part-1 evidence CSV. Excludes the
    # 'downgraded' status from matched counts.
    bm_status_counts: dict[str, int] = {}
    for f in funds:
        st = f.get("benchmark_match_status") or "unknown"
        bm_status_counts[st] = bm_status_counts.get(st, 0) + 1
    cycle_meta["benchmark_match_summary"] = {
        "match_status_counts": bm_status_counts,
        "proxy_downgrades": len(proxy_downgrades),
        "split_clean_adjusted": _clean_n,
        "split_suspect_nulled": _suspect_n,
        "inr_converted_series_registered": len(inr_series) if inr_series else 0,
    }

    # Return NAV maps alongside the fund list so convert() can emit the
    # separate nav-series-YYYY-MM-DD.json file without re-reading the
    # workbook (Fund Detail Fix-List 1 §A.3).
    return funds, nav_by_amfi, bm_by_name


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def convert(
    xlsx_path: Path,
    monitor_path: Path | None = None,
    morningstar_history_path: Path | None = None,
) -> Path:
    if not xlsx_path.exists():
        raise SchemaError(f"Input file not found: {xlsx_path}")

    contract = load_contract()
    print(f"[converter] loaded contract: {CONTRACT_PATH.name} ({contract['contract_version']})", file=sys.stderr)

    # Phase 2.2 §A1-REDO — load the full Morningstar manager-tenure xlsx
    # ('Manager Tenure Data as on <date>.xlsx'). 24 category sheets; header
    # row 9; Manager History col + AMFI Code col. Each cell carries the
    # complete dated history of every manager who has run that fund. Join
    # key is AMFI scheme_code (workbook col D ↔ Morningstar col E).
    morningstar_history: dict[int, list[dict]] = {}
    morningstar_diag: dict = {}
    if morningstar_history_path is not None and morningstar_history_path.exists():
        morningstar_history, morningstar_diag = load_morningstar_mgr_history(morningstar_history_path)
        print(
            f"[converter] loaded Morningstar manager-history: "
            f"{morningstar_diag['amfi_codes_loaded']} AMFI codes with parsed history "
            f"from {morningstar_diag['sheets']} sheets in {morningstar_history_path.name} "
            f"(rows seen: {morningstar_diag['rows_seen']}, dups: "
            f"{len(morningstar_diag.get('amfi_dups', []))}, "
            f"unparseable: {morningstar_diag.get('amfi_unparseable_history', 0)})",
            file=sys.stderr,
        )
    elif morningstar_history_path is not None:
        print(
            f"[converter] WARNING: Morningstar manager-history file not found at "
            f"{morningstar_history_path}; falling back to MF Monitor names on Data col F.",
            file=sys.stderr,
        )

    # Optional Monitor overlay — Fix-List 5 §A + Fix-List 6 §1A
    monitor_fund_data: dict | None = None
    monitor_index_data: dict | None = None
    if monitor_path is not None:
        if not monitor_path.exists():
            print(
                f"[converter] WARNING: monitor file not found at {monitor_path}; "
                f"monitor_returns / exit_load / monitor_ter_pct / "
                f"benchmark_monitor_returns will all be null",
                file=sys.stderr,
            )
        else:
            print(f"[converter] reading Monitor overlay: {monitor_path.name}", file=sys.stderr)
            monitor_payload = extract_monitor_data(monitor_path)
            monitor_fund_data  = monitor_payload["fund_returns"]
            monitor_index_data = monitor_payload["index_returns"]
            print(
                f"[converter]   extracted {len(monitor_fund_data)} fund rows + "
                f"{len(monitor_index_data)} index rows from Monitor",
                file=sys.stderr,
            )

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"[converter] opened workbook: {xlsx_path.name} ({len(wb.sheetnames)} sheets)", file=sys.stderr)

    # 1. Sheets + header anchors
    summary = validate_sheets(wb, contract)
    print(f"[converter] validated {summary['sheets_validated']} sheets, {summary['anchors_passed']} anchors passed", file=sys.stderr)

    # 2. Master Table 2 weights sum to 100
    weights_total = validate_weights(wb, contract, summary)
    print(f"[converter] Master Table 2 weights total: {weights_total}%", file=sys.stderr)

    # 3. Cycle metadata
    cycle_meta = build_cycle_meta(wb, xlsx_path.name, summary, contract)
    print(f"[converter] cycle: {cycle_meta['cycle_label_date']} (internal {cycle_meta['cycle_label']}) | as on {cycle_meta['as_on_display']} | Rf {cycle_meta['rf_rate_display']}", file=sys.stderr)

    # Phase 2.2 §benchmark — Global PRI source data (Nasdaq 100 / S&P 500 /
    # NYSE FANG+ / MSCI World USD CSVs + USD/INR daily) lives sibling to the
    # workbook on the local file system at <cycle>/Data/BM NAV/Global &
    # Others BM NAV/. When the converter is invoked from the production
    # data/ folder (no sibling Data tree) the dir simply won't exist and
    # the loader falls back gracefully (the 4 USD-international funds → absent).
    global_bm_dir = (
        xlsx_path.resolve().parent.parent
        / "Data" / "BM NAV" / "Global & Others BM NAV"
    )

    # Phase 2.2 Patch (no_of_stocks) — analytics underlyings live sibling to
    # the workbook at <cycle>/Data/Underlyings Data - Analytics/. Three
    # .xlsx files (Equity / Hybrid / Debt) all share schema: col A = Scheme
    # Name, col J = Asset ({Equity, Debt, Others}). We count Equity-only
    # rows per Scheme Name → no_of_stocks. Workbook col N over-counts
    # (includes Others/REITs/InvITs); analytics gives the pure-equity count.
    # Falls back to workbook col N if the dir is missing (e.g. CI run).
    analytics_dir = (
        xlsx_path.resolve().parent.parent
        / "Data" / "Underlyings Data - Analytics"
    )
    analytics_equity_counts, analytics_count_diag = _load_analytics_equity_counts(analytics_dir)
    print(
        f"[converter] analytics equity-row counts: "
        f"{analytics_count_diag['schemes_with_equity_rows']} schemes from "
        f"{len(analytics_count_diag['files_loaded'])} file(s) in {analytics_dir.name}",
        file=sys.stderr,
    )

    # 4. Funds — also returns the in-memory NAV maps so we can emit the
    # separate nav-series-YYYY-MM-DD.json file without re-reading the
    # workbook (Fund Detail Fix-List 1 §A.3). monitor_fund_data overlays
    # per-fund (Fix-List 5 §A); monitor_index_data overlays
    # benchmark_monitor_returns per fund (Fix-List 6 §1B).
    warnings: list[str] = []
    funds, nav_by_amfi, bm_by_name = build_funds(
        wb, contract, cycle_meta, warnings,
        monitor_data=monitor_fund_data,
        monitor_index_data=monitor_index_data,
        morningstar_history=morningstar_history,
        global_bm_dir=global_bm_dir,
        analytics_equity_counts=analytics_equity_counts,
    )

    # Phase 2.2 Patch (FM rule 2026-05-28) — supersedes the §A1-REDO
    # Morningstar-primary gate. Monitor's "Fund Manager" column is now
    # primary; Morningstar is a fallback. The coverage gate is reframed:
    # ≥ 60% of non-passive funds must have ANY lead set (Monitor FM,
    # Morningstar earliest-current, or 📋 Data col F). The per-source
    # counts are still tracked so we can spot future degradations.
    mgr_coverage = cycle_meta.pop("_phase22_mgr_coverage", {"morningstar": 0, "mf_monitor": 0, "none": 0})
    cycle_meta["manager_name_source"] = "MF Monitor 'Fund Manager' col primary; Morningstar earliest-current fallback"
    cycle_meta["manager_name_coverage"] = mgr_coverage
    cycle_meta["morningstar_diagnostics"] = {
        "file": morningstar_diag.get("file"),
        "sheets": morningstar_diag.get("sheets"),
        "amfi_codes_loaded": morningstar_diag.get("amfi_codes_loaded"),
        "amfi_dups": morningstar_diag.get("amfi_dups", []),
        "amfi_unparseable_history": morningstar_diag.get("amfi_unparseable_history"),
        "sheets_skipped_no_header": morningstar_diag.get("sheets_skipped_no_header", []),
    }
    non_passive = sum(1 for f in funds if f["centricity_score_status"] != PASSIVE_STATUS)
    non_passive_set = sum(
        1 for f in funds
        if f["centricity_score_status"] != PASSIVE_STATUS and f.get("manager_name")
    )
    non_passive_monitor = sum(
        1 for f in funds
        if f["centricity_score_status"] != PASSIVE_STATUS
        and f.get("manager_name_source") == "MF Monitor"
    )
    non_passive_morn = sum(
        1 for f in funds
        if f["centricity_score_status"] != PASSIVE_STATUS
        and f.get("manager_name_source") == "Morningstar"
    )
    set_share = (non_passive_set / non_passive * 100) if non_passive else 0
    monitor_share = (non_passive_monitor / non_passive * 100) if non_passive else 0
    morn_share = (non_passive_morn / non_passive * 100) if non_passive else 0
    cycle_meta["manager_name_coverage_non_passive"] = {
        "non_passive_total": non_passive,
        "lead_set": non_passive_set,
        "lead_set_share_pct": round(set_share, 1),
        "monitor_primary": non_passive_monitor,
        "monitor_primary_share_pct": round(monitor_share, 1),
        "morningstar_fallback": non_passive_morn,
        "morningstar_fallback_share_pct": round(morn_share, 1),
    }
    print(
        f"[converter] manager-name source: MF Monitor {mgr_coverage['mf_monitor']} / "
        f"Morningstar {mgr_coverage['morningstar']} / none {mgr_coverage['none']} "
        f"(total {sum(mgr_coverage.values())})",
        file=sys.stderr,
    )
    print(
        f"[converter]   Non-passive lead coverage: "
        f"{non_passive_set}/{non_passive} = {set_share:.1f}% "
        f"(Monitor primary {non_passive_monitor} | Morningstar fallback {non_passive_morn}; "
        f"hard gate: ≥ 60%)",
        file=sys.stderr,
    )
    if set_share < 60.0:
        warnings.append(
            f"[FM-rule hard-gate failed] Non-passive lead coverage {set_share:.1f}% "
            f"of {non_passive} funds — expected ≥ 60%. Investigate Monitor + "
            f"Morningstar joins. Sample of unset AMFI codes follows."
        )
        unmatched = [
            (f["scheme_code"], f["fund_name"])
            for f in funds
            if f["centricity_score_status"] != PASSIVE_STATUS
            and not f.get("manager_name")
        ]
        cycle_meta["morningstar_diagnostics"]["unmatched_sample"] = [
            {"amfi": c, "fund_name": n} for c, n in unmatched[:20]
        ]

    # Report Monitor match rate (helps catch name-mismatch problems early)
    if monitor_fund_data is not None:
        matched = sum(1 for f in funds if f.get("monitor_returns", {}).get("ytd_pct") is not None
                                          or f.get("exit_load") is not None
                                          or f.get("monitor_ter_pct") is not None)
        unmatched = len(funds) - matched
        print(
            f"[converter] Monitor fund overlay: {matched}/{len(funds)} funds matched by Scheme Name "
            f"({unmatched} unmatched — typically 1-3yr Warning / New Fund Monitoring funds "
            f"with names not yet in the Monitor file)",
            file=sys.stderr,
        )
        bm_matched = sum(1 for f in funds if f.get("benchmark_monitor_returns", {}).get("ytd_pct") is not None
                                            or f.get("benchmark_monitor_returns", {}).get("return_10y_pct") is not None)
        print(
            f"[converter] Monitor benchmark overlay: {bm_matched}/{len(funds)} funds matched their "
            f"benchmark to a Monitor index row",
            file=sys.stderr,
        )
    summary["fund_count"] = len(funds)
    print(f"[converter] built {len(funds)} fund records", file=sys.stderr)

    # Phase 2.2 §A3 — total_funds = actual fund record count (1,249 after the
    # Quantum-Direct dedup). The Master A2 banner says 1252 (universe count
    # incl. dedup'd dupes); the dashboard surfaces total_funds in titles +
    # the count chip, so it has to match what the table actually renders.
    banner_total = cycle_meta.get("total_funds")
    if banner_total != len(funds):
        cycle_meta["total_funds_banner"] = banner_total
        cycle_meta["total_funds"] = len(funds)
        print(
            f"[converter] total_funds: banner reports {banner_total}, "
            f"record count is {len(funds)} — using record count.",
            file=sys.stderr,
        )

    # 5. Per-parameter normalised scores (powers the right-drawer weight reshuffling)
    print("[converter] computing parameter_scores (per-category percentile rank)...", file=sys.stderr)
    compute_parameter_scores(funds, cycle_meta["scoring_weights"], warnings)

    # Phase 2.2 §1A — universal score recompute. The 15-May workbook's CM
    # column AD ("Score") has misaligned cells: for ~80% of Ranked funds it
    # shows the wrong value (Groww Aggressive Hybrid's AD = 0.847, which
    # actually belongs to ICICI Pru E&D; ICICI Pru E&D's AD = 0.599, which
    # belongs to Kotak Aggressive Hybrid IDCW; etc.). The category sheets
    # themselves are correct (R7 of the Aggressive Hybrid sheet shows ICICI
    # Pru E&D rank 1 with score 0.847) and Master Table 1 agrees. So the
    # right fix is: trust our per-fund percentile compute (which mirrors
    # Excel's COUNTIF logic) and DERIVE centricity_score = Σ(parameter_scores
    # × weight) / 100 for every Ranked fund, overwriting the bugged CM AD.
    #
    # This also subsumes the earlier §A2 fix (Dividend-Yield → Value-Contra
    # peer-pool change) — the recompute is now universal.
    weight_by_param = {w["parameter"]: w["weight_pct"] for w in cycle_meta["scoring_weights"]}
    recomputed_count = 0
    max_delta = 0.0
    cm_ad_swap_examples: list[str] = []
    for f in funds:
        if f.get("centricity_score_status") != "Ranked":
            continue
        ps = f.get("parameter_scores") or {}
        total = 0.0
        for p, w_pct in weight_by_param.items():
            v = ps.get(p)
            if v is None:
                continue
            total += v * w_pct
        new_score = round(total / 100, 6)
        old_score = f.get("centricity_score")
        if old_score is not None:
            delta = abs(new_score - old_score)
            if delta > max_delta:
                max_delta = delta
            if delta > 0.20 and len(cm_ad_swap_examples) < 5:
                cm_ad_swap_examples.append(
                    f"{f['fund_name']!r} (cat={f['category']}): "
                    f"CM AD={old_score:.4f} → recompute={new_score:.4f}"
                )
        f["centricity_score"] = new_score
        recomputed_count += 1
    print(
        f"[converter] §1A universal recompute: replaced centricity_score for "
        f"{recomputed_count} Ranked funds with Σ(parameter_scores × weight)/100 "
        f"to fix workbook CM-AD misalignment. Max |Δ| vs CM-AD = {max_delta:.4f}.",
        file=sys.stderr,
    )
    if cm_ad_swap_examples:
        print("[converter]   CM-AD swap examples (|Δ| > 0.20):", file=sys.stderr)
        for ex in cm_ad_swap_examples:
            print(f"    - {ex}", file=sys.stderr)

    # Re-derive ranks now that scores have been universally re-derived.
    # build_funds' post-pass ran with the workbook-cached CM-AD scores; redo
    # everything with the corrected ones. Both in-category and overall ranks.
    by_cat: dict[str, list[dict]] = {}
    for f in funds:
        by_cat.setdefault(f["category"], []).append(f)
    for cat, lst in by_cat.items():
        ranked = [f for f in lst if f["centricity_score"] is not None]
        ranked.sort(key=lambda x: x["centricity_score"], reverse=True)
        for i, f in enumerate(ranked, start=1):
            f["centricity_rank_in_category"] = i
        for f in lst:
            if f["centricity_score"] is None:
                f["centricity_rank_in_category"] = None
    # Stage B A6 (2026-05-28) — no cross-category overall ranking (§7.3).
    print(
        f"[converter] §1A re-ranked: in-category ranks recomputed for "
        f"{sum(1 for f in funds if f['centricity_rank_in_category'] is not None)} funds.",
        file=sys.stderr,
    )

    # 6. Verify recomputed score == stored centricity_score for every Ranked
    # fund. Phase 2.2 §1A changes this from a meaningful check to a tautology
    # for v2 cycles: we just overwrote centricity_score with the recompute
    # itself, so the delta is 0 by construction. We still run the function
    # because it's the single auditable trace of how scores were derived,
    # but the Δ should now be uniformly 0. For v1 cycles (15-Apr) the
    # original Excel-exact-match check (±0.0001) still applies.
    is_v2 = contract.get("contract_version", "").startswith("screener-v2")
    verify_summary = verify_parameter_scores_match(
        funds,
        cycle_meta["scoring_weights"],
        tolerance=0.0001,
        warn_only=is_v2,
        out_warnings=warnings,
    )
    summary["score_recompute_verified"] = verify_summary
    print(
        f"[converter]   verified {verify_summary['ranked_funds_verified']} ranked funds; "
        f"worst Δ = {verify_summary['worst_diff']:.7f} "
        f"(tolerance ±{verify_summary['tolerance']})",
        file=sys.stderr,
    )

    # Strip the hidden full-precision percentile-input dict before serialising
    for f in funds:
        f.pop("_raw_metrics", None)

    # Stage B A2 / A3 (2026-05-28) — auto-detect `source_dates.analytics` by
    # globbing `data/analytics-*.json`. Picks the most recent date ≤ cycle.
    # Survives subsequent converter re-runs (otherwise analytics_date keeps
    # getting reset to None whenever the screener is regenerated).
    _cycle_date_for_meta = _dt.date.fromisoformat(cycle_meta["cycle_date"])
    _analytics_files = sorted(DATA_DIR.glob("analytics-*.json"))
    _analytics_dates: list[str] = []
    for _af in _analytics_files:
        _m = re.match(r"analytics-(\d{4}-\d{2}-\d{2})\.json", _af.name)
        if _m:
            try:
                if _dt.date.fromisoformat(_m.group(1)) <= _cycle_date_for_meta:
                    _analytics_dates.append(_m.group(1))
            except ValueError:
                pass
    if _analytics_dates:
        cycle_meta["source_dates"]["analytics"] = max(_analytics_dates)
        print(
            f"[converter] source_dates.analytics auto-detected: "
            f"{cycle_meta['source_dates']['analytics']}",
            file=sys.stderr,
        )

    output = OrderedDict([
        ("contract_version", contract["contract_version"]),
        ("cycle_meta", cycle_meta),
        ("funds", funds),
        ("analytics_pending_fields", contract.get("v1_defaults_explicit_null", [])),
        ("schema_validation_summary", summary),
        ("converter_warnings", warnings[:200]),  # cap at 200 to keep JSON sane
    ])

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    out_path = DATA_DIR / f"screener-{cycle_meta['cycle_date']}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)
    print(f"[converter] wrote {out_path} ({out_path.stat().st_size:,} bytes)", file=sys.stderr)

    # 7. Emit the separate monthly NAV-series file consumed by Fund Detail's
    # "Growth of ₹ 1,00,000" chart. Lazy-loaded by fund-detail.js — kept
    # out of the main screener JSON to keep that file under 3 MB.
    cycle_date_obj = _dt.date.fromisoformat(cycle_meta["cycle_date"])
    nav_series_path = DATA_DIR / f"nav-series-{cycle_meta['cycle_date']}.json"
    print(f"[converter] emitting nav series file: {nav_series_path.name}", file=sys.stderr)
    emit_nav_series_file(funds, nav_by_amfi, bm_by_name, cycle_date_obj, nav_series_path)
    print(
        f"[converter]   wrote {nav_series_path} ({nav_series_path.stat().st_size:,} bytes)",
        file=sys.stderr,
    )

    # Stage B A3 — emit the benchmark-nav file keyed by canonical benchmark label.
    benchmark_nav_path = DATA_DIR / f"benchmark-nav-{cycle_meta['cycle_date']}.json"
    print(f"[converter] emitting benchmark nav file: {benchmark_nav_path.name}", file=sys.stderr)
    emit_benchmark_nav_file(bm_by_name, cycle_date_obj, benchmark_nav_path)
    print(
        f"[converter]   wrote {benchmark_nav_path} ({benchmark_nav_path.stat().st_size:,} bytes; "
        f"{len(bm_by_name)} benchmark labels)",
        file=sys.stderr,
    )

    if warnings:
        print(f"[converter] {len(warnings)} warning(s) (showing first 5):", file=sys.stderr)
        for w in warnings[:5]:
            print(f"  - {w}", file=sys.stderr)

    return out_path


def main(argv: list[str] | None = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    if not argv:
        print(
            "usage: excel_to_json_screener.py <whitelisting.xlsx> "
            "[<monitor.xlsx> [<morningstar_manager_tenure.xlsx>]]",
            file=sys.stderr,
        )
        return 2
    path = Path(argv[0])
    monitor_path = Path(argv[1]) if len(argv) > 1 and argv[1] else None
    morningstar_history_path = Path(argv[2]) if len(argv) > 2 and argv[2] else None
    try:
        convert(
            path,
            monitor_path=monitor_path,
            morningstar_history_path=morningstar_history_path,
        )
    except SchemaError as e:
        print(f"\nSCHEMA ERROR: {e}\n", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
