"""
Analytics Excel-to-JSON converter — Equity + Hybrid (Fund Detail Fix-List 5 §B).

Reads two source workbooks for one month-end snapshot:
    Cent-Claude/Data/Analytics File/<DD-MM-YYYY>/
        EQUITY MF.xlsx   (Sheet1, ~101k holdings rows)
        HYBRID FUNDS.xlsx (Sheet1, ~21k holdings rows)

Both share the same column structure (header row 1):
    Col 0 : Scheme Name              ← match key (string == fund_name in screener JSON)
    Col 1 : Company Name
    Col 3 : Holding(%)               ← portfolio weight
    Col 4 : Market Value (Rs in Cr)
    Col 8 : Sector
    Col 18: SEBI MCAP Type           ← 'Large Cap' / 'Mid Cap' / 'Small Cap' / 'Others'

Per fund (joined to screener JSON by Scheme Name → scheme_code):
    1. Sort holdings by Holding(%) desc; take top 20 (after the cash/TREPS
       exclusion below).
    2. Exclude rows where Company Name contains 'Repo', 'TREPS', 'T-Bill',
       or 'Treasury' (case-insensitive). Sum these into `cash_and_equiv_pct`.
    3. Sector allocation: group remaining rows by Sector, sum Holding(%),
       sort desc, round to 2dp.
    4. top_10_concentration_pct = sum of top 10 holdings' weights.
    5. top_3_sector_concentration_pct = sum of top 3 sectors' weights.

Output: data/analytics-YYYY-MM-DD.json (date derived from folder name).

Funds without a name match in the screener JSON are silently skipped — the
dashboard will fall back to the analytics_pending placeholder for them.

Usage:
    python scripts/excel_to_json_analytics.py <equity.xlsx> <hybrid.xlsx>
       [--analytics-date YYYY-MM-DD]
       [--cycle-json data/screener-YYYY-MM-DD.json]

If --cycle-json is omitted, the latest data/screener-*.json is used to
build the fund_name → scheme_code lookup. If --analytics-date is omitted,
today's ISO date is used.

Debt analytics deferred — see CLAUDE.md §4.1 row for `debt-analytics-v1`.
"""
from __future__ import annotations

import datetime as _dt
import json
import re
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"

# Holdings excluded from top-20 / sectors but rolled up into cash_and_equiv
CASH_EQUIV_PATTERNS = re.compile(r"\b(Repo|TREPS|T[- ]?Bill|Treasury)\b", re.IGNORECASE)


def _safe_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _safe_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _is_cash_equiv(company_name: str | None) -> bool:
    if not company_name:
        return False
    return bool(CASH_EQUIV_PATTERNS.search(company_name))


def _stream_analytics(xlsx_path: Path):
    """
    Yield (scheme_name, company_name, holding_pct, sector, mcap_type, asset)
    tuples for every holding row in Sheet1 of the file.

    `asset` is the col-9 ("Asset") categorical: in the 31-Mar-2026 files
    it carries one of {Equity, Debt, Others}. The Fix-List 9 Feature A
    full-holdings extractor uses it to keep equity-only positions when
    emitting the holdings-full JSON.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue                                # header
        if not row:
            continue
        scheme = _safe_str(row[0]) if len(row) > 0 else None
        if not scheme:
            continue
        company  = _safe_str(row[1])  if len(row) > 1  else None
        holding  = _safe_float(row[3]) if len(row) > 3  else None
        sector   = _safe_str(row[8])  if len(row) > 8  else None
        asset    = _safe_str(row[9])  if len(row) > 9  else None
        mcap     = _safe_str(row[18]) if len(row) > 18 else None
        yield scheme, company, holding, sector, mcap, asset
    wb.close()


def _build_per_fund(rows_by_scheme: dict[str, list[dict]]) -> dict[str, dict]:
    """For every scheme, build the per-fund analytics block."""
    out: dict[str, dict] = {}
    for scheme, holdings in rows_by_scheme.items():
        # Split: real equity holdings vs cash-equiv (Repo/TREPS/T-Bill/Treasury)
        non_cash = [h for h in holdings if not _is_cash_equiv(h["company"])]
        cash_eq = [h for h in holdings if _is_cash_equiv(h["company"])]

        # Sort by holding% desc
        non_cash.sort(key=lambda h: -(h["holding_pct"] or 0))

        top_20 = []
        for i, h in enumerate(non_cash[:20], start=1):
            top_20.append(OrderedDict([
                ("rank", i),
                ("company", h["company"]),
                ("holding_pct", round(h["holding_pct"] or 0, 4)),
                ("sector", h["sector"]),
                ("mcap_type", h["mcap"]),
            ]))

        # Sector allocation — sum holding% by sector, exclude cash-equiv
        by_sector: dict[str, float] = defaultdict(float)
        for h in non_cash:
            sec = h["sector"] or "—"
            by_sector[sec] += h["holding_pct"] or 0
        sector_alloc = [
            OrderedDict([("sector", s), ("holding_pct", round(p, 2))])
            for s, p in sorted(by_sector.items(), key=lambda kv: -kv[1])
        ]

        top_10_concentration_pct = round(
            sum((h["holding_pct"] or 0) for h in non_cash[:10]), 2
        )
        top_3_sector_concentration_pct = round(
            sum(s["holding_pct"] for s in sector_alloc[:3]), 2
        )
        cash_pct = round(sum((h["holding_pct"] or 0) for h in cash_eq), 2) if cash_eq else 0.0

        out[scheme] = OrderedDict([
            ("total_holdings_count", len(non_cash)),
            ("top_20_holdings", top_20),
            ("sector_allocation", sector_alloc),
            ("top_10_concentration_pct", top_10_concentration_pct),
            ("top_3_sector_concentration_pct", top_3_sector_concentration_pct),
            ("cash_and_equiv_pct", cash_pct),
        ])
    return out


def _build_full_holdings(rows_by_scheme: dict[str, list[dict]],
                         max_per_fund: int = 200) -> dict[str, list[dict]]:
    """
    Fix-List 9 Feature A — full equity-only holdings per fund.

    For each scheme:
      • Keep rows where Asset == 'Equity' (covers Domestic Equities,
        Overseas Equities, ADRs & GDRs in the source file).
      • Drop holdings with null / zero Holding(%).
      • Sort desc by Holding(%) and cap at `max_per_fund` (default 200 —
        covers every active equity fund in the universe; the only
        portfolios that exceed 200 lines are NIFTY-500 / Total-Market
        index funds where weight contribution beyond rank-200 is
        negligible).
      • Emit a flat list of {company, holding_pct, sector} dicts. The
        overlap calculator and Similar Funds widget read this list
        directly; no rank field is needed because order is implicit.

    Returns { scheme_name: [holdings] }; the convert() function re-keys
    to scheme_code at emit time.
    """
    out: dict[str, list[dict]] = {}
    for scheme, holdings in rows_by_scheme.items():
        equity = []
        for h in holdings:
            asset = (h.get("asset") or "").strip()
            if asset != "Equity":
                continue
            pct = h.get("holding_pct")
            if pct is None or pct <= 0:
                continue
            equity.append(OrderedDict([
                ("company", h.get("company") or "—"),
                ("holding_pct", round(pct, 4)),
                ("sector", h.get("sector") or "—"),
            ]))
        equity.sort(key=lambda h: -(h["holding_pct"] or 0))
        if len(equity) > max_per_fund:
            equity = equity[:max_per_fund]
        if equity:
            out[scheme] = equity
    return out


def _latest_screener_json() -> Path | None:
    candidates = sorted(DATA_DIR.glob("screener-*.json"), reverse=True)
    return candidates[0] if candidates else None


def _load_name_to_scheme(cycle_json: Path) -> tuple[dict[str, int], str]:
    with open(cycle_json, "r", encoding="utf-8") as f:
        j = json.load(f)
    out: dict[str, int] = {}
    for fund in j.get("funds", []):
        name = fund.get("fund_name")
        code = fund.get("scheme_code")
        if name and code:
            out[name] = int(code)
    cycle_date = (j.get("cycle_meta") or {}).get("cycle_date") or "unknown"
    return out, cycle_date


def convert(
    equity_path: Path,
    hybrid_path: Path,
    debt_path: Path | None = None,
    *,
    analytics_date: str | None = None,
    cycle_json_path: Path | None = None,
) -> Path:
    if cycle_json_path is None:
        cycle_json_path = _latest_screener_json()
    if cycle_json_path is None or not cycle_json_path.exists():
        raise SystemExit(
            "[analytics] No screener-*.json found in data/. "
            "Run scripts/excel_to_json_screener.py first."
        )

    name_to_scheme, screener_cycle_date = _load_name_to_scheme(cycle_json_path)
    print(
        f"[analytics] using screener JSON {cycle_json_path.name} "
        f"({len(name_to_scheme)} funds in lookup) ",
        file=sys.stderr,
    )

    if analytics_date is None:
        # Try to derive from the equity_path's parent folder name (DD-MM-YYYY)
        m = re.search(r"(\d{2})-(\d{2})-(\d{4})", str(equity_path.parent))
        if m:
            analytics_date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        else:
            analytics_date = _dt.date.today().isoformat()

    # Stream + group by scheme name
    rows_by_scheme: dict[str, list[dict]] = defaultdict(list)
    sources: list[str] = []
    # Stage B A2 (2026-05-28) — debt analytics now included alongside Equity +
    # Hybrid (kickoff scope). The Debt file shares the same Sheet1 column
    # layout (col 0 Scheme Name, col 1 Company, col 3 Holding%, col 8 Sector,
    # col 9 Asset, col 18 MCAP Type) so _stream_analytics handles it uniformly.
    # Per-fund holdings for debt funds are written to the same analytics JSON;
    # the dashboard renders them in the same shape as equity/hybrid funds.
    for path in (equity_path, hybrid_path, debt_path):
        if path is None or not path.exists():
            continue
        sources.append(path.name)
        print(f"[analytics] streaming {path.name}", file=sys.stderr)
        n = 0
        for scheme, company, holding, sector, mcap, asset in _stream_analytics(path):
            rows_by_scheme[scheme].append({
                "company": company, "holding_pct": holding,
                "sector": sector, "mcap": mcap, "asset": asset,
            })
            n += 1
        print(f"[analytics]   {n:,} holdings rows from {path.name}", file=sys.stderr)

    print(f"[analytics] grouping into {len(rows_by_scheme)} unique scheme names", file=sys.stderr)
    per_fund = _build_per_fund(rows_by_scheme)

    # Re-key by scheme_code (string-of-int) — drop schemes without a screener match
    matched: dict[str, dict] = {}
    unmatched: list[str] = []
    for scheme_name, payload in per_fund.items():
        code = name_to_scheme.get(scheme_name)
        if code is None:
            unmatched.append(scheme_name)
            continue
        matched[str(code)] = payload
    print(
        f"[analytics] matched {len(matched)}/{len(per_fund)} funds to scheme_code "
        f"({len(unmatched)} unmatched — typically Debt funds or sub-3y new funds)",
        file=sys.stderr,
    )

    output = OrderedDict([
        ("analytics_date", analytics_date),
        ("source", " + ".join(sources)),
        ("matched_funds", len(matched)),
        ("unmatched_count", len(unmatched)),
        ("funds", matched),
    ])

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    out_path = DATA_DIR / f"analytics-{analytics_date}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))
    print(
        f"[analytics] wrote {out_path} ({out_path.stat().st_size:,} bytes; "
        f"screener cycle {screener_cycle_date})",
        file=sys.stderr,
    )

    # ----- Fix-List 9 Feature A — emit holdings-full-YYYY-MM-DD.json -----
    # Same scheme-name → scheme-code lookup; same per-fund matching rule.
    # Output is a flat list per fund (no rank, no concentration roll-up;
    # those are still in the analytics file). Powers the overlap matrix
    # and Similar Funds widget on Fund Detail.
    full_per_fund_by_name = _build_full_holdings(rows_by_scheme, max_per_fund=200)
    matched_full: dict[str, list[dict]] = {}
    unmatched_full = 0
    total_holdings = 0
    for scheme_name, holdings in full_per_fund_by_name.items():
        code = name_to_scheme.get(scheme_name)
        if code is None:
            unmatched_full += 1
            continue
        matched_full[str(code)] = holdings
        total_holdings += len(holdings)
    full_payload = OrderedDict([
        ("holdings_date", analytics_date),
        ("source", " + ".join(sources)),
        ("matched_funds", len(matched_full)),
        ("unmatched_count", unmatched_full),
        ("total_holdings_rows", total_holdings),
        ("max_per_fund", 200),
        ("filter", "Asset == 'Equity'"),
        ("funds", matched_full),
    ])
    full_out_path = DATA_DIR / f"holdings-full-{analytics_date}.json"
    with open(full_out_path, "w", encoding="utf-8") as f:
        json.dump(full_payload, f, ensure_ascii=False, separators=(",", ":"))
    print(
        f"[analytics] wrote {full_out_path} ({full_out_path.stat().st_size:,} bytes; "
        f"{len(matched_full)} funds, {total_holdings:,} equity holdings)",
        file=sys.stderr,
    )
    return out_path


def main(argv: list[str] | None = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    positional: list[str] = []
    analytics_date: str | None = None
    cycle_json: str | None = None
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--analytics-date" and i + 1 < len(argv):
            analytics_date = argv[i + 1]; i += 2; continue
        if a == "--cycle-json" and i + 1 < len(argv):
            cycle_json = argv[i + 1]; i += 2; continue
        positional.append(a); i += 1
    if len(positional) < 2:
        print(
            "usage: excel_to_json_analytics.py <equity.xlsx> <hybrid.xlsx> [<debt.xlsx>] "
            "[--analytics-date YYYY-MM-DD] [--cycle-json data/screener-YYYY-MM-DD.json]",
            file=sys.stderr,
        )
        return 2
    debt = Path(positional[2]) if len(positional) > 2 and positional[2] else None
    convert(
        Path(positional[0]),
        Path(positional[1]),
        debt,
        analytics_date=analytics_date,
        cycle_json_path=Path(cycle_json) if cycle_json else None,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
