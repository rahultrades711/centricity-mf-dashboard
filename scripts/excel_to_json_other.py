"""
excel_to_json_other.py — Stage B E1 (2026-05-29)

Converts the Other-Funds whitelisting workbook (Commodity + FoF + Solution +
Other Misc, ~290 funds) into `data/other-<cycle>.json`, keyed by AMFI scheme
code. These funds are **Not Scored** this cycle (no Centricity score/rank, like
index funds) — the dashboard merges them into the Equity / Hybrid / Debt /
Commodity asset-class universes at load time.

Classification (E1 locked rules, Rahul 2026-05-29):
  • asset_class:
      - the whole `Commodity` sheet            -> "Commodity" (new asset class)
      - else from the workbook's `Underlying Focus` column:
          Equity / Solution (Equity)           -> "Equity"
          Hybrid / Solution (Hybrid)            -> "Hybrid"
          Debt   / Solution (Debt)              -> "Debt"
          (the 75%/75% underlying split is already applied in that column)
      - "Mixed/Unspecified" (no 75%+ majority)  -> "Hybrid" (decision #3
          in-between default; ALL flagged so Rahul can reclassify the few that
          are clearly equity (overseas/sectoral ETF FoFs) or debt (G-Sec FoFs))
  • sub_category (NEW peer groups — never merged into existing eq/hybrid/debt):
      - FoF (Overseas) + Underlying Focus=Equity (or any global/overseas equity)
                                               -> Equity / "Global"
      - other FoF (Domestic/Overseas)          -> "<AssetClass> FoF"
      - Solution Retirement                    -> "Solution - Retirement"
      - Solution Children                      -> "Solution - Children"
      - Commodity (from Underlying Metal)      -> "Commodity - Gold/Silver/Multi"
      - Other Misc                             -> "Other"
      - underivable                            -> "Other" (flagged for review)

ISIN -> AMFI: the workbook carries ISIN, not AMFI. Resolve via the ACE master
`Fund Details without Turnover Ratios.xlsx` (SD_Scheme ISIN -> SD_Scheme AMFI
Code). Never join by name. Unresolved ISINs are reported (and the fund skipped).
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
WB_BASE = Path(
    r"C:\Claude Folder\Cent-Claude\MF Screener"
    r"\Monthly Equity Whitelisting File (MAIN FILE)\15-May-2026"
)
OTHER_XLSX = WB_BASE / "Output" / "OtherFunds_Whitelisting_15May2026.xlsx"
ACE_XLSX = WB_BASE / "Data" / "Fund Details Data" / "Fund Details without Turnover Ratios.xlsx"
CYCLE_DATE = "2026-05-15"

DATA_SHEETS = ["Commodity", "FoF (Domestic)", "FoF (Overseas)",
               "Solution Retirement", "Solution Children", "Other Misc"]


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("%", "")
    if s in ("", "-", "—", "NA", "N/A", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def build_isin_to_amfi() -> dict[str, int]:
    wb = openpyxl.load_workbook(ACE_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    # header at row index 3 (Scheme Code / ... / SD_Scheme AMFI Code / SD_Scheme ISIN)
    hdr_idx, header = None, None
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if any("ISIN" in c.upper() for c in cells) and any("AMFI CODE" in c.upper() for c in cells):
            hdr_idx, header = i, cells
            break
    if header is None:
        raise SystemExit("[other] could not find ISIN/AMFI header in ACE master")
    amfi_col = next(j for j, h in enumerate(header) if "AMFI CODE" in h.upper())
    isin_col = next(j for j, h in enumerate(header) if h.upper().endswith("ISIN") or "SCHEME ISIN" in h.upper())
    mapping: dict[str, int] = {}
    for r in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
        isin = _str(r[isin_col]) if isin_col < len(r) else None
        amfi = r[amfi_col] if amfi_col < len(r) else None
        if not isin or amfi is None:
            continue
        try:
            mapping[isin.upper()] = int(amfi)
        except (TypeError, ValueError):
            continue
    return mapping


def _asset_class(sheet: str, focus: str | None) -> tuple[str, bool]:
    """Return (asset_class, flagged). flagged=True if we had to guess."""
    if sheet == "Commodity":
        return "Commodity", False
    f = (focus or "").lower()
    if "equity" in f:
        return "Equity", False
    if "hybrid" in f:
        return "Hybrid", False
    if "debt" in f:
        return "Debt", False
    # "Mixed/Unspecified" — the workbook could not assign a 75%+ majority.
    # Per decision #3 the in-between default is Hybrid; flagged so Rahul can
    # reclassify the few that are clearly equity (overseas/sectoral ETF FoFs)
    # or debt (G-Sec / gilt FoFs). We do NOT name-guess here.
    return "Hybrid", True


def _sub_category(sheet: str, focus: str | None, sebi: str | None,
                  metal: str | None) -> tuple[str, bool]:
    f = (focus or "").lower()
    if sheet == "Commodity":
        m = (metal or focus or "").lower()
        if "gold" in m:
            return "Commodity - Gold", False
        if "silver" in m:
            return "Commodity - Silver", False
        return "Commodity - Multi", False
    if sheet == "FoF (Overseas)":
        return ("Global", False) if "equity" in f else (
            "Hybrid FoF" if "hybrid" in f else "Debt FoF" if "debt" in f else "FoF", "equity" not in f and "hybrid" not in f and "debt" not in f)
    if sheet == "FoF (Domestic)":
        if "equity" in f:
            return "Equity FoF", False
        if "hybrid" in f:
            return "Hybrid FoF", False
        if "debt" in f:
            return "Debt FoF", False
        return "FoF", True
    if sheet == "Solution Retirement":
        return "Solution - Retirement", False
    if sheet == "Solution Children":
        return "Solution - Children", False
    if sheet == "Other Misc":
        return "Other", False
    return "Other", True


def convert() -> dict:
    isin_to_amfi = build_isin_to_amfi()
    wb = openpyxl.load_workbook(OTHER_XLSX, read_only=True, data_only=True)

    funds: list[dict] = []
    seen_amfi: set[int] = set()
    unresolved: list[dict] = []     # ISIN couldn't map to AMFI
    flagged: list[dict] = []        # asset/sub-category had to be guessed
    per_sheet: dict[str, int] = {}

    for sheet in DATA_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # header row = first containing "Fund Name"
        hdr_idx = next((i for i, r in enumerate(rows)
                        if any(str(c).strip().lower() == "fund name" for c in r if c is not None)), None)
        if hdr_idx is None:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
        col = {h.lower(): j for j, h in enumerate(header)}

        def g(r, name):
            j = col.get(name.lower())
            return r[j] if (j is not None and j < len(r)) else None

        count = 0
        for r in rows[hdr_idx + 1:]:
            fund_name = _str(g(r, "Fund Name"))
            # skip group-header rows ("━━ Gold (50 funds) ━━") + blanks
            if not fund_name or fund_name.startswith("━") or "─" in fund_name:
                continue
            isin = _str(g(r, "ISIN"))
            amfi = isin_to_amfi.get(isin.upper()) if isin else None
            focus = _str(g(r, "Underlying Focus"))
            sebi = _str(g(r, "SEBI Category"))
            metal = _str(g(r, "Underlying Metal"))
            asset_class, ac_flag = _asset_class(sheet, focus)
            sub_category, sc_flag = _sub_category(sheet, focus, sebi, metal)

            if amfi is None:
                unresolved.append({"sheet": sheet, "fund_name": fund_name, "isin": isin})
                continue
            if amfi in seen_amfi:
                continue   # de-dupe across sheets
            seen_amfi.add(amfi)
            if ac_flag or sc_flag:
                flagged.append({"amfi": amfi, "fund_name": fund_name, "sheet": sheet,
                                "underlying_focus": focus, "asset_class": asset_class,
                                "sub_category": sub_category})

            rec = {
                "scheme_code": amfi,
                "fund_name": fund_name,
                "amc": _str(g(r, "AMC")),
                "sebi_category": sebi,
                "category": sub_category,            # NEW peer-group, dashboard reads `category`
                "sub_category_class": asset_class,   # asset class (Equity/Hybrid/Debt/Commodity)
                "manager_name": _str(g(r, "Manager")),
                "aum_cr": _num(g(r, "AUM (Cr)")),
                "ter_pct": _num(g(r, "TER (%)")),
                "nav_latest_value": _num(g(r, "Latest NAV")),
                "inception_date": _str(g(r, "Inception")),
                "fund_tenure_yrs": _num(g(r, "Fund Tenure (Yrs)")),
                "exit_load": _str(g(r, "Exit Load")),
                "isin": isin,
                "trailing_returns": {
                    "return_1y_pct": _num(g(r, "1Y Return %")),
                    "return_3y_pct": _num(g(r, "3Y Return %")),
                    "return_5y_pct": _num(g(r, "5Y Return %")),
                    "return_si_pct": _num(g(r, "SI Return %")),
                },
                "centricity_score": None,
                "centricity_rank_in_category": None,
                "centricity_score_status": "Not Scored",
                "analytics_pending": True,
            }
            if sheet == "Commodity" and metal:
                rec["underlying_metal"] = metal
            lockin = _str(g(r, "Lock-in"))
            if lockin:
                rec["lock_in"] = lockin
            funds.append(rec)
            count += 1
        per_sheet[sheet] = count

    from collections import Counter
    by_asset = Counter(f["sub_category_class"] for f in funds)
    by_sub = Counter(f["category"] for f in funds)

    out = {
        "contract_version": "other-v1",
        "cycle_meta": {
            "product_family": "MF_Other",
            "cycle_date": CYCLE_DATE,
            "cycle_label": "U2 May 2026",
            "label_date": "15th May 2026",
            "as_on_display": "15 May 2026",
            "total_funds": len(funds),
            "status": "Not Scored",
            "by_asset_class": dict(by_asset),
            "by_sub_category": dict(by_sub),
        },
        "funds": funds,
    }

    out_path = DATA_DIR / f"other-{CYCLE_DATE}.json"
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, separators=(",", ":"))

    print(f"[other] wrote {out_path.name}: {len(funds)} funds", file=sys.stderr)
    print(f"[other] per sheet: {per_sheet}", file=sys.stderr)
    print(f"[other] by asset_class: {dict(by_asset)}", file=sys.stderr)
    print(f"[other] by sub_category: {dict(by_sub)}", file=sys.stderr)
    print(f"[other] unresolved ISIN -> AMFI: {len(unresolved)}", file=sys.stderr)
    for u in unresolved:
        print(f"         UNRESOLVED {u['sheet']} | {u['fund_name']} | ISIN={u['isin']}", file=sys.stderr)
    print(f"[other] flagged (guessed asset/sub-category): {len(flagged)}", file=sys.stderr)
    for x in flagged:
        print(f"         FLAG {x['fund_name']} ({x['sheet']}, focus={x['underlying_focus']}) -> {x['asset_class']}/{x['sub_category']}", file=sys.stderr)
    return out


if __name__ == "__main__":
    convert()
